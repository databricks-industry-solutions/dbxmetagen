"""FastAPI backend for dbxmetagen dashboard app."""

import contextvars
import io
import os
import re
import json
import time
import uuid as _uuid
import queue
import logging
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed

import yaml
from collections import Counter
from typing import Optional, Union
from contextlib import asynccontextmanager

from cachetools import TTLCache, cached
from fastapi import Body, FastAPI, HTTPException, Query, Request, UploadFile, File, Form
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from databricks.sdk import WorkspaceClient
from db import pg_execute, get_engine, pg_configured
from dbxmetagen.ddl_bundle_utils import rewrite_ddl_catalog_schema as _rewrite_ddl_catalog_schema, dq_grade as _dq_grade

logger = logging.getLogger(__name__)


class _PollLogFilter(logging.Filter):
    """Suppress repetitive access log lines for polling endpoints."""
    _POLL_FRAGMENTS = ("/api/agent/deep/task/", "/api/genie/tasks/")

    def filter(self, record: logging.LogRecord) -> bool:
        msg = record.getMessage()
        return not any(frag in msg for frag in self._POLL_FRAGMENTS)


logging.getLogger("uvicorn.access").addFilter(_PollLogFilter())

# ---------------------------------------------------------------------------
# TTL caches -- shared across the process lifetime of the Databricks App
# ---------------------------------------------------------------------------
_yaml_cache = TTLCache(maxsize=32, ttl=300)
_yaml_lock = threading.Lock()
_job_list_cache = TTLCache(maxsize=4, ttl=30)
_job_list_lock = threading.Lock()
_coverage_cache = TTLCache(maxsize=16, ttl=60)
_coverage_lock = threading.Lock()
_sl_context_cache = TTLCache(maxsize=8, ttl=120)
_sl_context_lock = threading.Lock()


def invalidate_query_caches():
    """Clear query-result caches after mutations (job submit, DDL apply, etc.)."""
    with _coverage_lock:
        _coverage_cache.clear()
    with _sl_context_lock:
        _sl_context_cache.clear()

_LLM_MODEL = os.environ.get("LLM_MODEL", "databricks-claude-sonnet-4-6")
_AVAILABLE_MODELS = [
    "databricks-claude-sonnet-4-6",
    "databricks-gpt-oss-120b",
]

# Background Genie builder tasks: task_id -> {status, stage, result, error, created}
_genie_tasks: dict[str, dict] = {}

# ---------------------------------------------------------------------------
# Databricks client
# ---------------------------------------------------------------------------

_ws: Optional[WorkspaceClient] = None
_OBO_ENABLED = os.environ.get("ENABLE_OBO", "false").lower() == "true"
_obo_token_var: contextvars.ContextVar[Optional[str]] = contextvars.ContextVar(
    "_obo_token_var", default=None
)


def get_workspace_client() -> WorkspaceClient:
    """Return the app service principal WorkspaceClient singleton."""
    global _ws
    if _ws is None:
        _ws = WorkspaceClient()
    return _ws


def _get_effective_client() -> WorkspaceClient:
    """Return user-scoped WS client when OBO is active for this request, else app SP.

    Background daemon threads don't inherit the ContextVar, so they
    automatically fall back to the app SP singleton.
    """
    if _OBO_ENABLED:
        token = _obo_token_var.get(None)
        if token:
            from databricks.sdk.core import Config
            cfg = Config(host=get_workspace_client().config.host, token=token)
            return WorkspaceClient(config=cfg)
    return get_workspace_client()


_NOT_FOUND_RE = re.compile(
    r"TABLE_OR_VIEW_NOT_FOUND|SCHEMA_NOT_FOUND|CATALOG_NOT_FOUND"
    r"|does not exist|INVALID_SCHEMA_OR_RELATION_NAME"
    r"|relation .+ does not exist",
    re.IGNORECASE,
)


def execute_sql(query: str, warehouse_id: Optional[str] = None, timeout: int = 30):
    """Execute SQL via Statement Execution API and return rows as list[dict].

    Returns [] for missing-table/schema/catalog errors (expected before
    pipelines have run).  Raises HTTPException for other failures.
    Polls for completion when the initial wait_timeout is exceeded.
    """
    wh = warehouse_id or os.environ.get("WAREHOUSE_ID", "")
    if not wh:
        raise HTTPException(500, detail="WAREHOUSE_ID not configured")
    try:
        ws = _get_effective_client()
        wait_s = min(timeout, 50)
        resp = ws.statement_execution.execute_statement(
            statement=query, warehouse_id=wh, wait_timeout=f"{wait_s}s"
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("SDK error executing SQL: %s", exc)
        raise HTTPException(500, detail=str(exc))

    # Poll if still running (AI_QUERY can exceed the initial wait_timeout)
    deadline = time.time() + timeout
    while (
        resp.status
        and resp.status.state
        and resp.status.state.value in ("PENDING", "RUNNING")
    ):
        if time.time() > deadline:
            raise HTTPException(504, detail=f"Query timed out after {timeout}s")
        time.sleep(3)
        try:
            resp = ws.statement_execution.get_statement(resp.statement_id)
        except Exception as exc:
            logger.error("Error polling statement %s: %s", resp.statement_id, exc)
            raise HTTPException(500, detail=str(exc))

    if resp.status and resp.status.state and resp.status.state.value == "FAILED":
        msg = resp.status.error.message if resp.status.error else "SQL failed"
        if _NOT_FOUND_RE.search(msg):
            logger.warning("Table/schema not found: %s", msg)
            raise HTTPException(404, detail=msg)
        raise HTTPException(500, detail=msg)
    cols = [c.name for c in resp.manifest.schema.columns] if resp.manifest else []
    rows = []
    if resp.result and resp.result.data_array:
        for row in resp.result.data_array:
            rows.append(dict(zip(cols, row)))
    return rows


# ---------------------------------------------------------------------------
# App config
# ---------------------------------------------------------------------------

CATALOG = os.environ.get("CATALOG_NAME", "")
SCHEMA = os.environ.get("SCHEMA_NAME", "metadata_results")


def fq(table: str) -> str:
    return f"`{CATALOG}`.`{SCHEMA}`.`{table}`"


# Safe for use in LIKE/WHERE: alphanumeric, underscore, dot, hyphen, space, %
_SAFE_IDENT_RE = re.compile(r"^[a-zA-Z0-9_.\- %]*$")


def _ensure_column(table_fqn: str, col_name: str, col_type: str = "STRING"):
    """Add a column to a table if it doesn't already exist (schema evolution helper)."""
    try:
        cols = execute_sql(f"DESCRIBE TABLE {table_fqn}", timeout=15)
        if any(r.get("col_name") == col_name for r in cols):
            return
        execute_sql(f"ALTER TABLE {table_fqn} ADD COLUMN {col_name} {col_type}", timeout=15)
    except Exception as e:
        logger.debug("_ensure_column(%s, %s) skipped: %s", table_fqn, col_name, e)


def _safe_sql_str(s: Optional[str]) -> str:
    """Escape single quotes for SQL string literal."""
    if s is None:
        return "NULL"
    return "'" + str(s).replace("\\", "\\\\").replace("'", "''") + "'"


def _esc_sql(s) -> str:
    """Escape single quotes for use inside SQL string literal."""
    return str(s or "").replace("'", "''")


_labeled_table_ensured = False
_review_column_ensured: set[str] = set()


def _ensure_labeled_updates_table():
    global _labeled_table_ensured
    if _labeled_table_ensured:
        return
    execute_sql(
        f"""
        CREATE TABLE IF NOT EXISTS {fq('metadata_labeled_updates')} (
            update_id STRING NOT NULL,
            source_kb STRING NOT NULL,
            entity_identifier STRING NOT NULL,
            field_name STRING NOT NULL,
            old_value STRING,
            new_value STRING,
            updated_at TIMESTAMP,
            updated_by STRING
        ) COMMENT 'History of human corrections from metadata review app'
        """
    )
    _labeled_table_ensured = True


def _ensure_review_updated_at(table_key: str):
    global _review_column_ensured
    if table_key in _review_column_ensured:
        return
    try:
        execute_sql(f"ALTER TABLE {fq(table_key)} ADD COLUMN review_updated_at TIMESTAMP")
        _review_column_ensured.add(table_key)
    except Exception as e:
        if "already exists" in str(e).lower():
            _review_column_ensured.add(table_key)
        else:
            logger.warning("_ensure_review_updated_at(%s) failed: %s", table_key, e)


_pg_fallback_warned = False


def graph_query(sql: str) -> list[dict]:
    """Query graph tables: try Lakebase PG first, fall back to UC Delta tables."""
    global _pg_fallback_warned
    if pg_configured():
        try:
            return pg_execute(sql)
        except HTTPException:
            if not _pg_fallback_warned:
                logger.warning("Lakebase PG not connected, using UC Delta for graph queries")
                _pg_fallback_warned = True
        except Exception as e:
            if not _pg_fallback_warned:
                logger.warning("Lakebase PG not connected (%s), using UC Delta for graph queries", e)
                _pg_fallback_warned = True
    elif not _pg_fallback_warned:
        logger.info("Lakebase not configured (PGHOST not set), using UC Delta for graph queries")
        _pg_fallback_warned = True
    uc_sql = sql.replace("public.graph_nodes", fq("graph_nodes")).replace(
        "public.graph_edges", fq("graph_edges")
    )
    return execute_sql(uc_sql)


def multi_hop_traverse(
    start_node: str,
    max_hops: int = 3,
    relationship: str | None = None,
    edge_type: str | None = None,
    edge_types: list[str] | None = None,
    direction: str = "outgoing",
) -> dict:
    """Iterative BFS-style graph traversal with support for edge_type filtering.

    Accepts either a single ``edge_type`` or a list of ``edge_types`` (OR filter).
    If both are provided, ``edge_types`` takes precedence.
    """
    _validate_filter(relationship, "relationship")
    if edge_types:
        for et in edge_types:
            _validate_filter(et, "edge_type")
    else:
        _validate_filter(edge_type, "edge_type")

    visited_nodes: dict[str, dict] = {}
    edges_found: list[dict] = []
    frontier = {start_node}

    filters = []
    if relationship:
        filters.append(f"e.relationship = {_safe_sql_str(relationship)}")
    if edge_types:
        et_list = ", ".join(_safe_sql_str(et) for et in edge_types)
        filters.append(f"e.edge_type IN ({et_list})")
    elif edge_type:
        filters.append(f"e.edge_type = {_safe_sql_str(edge_type)}")
    filter_clause = (" AND " + " AND ".join(filters)) if filters else ""

    cols = (
        "e.src, e.dst, e.relationship, e.edge_type, e.weight, "
        "e.join_expression, e.join_confidence, e.ontology_rel, e.source_system"
    )

    for hop in range(max_hops):
        if not frontier:
            break
        id_list = ", ".join(_safe_sql_str(n) for n in frontier)
        if direction == "outgoing":
            q = f"SELECT {cols} FROM public.graph_edges e WHERE e.src IN ({id_list}) {filter_clause}"
        elif direction == "incoming":
            q = f"SELECT {cols} FROM public.graph_edges e WHERE e.dst IN ({id_list}) {filter_clause}"
        else:
            q = (
                f"SELECT {cols} FROM public.graph_edges e "
                f"WHERE (e.src IN ({id_list}) OR e.dst IN ({id_list})) {filter_clause}"
            )
        rows = graph_query(q)
        next_frontier = set()
        for r in rows:
            edges_found.append(r)
            for side in ("src", "dst"):
                nid = r.get(side)
                if nid and nid not in visited_nodes:
                    next_frontier.add(nid)
        frontier = next_frontier - set(visited_nodes.keys()) - {start_node}
        # Fetch node details for new frontier
        if frontier:
            nid_list = ", ".join(_safe_sql_str(n) for n in frontier)
            nq = (
                f"SELECT id, node_type, domain, display_name, short_description, "
                f"sensitivity, status FROM public.graph_nodes WHERE id IN ({nid_list})"
            )
            for nr in graph_query(nq):
                visited_nodes[nr["id"]] = nr

    # Also fetch start node details
    start_rows = graph_query(
        f"SELECT id, node_type, domain, display_name, short_description "
        f"FROM public.graph_nodes WHERE id = {_safe_sql_str(start_node)}"
    )
    if start_rows:
        visited_nodes[start_node] = start_rows[0]

    return {
        "start_node": start_node,
        "hops": max_hops,
        "nodes": visited_nodes,
        "edges": edges_found,
        "node_count": len(visited_nodes),
        "edge_count": len(edges_found),
    }


# ---------------------------------------------------------------------------
# FastAPI app
# ---------------------------------------------------------------------------


@asynccontextmanager
async def lifespan(app: FastAPI):
    logger.info("dbxmetagen API starting – catalog=%s schema=%s obo=%s", CATALOG, SCHEMA, _OBO_ENABLED)
    if _OBO_ENABLED:
        logger.info(
            "OBO mode active – ensure workspace preview "
            "'Databricks Apps - On-Behalf-Of User Authorization' is enabled"
        )
    if pg_configured():
        try:
            logger.info(
                "Lakebase PG connection configured -> %s:%s/%s",
                os.environ.get("PGHOST"),
                os.environ.get("PGPORT", "5432"),
                os.environ.get("PGDATABASE"),
            )
            get_engine()
            logger.info("Lakebase engine created OK")
        except Exception as e:
            logger.error("Lakebase engine creation failed (non-fatal): %s", e)
    else:
        logger.warning("PGHOST not set – add Lakebase database resource in Apps UI")

    route_count = len([r for r in app.routes if hasattr(r, "methods")])
    mount_count = len([r for r in app.routes if not hasattr(r, "methods")])
    logger.info("Routes registered: %d endpoints, %d mounts", route_count, mount_count)
    for r in app.routes:
        if hasattr(r, "methods"):
            logger.info("  %s %s", r.methods, r.path)
    yield


app = FastAPI(title="dbxmetagen API", version="0.8.9", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request
from starlette.responses import Response


class OBOMiddleware(BaseHTTPMiddleware):
    """Set the OBO user token ContextVar for the duration of each request."""
    async def dispatch(self, request: Request, call_next):
        if _OBO_ENABLED:
            _obo_token_var.set(request.headers.get("x-forwarded-access-token"))
        return await call_next(request)


class DebugRoutingMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        method = request.method
        if path.startswith("/api/"):
            logger.info(">> %s %s (matched routes: %d)", method, path,
                        len([r for r in app.routes if hasattr(r, "methods")]))
        response: Response = await call_next(request)
        if path.startswith("/api/") and response.status_code >= 400:
            logger.warning("<< %s %s -> %d", method, path, response.status_code)
        return response


app.add_middleware(DebugRoutingMiddleware)
app.add_middleware(OBOMiddleware)


@app.get("/api/health")
def health():
    """Simple health check that verifies routes are loaded."""
    route_count = len([r for r in app.routes if hasattr(r, "methods")])
    return {
        "status": "ok",
        "routes": route_count,
        "route_list": [
            {"path": r.path, "methods": list(r.methods)}
            for r in app.routes if hasattr(r, "methods")
        ][:20],
    }


def _resolve_user_identity() -> Optional[str]:
    """Return the OBO user's email when available, else None."""
    if not _OBO_ENABLED:
        return None
    try:
        return _get_effective_client().current_user.me().user_name
    except Exception:
        return None


@app.get("/api/config")
def get_config():
    """Return current catalog/schema defaults and processing settings for frontend."""
    host = ""
    try:
        host = get_workspace_client().config.host or ""
    except Exception:
        host = os.environ.get("DATABRICKS_HOST", "")
    return {
        "catalog_name": CATALOG,
        "schema_name": SCHEMA,
        "model": _LLM_MODEL,
        "sample_size": int(os.environ.get("SAMPLE_SIZE", "5")),
        "apply_ddl": os.environ.get("APPLY_DDL", "false").lower() == "true",
        "use_kb_comments": os.environ.get("USE_KB_COMMENTS", "false").lower() == "true",
        "include_lineage": os.environ.get("INCLUDE_LINEAGE", "false").lower() == "true",
        "workspace_host": host.rstrip("/"),
        "available_models": _AVAILABLE_MODELS,
        "lakebase_configured": pg_configured(),
        "obo_enabled": _OBO_ENABLED,
    }


@app.get("/api/auth/check")
def auth_check():
    """Verify the current caller's UC access to the configured catalog/schema."""
    if not _OBO_ENABLED:
        return {"obo_enabled": False, "message": "OBO is disabled; all operations use the app service principal"}
    identity = _resolve_user_identity()
    result = {"obo_enabled": True, "user_identity": identity, "has_catalog_access": False, "has_schema_access": False}
    try:
        execute_sql(f"USE CATALOG `{CATALOG}`")
        result["has_catalog_access"] = True
    except Exception:
        return result
    try:
        rows = execute_sql(f"SHOW SCHEMAS IN `{CATALOG}` LIKE '{SCHEMA}'")
        result["has_schema_access"] = len(rows) > 0
    except Exception:
        pass
    return result


# ---------------------------------------------------------------------------
# Request / response models
# ---------------------------------------------------------------------------


class JobRunRequest(BaseModel):
    job_id: Optional[int] = None
    job_name: Optional[str] = None
    # Metadata job params
    table_names: Optional[str] = None
    mode: Optional[str] = None
    apply_ddl: bool = False
    use_kb_comments: bool = False
    include_lineage: bool = False
    # Analytics pipeline params
    catalog_name: Optional[str] = None
    schema_name: Optional[str] = None
    ontology_bundle: Optional[str] = None
    domain_config: Optional[str] = None
    extra_params: dict = {}


class GraphQueryRequest(BaseModel):
    question: str


class GraphTraverseRequest(BaseModel):
    start_node: str
    max_hops: int = 3
    relationship: Optional[str] = None
    direction: str = "outgoing"  # outgoing | incoming | both


class SemanticLayerQuestionsRequest(BaseModel):
    questions: list[str]


class SemanticProfileRequest(BaseModel):
    profile_name: str
    questions: list[str]
    table_patterns: list[str] = []
    business_context: Optional[str] = None


class SemanticGenerateRequest(BaseModel):
    tables: list[str]
    questions: list[str]
    catalog_name: Optional[str] = None
    schema_name: Optional[str] = None
    model_endpoint: str = _LLM_MODEL
    project_id: Optional[str] = None
    profile_id: Optional[str] = None
    mode: str = (
        "replace"  # "replace" (supersede matching), "additive" (skip supersede), "replace_all" (supersede ALL in project)
    )
    business_context: Optional[str] = None


class SemanticProjectRequest(BaseModel):
    project_name: str
    description: str = ""


class MetricViewCreateRequest(BaseModel):
    target_catalog: str
    target_schema: str


class GenieGenerateRequest(BaseModel):
    table_identifiers: list[str]
    questions: list[str] = []
    metric_view_names: list[str] = []
    kpi_names: list[str] = []
    model_endpoint: str = _LLM_MODEL
    business_context: Optional[str] = None
    refinement_feedback: Optional[str] = None
    prior_result: Optional[dict] = None


class SuggestQuestionsRequest(BaseModel):
    table_identifiers: list[str]
    metric_view_names: list[str] = []
    model_endpoint: str = _LLM_MODEL
    count: int = 8
    purpose: str = "genie"  # "genie" or "metric_views"
    business_context: Optional[str] = None


class GenieCreateRequest(BaseModel):
    title: str
    description: Optional[str] = None
    serialized_space: dict
    warehouse_id: Optional[str] = None
    space_id: Optional[str] = None  # if provided, update instead of create


class GenieUpdateAssistRequest(BaseModel):
    section: str  # joins, instructions, questions, measures, filters, expressions, example_sql, synonyms
    table_identifiers: list[str] = []
    existing_items: Optional[list | dict] = None
    user_prompt: str = ""
    model_endpoint: str = _LLM_MODEL


class GenieEnrichDescriptionRequest(BaseModel):
    table_identifier: str
    existing_description: Optional[str] = None


# ---------------------------------------------------------------------------
# Jobs endpoints
# ---------------------------------------------------------------------------


def _get_job_with_retry(ws, job_id: int, retries: int = 3):
    """Call ws.jobs.get with retry on transient errors (503, timeouts)."""
    import time

    transient_markers = ("temporarily unavailable", "503", "timed out", "connection")
    for attempt in range(retries + 1):
        try:
            return ws.jobs.get(job_id)
        except Exception as e:
            msg = str(e).lower()
            if attempt < retries and any(m in msg for m in transient_markers):
                wait = 0.5 * (2 ** attempt)
                logger.debug("Transient error fetching job %d, retry %d in %.1fs: %s", job_id, attempt + 1, wait, e)
                time.sleep(wait)
            else:
                raise


def _list_dbxmetagen_jobs(ws):
    """Return project jobs using known IDs (env var), falling back to list(). Cached 30s."""
    with _job_list_lock:
        if "jobs" in _job_list_cache:
            return _job_list_cache["jobs"]
    if _KNOWN_JOB_IDS:
        jobs = []
        for name, job_id in _KNOWN_JOB_IDS.items():
            try:
                j = _get_job_with_retry(ws, job_id)
                jobs.append(j)
            except Exception as e:
                logger.warning("ws.jobs.get(%s=%d) failed: %s", name, job_id, e)
        logger.info(
            "Job discovery via valueFrom: %d/%d reachable",
            len(jobs),
            len(_KNOWN_JOB_IDS),
        )
        with _job_list_lock:
            _job_list_cache["jobs"] = jobs
        return jobs

    logger.info("No job IDs via valueFrom; falling back to ws.jobs.list()")
    try:
        all_jobs = list(ws.jobs.list())
    except Exception as e:
        logger.error("ws.jobs.list() failed: %s", e)
        raise HTTPException(
            503,
            detail=f"Failed to list jobs from Databricks API: {e}. "
            "Check app SPN permissions and workspace connectivity.",
        )
    matched = [
        j
        for j in all_jobs
        if j.settings
        and j.settings.name
        and any(kw in j.settings.name.lower() for kw in _JOB_NAME_KEYWORDS)
    ]
    logger.info(
        "Job discovery via list(): %d total, %d matched", len(all_jobs), len(matched)
    )
    with _job_list_lock:
        _job_list_cache["jobs"] = matched
    return matched


# ---------------------------------------------------------------------------
# Job configuration -- IDs injected via app.yaml valueFrom references
# ---------------------------------------------------------------------------

_JOB_ENV_MAP = {
    "metadata_generator": "METADATA_GENERATOR_JOB_ID",
    "metadata_parallel_modes": "METADATA_PARALLEL_MODES_JOB_ID",
    "sync_ddl": "SYNC_DDL_JOB_ID",
    "full_analytics_pipeline": "FULL_ANALYTICS_PIPELINE_JOB_ID",
    "fk_prediction": "FK_PREDICTION_JOB_ID",
    "sync_graph_lakebase": "SYNC_GRAPH_LAKEBASE_JOB_ID",
    "ontology_prediction": "ONTOLOGY_PREDICTION_JOB_ID",
    "knowledge_base_builder": "KNOWLEDGE_BASE_BUILDER_JOB_ID",
    "profiling": "PROFILING_JOB_ID",
    "metagen_with_kb": "METAGEN_WITH_KB_JOB_ID",
    "semantic_layer": "SEMANTIC_LAYER_JOB_ID",
    "metadata_kb_build": "METADATA_KB_BUILD_JOB_ID",
    "metadata_parallel_kb_build": "METADATA_PARALLEL_KB_BUILD_JOB_ID",
    "metadata_serverless": "METADATA_SERVERLESS_JOB_ID",
    "metadata_parallel_serverless": "METADATA_PARALLEL_SERVERLESS_JOB_ID",
    "kb_enriched_modes": "KB_ENRICHED_MODES_JOB_ID",
}

_KNOWN_JOB_IDS: dict[str, int] = {}
for _name, _env_var in _JOB_ENV_MAP.items():
    _val = os.environ.get(_env_var)
    if _val:
        try:
            _KNOWN_JOB_IDS[_name] = int(_val)
        except ValueError:
            logger.warning("Invalid %s value: %s", _env_var, _val)

if _KNOWN_JOB_IDS:
    logger.info(
        "Loaded %d job IDs via valueFrom: %s",
        len(_KNOWN_JOB_IDS),
        list(_KNOWN_JOB_IDS.keys()),
    )
else:
    logger.warning("No job IDs found in env vars; will fall back to ws.jobs.list()")


@app.get("/api/jobs")
def list_jobs():
    """List dbxmetagen jobs visible to the app."""
    ws = get_workspace_client()
    jobs = _list_dbxmetagen_jobs(ws)
    return [{"job_id": j.job_id, "name": j.settings.name} for j in jobs]


@app.post("/api/jobs/run")
def run_job(req: JobRunRequest):
    """Trigger a dbxmetagen job by job_id (preferred) or job_name suffix match."""
    logger.info("run_job request: job_id=%s, job_name=%s", req.job_id, req.job_name)
    ws = get_workspace_client()
    if req.job_id:
        target_job_id = req.job_id
    elif req.job_name:
        # Prefer direct lookup in known job IDs (exact or substring match)
        target_job_id = None
        if _KNOWN_JOB_IDS:
            if req.job_name in _KNOWN_JOB_IDS:
                target_job_id = _KNOWN_JOB_IDS[req.job_name]
            else:
                for name, jid in _KNOWN_JOB_IDS.items():
                    if req.job_name in name or name in req.job_name:
                        target_job_id = jid
                        break
            if target_job_id:
                logger.info(
                    "Resolved job_name '%s' via known IDs -> %d",
                    req.job_name,
                    target_job_id,
                )

        if not target_job_id:
            all_jobs = _list_dbxmetagen_jobs(ws)
            matching = [
                j
                for j in all_jobs
                if j.settings
                and j.settings.name
                and j.settings.name.endswith(req.job_name)
            ]
            if not matching:
                matching = [
                    j
                    for j in all_jobs
                    if j.settings
                    and j.settings.name
                    and req.job_name in j.settings.name
                ]
            if matching:
                target_job_id = matching[0].job_id
            else:
                available = [j.settings.name for j in all_jobs if j.settings]
                raise HTTPException(
                    404,
                    detail=f"Job '{req.job_name}' not found. "
                    f"Run 'databricks bundle deploy' to create jobs, then restart the app. "
                    f"Available jobs ({len(available)}): {available}",
                )
    else:
        raise HTTPException(400, detail="Provide job_id or job_name")
    params = {}
    if req.table_names:
        params["table_names"] = req.table_names
    if req.mode:
        params["mode"] = req.mode
    if req.apply_ddl:
        params["apply_ddl"] = "true"
    if req.use_kb_comments:
        params["use_kb_comments"] = "true"
    if req.include_lineage:
        params["include_lineage"] = "true"
    if req.catalog_name:
        params["catalog_name"] = req.catalog_name
    if req.schema_name:
        params["schema_name"] = req.schema_name
    if req.ontology_bundle:
        params["ontology_bundle"] = req.ontology_bundle
    if req.domain_config:
        params["domain_config_path"] = _resolve_domain_config_path(req.domain_config)
    params.update(req.extra_params)
    try:
        run = ws.jobs.run_now(job_id=target_job_id, job_parameters=params)
    except Exception as e:
        logger.error("jobs.run_now(job_id=%s) failed: %s", target_job_id, e)
        raise HTTPException(
            500,
            detail=f"Failed to trigger job {target_job_id}: {e}. "
            "The SPN may lack CAN_MANAGE_RUN permission on this job.",
        )
    with _job_list_lock:
        _job_list_cache.clear()
    invalidate_query_caches()
    return {"run_id": run.run_id}


@app.get("/api/jobs/{run_id}/status")
def get_run_status(run_id: int):
    """Get status of a job run with task-level detail."""
    ws = get_workspace_client()
    try:
        run = ws.jobs.get_run(run_id=run_id)
    except Exception as e:
        logger.error("get_run(run_id=%s) failed: %s", run_id, e)
        raise HTTPException(502, detail=f"Failed to fetch run {run_id}: {e}")

    tasks = []
    for t in run.tasks or []:
        ts = t.state if t else None
        tasks.append(
            {
                "task_key": t.task_key,
                "state": (
                    ts.life_cycle_state.value
                    if ts and ts.life_cycle_state
                    else "UNKNOWN"
                ),
                "result": ts.result_state.value if ts and ts.result_state else None,
            }
        )

    return {
        "run_id": run.run_id,
        "state": run.state.life_cycle_state.value if run.state else "UNKNOWN",
        "result": (
            run.state.result_state.value
            if run.state and run.state.result_state
            else None
        ),
        "state_message": (
            getattr(run.state, "state_message", None) if run.state else None
        ),
        "run_page_url": getattr(run, "run_page_url", None),
        "start_time": getattr(run, "start_time", None),
        "end_time": getattr(run, "end_time", None),
        "tasks": tasks,
    }


@app.get("/api/jobs/runs")
def list_recent_runs(limit: int = 50):
    """Return recent runs across all dbxmetagen jobs."""
    ws = get_workspace_client()
    try:
        dbx_jobs = _list_dbxmetagen_jobs(ws)
    except HTTPException:
        return []
    job_name_map = {j.job_id: j.settings.name for j in dbx_jobs if j.settings}
    runs = []
    for j in dbx_jobs:
        try:
            for r in ws.jobs.list_runs(job_id=j.job_id, limit=10):
                st = r.state if r else None
                runs.append(
                    {
                        "run_id": r.run_id,
                        "job_id": j.job_id,
                        "job_name": job_name_map.get(j.job_id, ""),
                        "state": (
                            st.life_cycle_state.value
                            if st and st.life_cycle_state
                            else "UNKNOWN"
                        ),
                        "result": (
                            st.result_state.value if st and st.result_state else None
                        ),
                        "state_message": (
                            getattr(st, "state_message", None) if st else None
                        ),
                        "start_time": getattr(r, "start_time", None),
                        "run_page_url": getattr(r, "run_page_url", None),
                    }
                )
        except Exception as e:
            logger.warning("list_runs(job_id=%s) failed: %s", j.job_id, e)
    runs.sort(key=lambda r: r.get("start_time") or 0, reverse=True)
    return runs[:limit]


_JOB_NAME_KEYWORDS = {
    "metadata",
    "metagen",
    "dbxmetagen",
    "profiling",
    "ontology",
    "semantic",
    "fk_prediction",
    "sync",
}


@app.get("/api/jobs/health")
def jobs_health_check():
    """Diagnostic preflight: check SPN connectivity and job visibility."""
    report = {
        "known_job_ids_configured": len(_KNOWN_JOB_IDS),
        "jobs_reachable": {},
        "project_jobs_found": 0,
        "project_job_names": [],
        "errors": [],
    }
    ws = get_workspace_client()

    if _KNOWN_JOB_IDS:
        reachable = {}
        for name, job_id in _KNOWN_JOB_IDS.items():
            try:
                j = _get_job_with_retry(ws, job_id)
                reachable[name] = {
                    "id": job_id,
                    "status": "ok",
                    "job_name": j.settings.name if j.settings else None,
                }
                report["project_job_names"].append(
                    j.settings.name if j.settings else f"id:{job_id}"
                )
            except Exception as e:
                reachable[name] = {"id": job_id, "status": "error", "error": str(e)}
                report["errors"].append(f"Job '{name}' (id={job_id}) unreachable: {e}")
        report["jobs_reachable"] = reachable
        report["project_jobs_found"] = sum(
            1 for v in reachable.values() if v["status"] == "ok"
        )
    else:
        report["errors"].append(
            "No job IDs found. Ensure app.yaml has valueFrom entries for each job resource "
            "and dbxmetagen_app.yml declares matching resources, "
            "then run 'databricks bundle deploy' and restart the app."
        )

    return report


# ---------------------------------------------------------------------------
# Metadata endpoints
# ---------------------------------------------------------------------------


def _validate_filter(val: Optional[str], param: str) -> None:
    if val is None or val == "":
        return
    if not _SAFE_IDENT_RE.match(val):
        raise HTTPException(400, f"Invalid {param}: only alphanumeric, underscore, dot, hyphen, space allowed")


@app.get("/api/metadata/log")
def get_metadata_log(limit: int = 100, table_name: Optional[str] = None):
    _validate_filter(table_name, "table_name")
    where = f"WHERE table_name LIKE {_safe_sql_str(f'%{table_name}%')}" if table_name else ""
    q = f"SELECT * FROM {fq('metadata_generation_log')} {where} ORDER BY _created_at DESC LIMIT {limit}"
    return execute_sql(q)


@app.get("/api/metadata/knowledge-base")
def get_knowledge_base(table_name: Optional[str] = None, schema_name: Optional[str] = None, limit: int = 100):
    _validate_filter(table_name, "table_name")
    _validate_filter(schema_name, "schema_name")
    clauses = []
    if table_name:
        clauses.append(f"table_name LIKE {_safe_sql_str(f'%{table_name}%')}")
    if schema_name:
        clauses.append(f"`schema` = {_safe_sql_str(schema_name)}")
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    q = f"SELECT * FROM {fq('table_knowledge_base')} {where} ORDER BY table_name LIMIT {limit}"
    return execute_sql(q)


@app.get("/api/metadata/column-kb")
def get_column_kb(table_name: Optional[str] = None, column_name: Optional[str] = None, limit: int = 200):
    _validate_filter(table_name, "table_name")
    _validate_filter(column_name, "column_name")
    clauses = []
    if table_name:
        clauses.append(f"table_name LIKE {_safe_sql_str(f'%{table_name}%')}")
    if column_name:
        clauses.append(f"column_name LIKE {_safe_sql_str(f'%{column_name}%')}")
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    q = f"SELECT * FROM {fq('column_knowledge_base')} {where} ORDER BY table_name, column_name LIMIT {limit}"
    return execute_sql(q)


@app.get("/api/metadata/schema-kb")
def get_schema_kb(schema_name: Optional[str] = None):
    _validate_filter(schema_name, "schema_name")
    where = f"WHERE schema_name = {_safe_sql_str(schema_name)}" if schema_name else ""
    q = f"SELECT * FROM {fq('schema_knowledge_base')} {where} ORDER BY schema_name"
    return execute_sql(q)


@app.get("/api/metadata/geo-classifications")
def get_geo_classifications(table_name: Optional[str] = None, classification: Optional[str] = None, limit: int = 500):
    clauses = []
    if table_name:
        _validate_filter(table_name, "table_name")
        clauses.append(f"table_name = {_safe_sql_str(table_name)}")
    if classification:
        _validate_filter(classification, "classification")
        clauses.append(f"classification = {_safe_sql_str(classification)}")
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    q = f"SELECT * FROM {fq('geo_classifications')} {where} ORDER BY table_name, column_name LIMIT {limit}"
    return execute_sql(q)


# --- PATCH KB: request bodies ---
class TableKBRow(BaseModel):
    table_name: str
    comment: Optional[str] = None
    domain: Optional[str] = None
    subdomain: Optional[str] = None
    has_pii: Optional[bool] = None
    has_phi: Optional[bool] = None


class ColumnKBRow(BaseModel):
    column_id: Optional[str] = None
    table_name: Optional[str] = None
    column_name: Optional[str] = None
    comment: Optional[str] = None
    classification: Optional[str] = None
    classification_type: Optional[str] = None


class SchemaKBRow(BaseModel):
    schema_id: str
    comment: Optional[str] = None
    domain: Optional[str] = None


def _record_labeled_update(
    source_kb: str,
    entity_id: str,
    field_name: str,
    old_value: Optional[str],
    new_value: Optional[str],
    updated_by: Optional[str] = None,
):
    _ensure_labeled_updates_table()
    uid = str(_uuid.uuid4())
    execute_sql(
        f"""
        INSERT INTO {fq('metadata_labeled_updates')}
        (update_id, source_kb, entity_identifier, field_name, old_value, new_value, updated_at, updated_by)
        VALUES ({_safe_sql_str(uid)}, {_safe_sql_str(source_kb)}, {_safe_sql_str(entity_id)}, {_safe_sql_str(field_name)},
                {_safe_sql_str(old_value)}, {_safe_sql_str(new_value)}, current_timestamp(), {_safe_sql_str(updated_by)})
        """
    )


@app.patch("/api/metadata/knowledge-base")
def patch_knowledge_base(body: list[TableKBRow]):
    _ensure_labeled_updates_table()
    _ensure_review_updated_at("table_knowledge_base")
    tbl = fq("table_knowledge_base")
    for row in body:
        if not row.table_name or not _SAFE_IDENT_RE.match(row.table_name):
            raise HTTPException(400, "Invalid table_name")
        current = execute_sql(
            f"SELECT table_name, comment, domain, subdomain FROM {tbl} WHERE table_name = {_safe_sql_str(row.table_name)} LIMIT 1"
        )
        old = current[0] if current else {}
        updates = []
        if row.comment is not None:
            updates.append(f"comment = {_safe_sql_str(row.comment)}")
            if old.get("comment") != row.comment:
                _record_labeled_update("table_kb", row.table_name, "comment", old.get("comment"), row.comment)
        if row.domain is not None:
            updates.append(f"domain = {_safe_sql_str(row.domain)}")
            if old.get("domain") != row.domain:
                _record_labeled_update("table_kb", row.table_name, "domain", old.get("domain"), row.domain)
        if row.subdomain is not None:
            updates.append(f"subdomain = {_safe_sql_str(row.subdomain)}")
            if old.get("subdomain") != row.subdomain:
                _record_labeled_update("table_kb", row.table_name, "subdomain", old.get("subdomain"), row.subdomain)
        if row.has_pii is not None:
            updates.append(f"has_pii = {str(row.has_pii).lower()}")
        if row.has_phi is not None:
            updates.append(f"has_phi = {str(row.has_phi).lower()}")
        if not updates:
            continue
        updates.append("updated_at = current_timestamp()")
        updates.append("review_updated_at = current_timestamp()")
        execute_sql(
            f"UPDATE {tbl} SET {', '.join(updates)} WHERE table_name = {_safe_sql_str(row.table_name)}"
        )
    return {"updated": len(body)}


@app.patch("/api/metadata/column-kb")
def patch_column_kb(body: list[ColumnKBRow]):
    _ensure_labeled_updates_table()
    _ensure_review_updated_at("column_knowledge_base")
    tbl = fq("column_knowledge_base")
    for row in body:
        ident = row.column_id or (f"{row.table_name}.{row.column_name}" if row.table_name and row.column_name else None)
        if not ident or not _SAFE_IDENT_RE.match(ident.replace(".", "x")):
            raise HTTPException(400, "Provide column_id or (table_name, column_name)")
        where = f"column_id = {_safe_sql_str(row.column_id)}" if row.column_id else f"table_name = {_safe_sql_str(row.table_name)} AND column_name = {_safe_sql_str(row.column_name)}"
        current = execute_sql(f"SELECT column_id, comment, classification FROM {tbl} WHERE {where} LIMIT 1")
        old = current[0] if current else {}
        entity_id = old.get("column_id") or ident
        updates = []
        if row.comment is not None:
            updates.append(f"comment = {_safe_sql_str(row.comment)}")
            if old.get("comment") != row.comment:
                _record_labeled_update("column_kb", entity_id, "comment", old.get("comment"), row.comment)
        if row.classification is not None:
            updates.append(f"classification = {_safe_sql_str(row.classification)}")
            if old.get("classification") != row.classification:
                _record_labeled_update("column_kb", entity_id, "classification", old.get("classification"), row.classification)
        if row.classification_type is not None:
            updates.append(f"classification_type = {_safe_sql_str(row.classification_type)}")
        if not updates:
            continue
        updates.append("updated_at = current_timestamp()")
        updates.append("review_updated_at = current_timestamp()")
        execute_sql(f"UPDATE {tbl} SET {', '.join(updates)} WHERE {where}")
    return {"updated": len(body)}


@app.patch("/api/metadata/schema-kb")
def patch_schema_kb(body: list[SchemaKBRow]):
    _ensure_labeled_updates_table()
    _ensure_review_updated_at("schema_knowledge_base")
    tbl = fq("schema_knowledge_base")
    for row in body:
        if not row.schema_id or not _SAFE_IDENT_RE.match(row.schema_id):
            raise HTTPException(400, "Invalid schema_id")
        current = execute_sql(
            f"SELECT schema_id, comment, domain FROM {tbl} WHERE schema_id = {_safe_sql_str(row.schema_id)} LIMIT 1"
        )
        old = current[0] if current else {}
        updates = []
        if row.comment is not None:
            updates.append(f"comment = {_safe_sql_str(row.comment)}")
            if old.get("comment") != row.comment:
                _record_labeled_update("schema_kb", row.schema_id, "comment", old.get("comment"), row.comment)
        if row.domain is not None:
            updates.append(f"domain = {_safe_sql_str(row.domain)}")
            if old.get("domain") != row.domain:
                _record_labeled_update("schema_kb", row.schema_id, "domain", old.get("domain"), row.domain)
        if not updates:
            continue
        updates.append("updated_at = current_timestamp()")
        updates.append("review_updated_at = current_timestamp()")
        execute_sql(
            f"UPDATE {tbl} SET {', '.join(updates)} WHERE schema_id = {_safe_sql_str(row.schema_id)}"
        )
    return {"updated": len(body)}


# --- Generate / Apply DDL from KB ---
_DOMAIN_TAG = "domain"
_SUBDOMAIN_TAG = "subdomain"
_PI_CLASS_TAG = "data_classification"
_PI_SUBCLASS_TAG = "data_subclassification"


def _full_table_name(row: dict) -> str:
    t = (row.get("table_name") or "").strip()
    if not t:
        return ""
    parts = t.split(".")
    if len(parts) == 3:
        return ".".join(f"`{p}`" for p in parts)
    c = (row.get("catalog") or "").strip()
    s = (row.get("schema") or "").strip()
    return f"`{c}`.`{s}`.`{t}`" if c and s else t


def _escape_comment(t: Optional[str]) -> str:
    if t is None or t == "":
        return ""
    return str(t).replace('"', "'").replace("\\", "\\\\")


def _generate_table_ddl_rows(
    rows: list[dict], ddl_type: str = "all",
    domain_tag: str = _DOMAIN_TAG, subdomain_tag: str = _SUBDOMAIN_TAG,
) -> list[str]:
    stmts = []
    for r in rows:
        full = _full_table_name(r)
        if not full:
            continue
        if ddl_type in ("all", "comments"):
            comment = _escape_comment(r.get("comment"))
            if comment:
                stmts.append(f'COMMENT ON TABLE {full} IS "{comment}";')
        if ddl_type in ("all", "domain"):
            domain = (r.get("domain") or "").strip().replace("'", "''")
            subdomain = (r.get("subdomain") or "").strip().replace("'", "''")
            if domain:
                if subdomain:
                    stmts.append(f"ALTER TABLE {full} SET TAGS ('{domain_tag}' = '{domain}', '{subdomain_tag}' = '{subdomain}');")
                else:
                    stmts.append(f"ALTER TABLE {full} SET TAGS ('{domain_tag}' = '{domain}');")
    return stmts


def _generate_column_ddl_rows(
    rows: list[dict], ddl_type: str = "all",
    pi_class_tag: str = _PI_CLASS_TAG, pi_subclass_tag: str = _PI_SUBCLASS_TAG,
) -> list[str]:
    stmts = []
    for r in rows:
        full = _full_table_name(r)
        col = (r.get("column_name") or "").strip()
        if not full or not col:
            continue
        if ddl_type in ("all", "comments"):
            comment = _escape_comment(r.get("comment"))
            if comment:
                stmts.append(f'COMMENT ON COLUMN {full}.`{col}` IS "{comment}";')
        if ddl_type in ("all", "sensitivity"):
            classification = (r.get("classification") or "").strip().replace("'", "''")
            if classification and classification.lower() != "none":
                subclass = (r.get("classification_type") or "").strip().replace("'", "''")
                subclass = subclass if subclass and subclass.lower() != "none" else classification
                stmts.append(
                    f"ALTER TABLE {full} ALTER COLUMN `{col}` SET TAGS "
                    f"('{pi_class_tag}' = '{classification}', '{pi_subclass_tag}' = '{subclass}');"
                )
    return stmts


class GenerateDDLBody(BaseModel):
    scope: str  # "table" | "schema" | "column" | "geo"
    identifiers: Optional[list[str]] = None
    tag_key: Optional[str] = None  # legacy sensitivity tag override
    ddl_type: Optional[str] = None  # "comments" | "domain" | "sensitivity" | None (= all)
    domain_tag_key: Optional[str] = None
    subdomain_tag_key: Optional[str] = None
    sensitivity_tag_key: Optional[str] = None
    sensitivity_type_tag_key: Optional[str] = None


def _table_where(identifiers: list[str]) -> str:
    safe = [_safe_sql_str(x) for x in identifiers if _SAFE_IDENT_RE.match(x)]
    if not safe:
        raise HTTPException(400, "No valid table identifiers")
    return " OR ".join([f"table_name = {s}" for s in safe])


_TBL_LIMIT = 5000
_COL_LIMIT = 10000


@app.post("/api/metadata/generate-ddl")
def generate_ddl(body: GenerateDDLBody):
    scope = (body.scope or "table").lower()
    identifiers = body.identifiers or []
    ddl_type = (body.ddl_type or "all").lower()
    domain_tag = body.domain_tag_key or _DOMAIN_TAG
    subdomain_tag = body.subdomain_tag_key or _SUBDOMAIN_TAG
    pi_class = body.sensitivity_tag_key or body.tag_key or _PI_CLASS_TAG
    pi_subclass = body.sensitivity_type_tag_key or (_PI_SUBCLASS_TAG if pi_class == _PI_CLASS_TAG else pi_class)
    tbl_kb = fq("table_knowledge_base")
    col_kb = fq("column_knowledge_base")
    stmts: list[str] = []
    warnings: list[str] = []

    if scope in ("table", "schema"):
        if scope == "schema" and identifiers:
            safe = [_safe_sql_str(x) for x in identifiers if _SAFE_IDENT_RE.match(x)]
            where = " OR ".join([f"`schema` = {s}" for s in safe]) if safe else "1=0"
        elif identifiers:
            where = _table_where(identifiers)
        else:
            where = "1=1"
        if ddl_type in ("all", "comments", "domain"):
            tbl_rows = execute_sql(
                f"SELECT catalog, `schema`, table_name, comment, domain, subdomain FROM {tbl_kb} WHERE {where} LIMIT {_TBL_LIMIT}"
            )
            if len(tbl_rows) == _TBL_LIMIT:
                warnings.append(f"Table results truncated at {_TBL_LIMIT} rows")
            stmts += _generate_table_ddl_rows(tbl_rows, ddl_type, domain_tag, subdomain_tag)
        if ddl_type in ("all", "comments", "sensitivity"):
            col_where = where if scope == "schema" else (
                " OR ".join([f"table_name = {_safe_sql_str(x)}" for x in identifiers if _SAFE_IDENT_RE.match(x)])
                if identifiers else "1=1"
            )
            col_rows = execute_sql(
                f"SELECT catalog, `schema`, table_name, column_name, comment, classification, classification_type FROM {col_kb} WHERE {col_where} LIMIT {_COL_LIMIT}"
            )
            if len(col_rows) == _COL_LIMIT:
                warnings.append(f"Column results truncated at {_COL_LIMIT} rows")
            stmts += _generate_column_ddl_rows(col_rows, ddl_type, pi_class, pi_subclass)

    elif scope == "column":
        if identifiers:
            safe = [_safe_sql_str(x) for x in identifiers if _SAFE_IDENT_RE.match(x)]
            if not safe:
                raise HTTPException(400, "No valid identifiers")
            where = " OR ".join([f"table_name = {s} OR column_id = {s}" for s in safe])
        else:
            where = "1=1"
        col_rows = execute_sql(
            f"SELECT catalog, `schema`, table_name, column_name, comment, classification, classification_type FROM {col_kb} WHERE {where} LIMIT {_TBL_LIMIT}"
        )
        if len(col_rows) == _TBL_LIMIT:
            warnings.append(f"Column results truncated at {_TBL_LIMIT} rows")
        stmts = _generate_column_ddl_rows(col_rows, ddl_type, pi_class, pi_subclass)

    elif scope == "geo":
        stmts = _build_geo_tag_stmts(identifiers=identifiers or None, tag_key=body.tag_key or "geo_classification")
    else:
        raise HTTPException(400, "scope must be table, schema, column, or geo")

    diagnostic = None
    if not stmts and ddl_type == "sensitivity":
        try:
            diag_rows = execute_sql(
                f"SELECT COUNT(*) AS total, "
                f"COUNT(classification) AS with_class, "
                f"SUM(CASE WHEN classification IS NOT NULL AND LOWER(classification) != 'none' AND classification != '' THEN 1 ELSE 0 END) AS usable "
                f"FROM {col_kb} LIMIT 1"
            )
            if diag_rows:
                d = diag_rows[0]
                diagnostic = (
                    f"column_knowledge_base has {d.get('total', 0)} rows, "
                    f"{d.get('with_class', 0)} with classification set, "
                    f"{d.get('usable', 0)} usable (non-null, non-None). "
                    "If 0 usable, run the PI classification pipeline step first."
                )
        except Exception:
            pass

    sql = "\n".join(stmts) if stmts else (
        f"-- No DDL generated ({diagnostic})" if diagnostic else "-- No DDL generated"
    )

    vol_path = None
    if stmts:
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        current_date = datetime.now().strftime("%Y%m%d")
        volume_name = os.environ.get("VOLUME_NAME", "generated_metadata")
        ws = _get_effective_client()
        current_user = "app"
        try:
            current_user = ws.current_user.me().user_name.split("@")[0]
        except Exception:
            pass
        vol_path = f"/Volumes/{CATALOG}/{SCHEMA}/{volume_name}/{current_user}/{current_date}/generated_ddl_{ddl_type}_{ts}.sql"
        try:
            ws.files.upload(vol_path, io.BytesIO(sql.encode("utf-8")), overwrite=True)
            logger.info("DDL written to volume: %s", vol_path)
        except Exception as e:
            logger.warning("Failed to write DDL to volume: %s", e)
            vol_path = None

    return {"sql": sql, "statements": stmts, "volume_path": vol_path, "warnings": warnings if warnings else None}


_GOVERNED_TAG_HINT = (
    "This may be a governed tag requiring policy updates. "
    "You can use a custom tag key (e.g. 'discovered_classification') to write a discovery tag instead, "
    "or update the governed tag policy in Unity Catalog > Tags to allow this value."
)


def _execute_stmts_batched(
    stmts: list[str], batch: bool = True, timeout: int = 60,
) -> tuple[int, list[dict]]:
    """Execute DDL stmts with table-level column-ALTER batching. Returns (applied, errors).

    When *batch* is True, column-level ALTER TABLEs for the same table are merged
    into a single statement. Falls back to per-statement execution on batch failure.
    """
    applied = 0
    errors: list[dict] = []

    def _run_one(sql: str) -> None:
        nonlocal applied
        clean = sql.rstrip(";").strip()
        if not clean or clean.startswith("--"):
            return
        try:
            execute_sql(clean, timeout=timeout)
            applied += 1
        except Exception as e:
            err_str = str(e)
            detail: dict = {"statement": clean[:200], "error": err_str}
            if "PERMISSION_DENIED" in err_str and "tag" in err_str.lower():
                detail["governed_tag"] = True
                detail["hint"] = _GOVERNED_TAG_HINT
            errors.append(detail)

    if not batch or len(stmts) <= 1:
        for s in stmts:
            _run_one(s)
        return applied, errors

    groups: dict[str, list[str]] = {}
    table_level: list[str] = []
    for s in stmts:
        upper = s.strip().upper()
        if upper.startswith("ALTER TABLE") and "ALTER COLUMN" in upper:
            parts = s.strip().split(None, 3)
            tbl = parts[2] if len(parts) > 2 else None
            if tbl:
                groups.setdefault(tbl, []).append(s)
            else:
                table_level.append(s)
        else:
            table_level.append(s)

    for s in table_level:
        _run_one(s)

    for tbl, tbl_stmts in groups.items():
        try:
            alter_clauses = []
            for s in tbl_stmts:
                idx = s.upper().find("ALTER COLUMN")
                if idx > 0:
                    alter_clauses.append(s[idx:].rstrip(";").strip())
            if alter_clauses:
                batch_sql = f"ALTER TABLE {tbl} {', '.join(alter_clauses)};"
                execute_sql(batch_sql, timeout=120)
                applied += len(tbl_stmts)
        except Exception:
            for s in tbl_stmts:
                _run_one(s)

    return applied, errors


@app.post("/api/metadata/apply-ddl")
def apply_ddl(body: GenerateDDLBody, batch: bool = True):
    out = generate_ddl(body)
    stmts = out.get("statements") or []
    warnings = out.get("warnings") or []
    applied, errors = _execute_stmts_batched(stmts, batch=batch)
    result: dict = {"applied": applied}
    if errors:
        result["message"] = "Some DDL statements failed"
        result["errors"] = errors
    if warnings:
        result["warnings"] = warnings
    return result


# ---------------------------------------------------------------------------
# DDL Bundle builders (advanced metadata)
# ---------------------------------------------------------------------------


def _fetch_fk_rows(identifiers: Optional[list[str]] = None) -> list[dict]:
    """Fetch parsed FK prediction rows (shared by tag and constraint builders)."""
    fk_tbl = fq("fk_predictions")
    where = "src_table != dst_table AND final_confidence >= 0.5"
    if identifiers:
        safe = [_safe_sql_str(x) for x in identifiers if _SAFE_IDENT_RE.match(x)]
        if safe:
            tbl_cond = " OR ".join([f"src_table = {s}" for s in safe])
            where += f" AND ({tbl_cond})"
    try:
        rows = execute_sql(
            f"SELECT src_table, src_column, dst_table, dst_column FROM {fk_tbl} WHERE {where} ORDER BY final_confidence DESC LIMIT {_TBL_LIMIT}"
        )
    except Exception:
        return []
    parsed = []
    for r in rows:
        src_tbl = (r.get("src_table") or "").strip()
        src_col = (r.get("src_column") or "").strip().split(".")[-1]
        dst_tbl = (r.get("dst_table") or "").strip()
        dst_col = (r.get("dst_column") or "").strip().split(".")[-1]
        if src_tbl and src_col and dst_tbl and dst_col:
            parsed.append({"src_tbl": src_tbl, "src_col": src_col, "dst_tbl": dst_tbl, "dst_col": dst_col})
    return parsed


def _build_fk_tag_ddl(identifiers: Optional[list[str]] = None) -> list[str]:
    """Build FK-as-tags DDL from fk_predictions without executing."""
    return [
        f"ALTER TABLE {r['src_tbl']} ALTER COLUMN `{r['src_col']}` SET TAGS ('fk_references' = '{_esc_sql(r['dst_tbl'] + '.' + r['dst_col'])}');"
        for r in _fetch_fk_rows(identifiers)
    ]


def _build_fk_constraint_ddl(identifiers: Optional[list[str]] = None) -> list[str]:
    """Build FK constraint DDL from fk_predictions without executing."""
    import re as _re
    stmts = []
    for r in _fetch_fk_rows(identifiers):
        tbl_short = _re.sub(r'[^a-zA-Z0-9]', '_', r['src_tbl'].split('.')[-1])
        src_col_safe = _re.sub(r'[^a-zA-Z0-9]', '_', r['src_col'])
        dst_col_safe = _re.sub(r'[^a-zA-Z0-9]', '_', r['dst_col'])
        name = f"fk_{tbl_short}_{src_col_safe}_{dst_col_safe}"
        stmts.append(
            f"ALTER TABLE {r['src_tbl']} ADD CONSTRAINT IF NOT EXISTS {name} "
            f"FOREIGN KEY (`{r['src_col']}`) REFERENCES {r['dst_tbl']}(`{r['dst_col']}`);"
        )
    return stmts


def _build_data_quality_ddl(
    identifiers: Optional[list[str]] = None,
) -> list[str]:
    """Build data quality tag DDL from data_quality_scores."""
    dq_tbl = fq("data_quality_scores")
    where = "1=1"
    if identifiers:
        safe = [_safe_sql_str(x) for x in identifiers if _SAFE_IDENT_RE.match(x)]
        if safe:
            tbl_cond = " OR ".join([f"table_name = {s}" for s in safe])
            where = tbl_cond
    try:
        rows = execute_sql(
            f"SELECT table_name, overall_score FROM {dq_tbl} WHERE {where} LIMIT 500"
        )
    except Exception:
        return []
    stmts: list[str] = []
    for r in rows:
        tbl = (r.get("table_name") or "").strip()
        score = r.get("overall_score")
        if not tbl or score is None:
            continue
        if not _SAFE_IDENT_RE.match(tbl.replace(".", "x")):
            continue
        score_f = round(float(score), 1)
        grade = _dq_grade(score_f)
        stmts.append(
            f"ALTER TABLE {tbl} SET TAGS ('data_quality_score' = '{score_f}', 'data_quality_grade' = '{grade}');"
        )
    return stmts


def _build_metric_view_ddl(
    identifiers: Optional[list[str]] = None,
    target_catalog: Optional[str] = None,
    target_schema: Optional[str] = None,
) -> list[str]:
    """Build CREATE VIEW WITH METRICS DDL from metric_view_definitions."""
    try:
        _ensure_semantic_layer_tables()
    except Exception:
        return []
    try:
        rows = execute_sql(
            f"SELECT * FROM {fq('metric_view_definitions')} WHERE status = 'applied'"
        )
    except Exception:
        return []
    if not rows:
        return []
    default_cat = target_catalog or CATALOG
    default_sch = target_schema or SCHEMA
    stmts: list[str] = []
    for row in rows:
        defn = json.loads(row["json_definition"]) if isinstance(row["json_definition"], str) else row["json_definition"]
        mv_name = defn.get("name") or row.get("metric_view_name", "")
        if not mv_name:
            continue
        if identifiers:
            source = defn.get("source", "")
            if source and not any(i in source for i in identifiers):
                continue
        mv_cat = row.get("deployed_catalog") or default_cat
        mv_sch = row.get("deployed_schema") or default_sch
        fq_mv = f"`{mv_cat}`.`{mv_sch}`.`{mv_name}`"
        yaml_body = _definition_to_yaml(defn)
        stmts.append(f"CREATE OR REPLACE VIEW {fq_mv}\nWITH METRICS LANGUAGE YAML AS $$\n{yaml_body}$$;")
    return stmts


def _build_geo_tag_stmts(
    identifiers: Optional[list[str]] = None,
    tag_key: str = "geo_classification",
) -> list[str]:
    """Build geo classification tag DDL statements from geo_classifications table."""
    geo_tbl = fq("geo_classifications")
    _validate_filter(tag_key, "tag_key")
    if identifiers:
        safe = [_safe_sql_str(x) for x in identifiers if _SAFE_IDENT_RE.match(x)]
        where = " OR ".join([f"table_name = {s}" for s in safe]) if safe else "1=0"
    else:
        where = "1=1"
    try:
        rows = execute_sql(
            f"SELECT table_name, column_name, classification FROM {geo_tbl} WHERE ({where}) AND confidence >= 0.5 LIMIT {_COL_LIMIT}"
        )
    except Exception:
        return []
    stmts: list[str] = []
    for r in rows:
        tn = (r.get("table_name") or "").strip()
        cn = (r.get("column_name") or "").strip()
        cls = (r.get("classification") or "").strip()
        if tn and cn and cls:
            stmts.append(f"ALTER TABLE {tn} ALTER COLUMN `{cn}` SET TAGS ('{tag_key}' = '{cls}');")
    return stmts


# ---------------------------------------------------------------------------
# DDL Bundle endpoints
# ---------------------------------------------------------------------------


_BUNDLE_DDL_TYPES = [
    "comments", "domain", "sensitivity", "ontology",
    "fk",
]


class DDLBundleBody(BaseModel):
    types: list[str] = _BUNDLE_DDL_TYPES
    fk_mode: Optional[str] = "tags"
    identifiers: Optional[list[str]] = None
    target_catalog: Optional[str] = None
    target_schema: Optional[str] = None
    domain_tag_key: Optional[str] = None
    subdomain_tag_key: Optional[str] = None
    sensitivity_tag_key: Optional[str] = None
    geo_tag_key: Optional[str] = None


@app.post("/api/metadata/generate-ddl-bundle")
def generate_ddl_bundle(body: DDLBundleBody):
    """Generate a unified DDL bundle combining core + advanced metadata types.

    Each requested type produces a section of SQL statements. The combined script
    is written to a UC volume and returned as JSON with per-section detail.
    """
    requested = {t.lower() for t in body.types}
    ids = body.identifiers
    sections: dict[str, list[str]] = {}
    warnings: list[str] = []

    core_types = requested & {"comments", "domain", "sensitivity"}
    if core_types:
        for ct in sorted(core_types):
            core_body = GenerateDDLBody(
                scope="table",
                identifiers=ids,
                ddl_type=ct,
                domain_tag_key=body.domain_tag_key,
                subdomain_tag_key=body.subdomain_tag_key,
                sensitivity_tag_key=body.sensitivity_tag_key,
            )
            core_out = generate_ddl(core_body)
            core_stmts = core_out.get("statements") or []
            if core_stmts:
                sections[ct] = core_stmts
            if core_out.get("warnings"):
                warnings.extend(core_out["warnings"])

    if "ontology" in requested:
        sections["ontology"] = _build_ontology_tag_ddl(identifiers=ids)

    # Unified FK toggle: dispatch based on fk_mode
    if "fk" in requested:
        fk_mode = (body.fk_mode or "tags").lower()
        if fk_mode == "constraints":
            fk_stmts = _build_fk_constraint_ddl(identifiers=ids)
            if fk_stmts:
                sections["fk_constraints"] = fk_stmts
        else:
            fk_stmts = _build_fk_tag_ddl(identifiers=ids)
            if fk_stmts:
                sections["fk_tags"] = fk_stmts

    # Backward compat for old clients sending fk_tags/fk_constraints directly
    if "fk_tags" in requested and "fk" not in requested:
        sections["fk_tags"] = _build_fk_tag_ddl(identifiers=ids)
    if "fk_constraints" in requested and "fk" not in requested:
        sections["fk_constraints"] = _build_fk_constraint_ddl(identifiers=ids)

    if "geo" in requested:
        sections["geo"] = _build_geo_tag_stmts(identifiers=ids, tag_key=body.geo_tag_key or "geo_classification")

    if "metric_views" in requested:
        sections["metric_views"] = _build_metric_view_ddl(
            identifiers=ids,
            target_catalog=body.target_catalog,
            target_schema=body.target_schema,
        )

    if body.target_catalog or body.target_schema:
        target_cat = body.target_catalog or CATALOG
        target_sch = body.target_schema or SCHEMA
        for key in sections:
            if key != "metric_views":
                sections[key] = _rewrite_ddl_catalog_schema(sections[key], CATALOG, SCHEMA, target_cat, target_sch)

    parts = []
    total_count = 0
    for section_name, section_stmts in sections.items():
        if section_stmts:
            parts.append(f"-- =============================================================")
            parts.append(f"-- {section_name.upper().replace('_', ' ')} ({len(section_stmts)} statements)")
            parts.append(f"-- =============================================================\n")
            parts.extend(section_stmts)
            parts.append("")
            total_count += len(section_stmts)

    sql = "\n".join(parts) if parts else "-- No DDL generated"

    vol_path = None
    if total_count > 0:
        from datetime import datetime
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        current_date = datetime.now().strftime("%Y%m%d")
        volume_name = os.environ.get("VOLUME_NAME", "generated_metadata")
        ws = _get_effective_client()
        current_user = "app"
        try:
            current_user = ws.current_user.me().user_name.split("@")[0]
        except Exception:
            pass
        type_slug = "_".join(sorted(requested))[:60]
        vol_path = f"/Volumes/{CATALOG}/{SCHEMA}/{volume_name}/{current_user}/{current_date}/ddl_bundle_{type_slug}_{ts}.sql"
        try:
            ws.files.upload(vol_path, io.BytesIO(sql.encode("utf-8")), overwrite=True)
            logger.info("DDL bundle written to volume: %s", vol_path)
        except Exception as e:
            logger.warning("Failed to write DDL bundle to volume: %s", e)
            vol_path = None

    return {
        "sql": sql,
        "sections": {k: v for k, v in sections.items() if v},
        "counts": {k: len(v) for k, v in sections.items()},
        "total_statements": total_count,
        "volume_path": vol_path,
        "warnings": warnings if warnings else None,
    }


_bundle_apply_tasks: TTLCache = TTLCache(maxsize=64, ttl=3600)


def _run_bundle_apply(task_id: str, sections: dict[str, list[str]], volume_path: str = None):
    """Background worker for applying DDL bundle sections."""
    task = _bundle_apply_tasks[task_id]
    results: dict[str, dict] = {}
    total_applied = 0
    total_errors = 0
    try:
        for section_name, stmts in sections.items():
            task["current_section"] = section_name
            task["total_applied"] = total_applied
            sec_applied, sec_errors = _execute_stmts_batched(stmts, batch=True)
            results[section_name] = {"applied": sec_applied, "errors": len(sec_errors)}
            total_applied += sec_applied
            total_errors += len(sec_errors)
        task.update({
            "status": "done", "results": results, "total_applied": total_applied,
            "total_errors": total_errors, "volume_path": volume_path,
            "current_section": None,
        })
    except Exception as e:
        task.update({"status": "error", "error": str(e)})


@app.post("/api/metadata/apply-ddl-bundle")
def apply_ddl_bundle(body: DDLBundleBody):
    """Generate and apply a unified DDL bundle asynchronously."""
    out = generate_ddl_bundle(body)
    sections = out.get("sections") or {}
    if not sections:
        return {"task_id": None, "applied": 0, "errors": 0, "results": {}}
    task_id = str(_uuid.uuid4())[:12]
    _bundle_apply_tasks[task_id] = {
        "status": "running", "current_section": None,
        "total_applied": 0, "total_errors": 0,
    }
    threading.Thread(
        target=_run_bundle_apply, args=(task_id, sections, out.get("volume_path")),
        daemon=True,
    ).start()
    return {"task_id": task_id}


class ApplyBundleSqlBody(BaseModel):
    sections: dict[str, list[str]]


@app.post("/api/metadata/apply-ddl-bundle-sql")
def apply_ddl_bundle_sql(body: ApplyBundleSqlBody):
    """Apply pre-generated DDL sections asynchronously (skips regeneration)."""
    if not body.sections:
        return {"task_id": None, "applied": 0, "errors": 0}
    task_id = str(_uuid.uuid4())[:12]
    _bundle_apply_tasks[task_id] = {
        "status": "running", "current_section": None,
        "total_applied": 0, "total_errors": 0,
    }
    threading.Thread(
        target=_run_bundle_apply, args=(task_id, body.sections),
        daemon=True,
    ).start()
    return {"task_id": task_id}


@app.get("/api/metadata/apply-ddl-bundle/status/{task_id}")
def apply_ddl_bundle_status(task_id: str):
    """Poll apply-ddl-bundle progress."""
    task = _bundle_apply_tasks.get(task_id)
    if not task:
        raise HTTPException(404, "Task not found")
    return task


# ---------------------------------------------------------------------------
# Review Editor combined endpoint
# ---------------------------------------------------------------------------

class ReviewCombinedRequest(BaseModel):
    tables: Optional[list[str]] = None
    schemas: Optional[list[str]] = None


@app.post("/api/metadata/review-combined")
def review_combined(body: ReviewCombinedRequest):
    """Fetch combined table + column KB data, with ontology and FK info per table."""
    tbl_kb = fq("table_knowledge_base")
    col_kb = fq("column_knowledge_base")
    ent_tbl = fq("ontology_entities")
    fk_tbl = fq("fk_predictions")

    where_parts = []
    if body.tables:
        safe = [_safe_sql_str(t) for t in body.tables if _SAFE_IDENT_RE.match(t)]
        if safe:
            where_parts.append("(" + " OR ".join(f"table_name = {s}" for s in safe) + ")")
    if body.schemas:
        for s in body.schemas:
            parts = s.split(".")
            if len(parts) == 2 and all(_SAFE_IDENT_RE.match(p) for p in parts):
                where_parts.append(f"(catalog = '{parts[0]}' AND `schema` = '{parts[1]}')")
    if not where_parts:
        raise HTTPException(400, "Provide at least one table or schema")
    where = " OR ".join(where_parts)

    try:
        return _review_combined_impl(tbl_kb, col_kb, ent_tbl, fk_tbl, where)
    except HTTPException:
        raise
    except Exception as e:
        logger.error("review_combined failed: %s", e)
        raise HTTPException(500, detail=str(e))


def _review_combined_impl(tbl_kb, col_kb, ent_tbl, fk_tbl, where):
    _has_review_status = False
    try:
        cols = execute_sql(f"DESCRIBE TABLE {tbl_kb}", timeout=15)
        _has_review_status = any(r.get("col_name") == "review_status" for r in cols)
    except Exception as e:
        logger.debug("review_combined DESCRIBE TABLE: %s", e)
    if not _has_review_status:
        try:
            execute_sql(f"ALTER TABLE {tbl_kb} ADD COLUMN review_status STRING", timeout=15)
            _has_review_status = True
        except Exception as e:
            logger.debug("review_combined ADD COLUMN review_status: %s", e)
    rs_expr = "COALESCE(review_status, 'unreviewed') AS review_status" if _has_review_status else "'unreviewed' AS review_status"
    try:
        count_rows = execute_sql(f"SELECT COUNT(*) AS cnt FROM {tbl_kb} WHERE {where}", timeout=10)
        total_count = int(count_rows[0]["cnt"]) if count_rows else 0
    except Exception:
        total_count = None
    tbl_rows = execute_sql(f"""
        SELECT table_name, catalog, `schema`, table_short_name, comment,
               domain, subdomain, has_pii, has_phi,
               {rs_expr}
        FROM {tbl_kb} WHERE {where} LIMIT 200
    """)
    if not tbl_rows:
        return {"tables": [], "total_count": total_count or 0, "truncated": False}

    tbl_names = [r["table_name"] for r in tbl_rows]
    safe_names = [_safe_sql_str(n) for n in tbl_names]
    in_clause = ", ".join(safe_names)

    col_rows = execute_sql(f"""
        SELECT column_id, table_name, column_name, data_type, comment,
               classification, classification_type, confidence
        FROM {col_kb} WHERE table_name IN ({in_clause})
    """)

    onto_rows, fk_rows, col_prop_rows = [], [], []
    _onto_where = f"SIZE(source_tables) > 0 AND EXISTS(source_tables, t -> t IN ({in_clause}))"
    try:
        onto_rows = execute_sql(f"""
            SELECT entity_id, entity_type, entity_name, confidence,
                   source_columns, validation_notes, validated,
                   COALESCE(entity_role, 'primary') AS entity_role,
                   discovery_confidence, entity_uri, source_ontology,
                   EXPLODE(source_tables) as table_name
            FROM {ent_tbl}
            WHERE {_onto_where}
        """)
    except Exception as e:
        logger.warning("Enriched ontology query failed (%s), falling back to simple query", e)
        try:
            onto_rows = execute_sql(f"""
                SELECT entity_type, entity_name, confidence,
                       NULL AS entity_id, NULL AS source_columns,
                       NULL AS validation_notes, false AS validated,
                       'primary' AS entity_role, NULL AS discovery_confidence,
                       NULL AS entity_uri, NULL AS source_ontology,
                       EXPLODE(source_tables) as table_name
                FROM {ent_tbl}
                WHERE {_onto_where}
            """)
        except Exception as e:
            logger.debug("Ontology fallback query also failed: %s", e)

    # Fetch column properties
    cp_tbl = fq("ontology_column_properties")
    try:
        col_prop_rows = execute_sql(f"""
            SELECT property_id, table_name, column_name, property_name,
                   property_role, owning_entity_id, owning_entity_type,
                   linked_entity_type, confidence
            FROM {cp_tbl}
            WHERE table_name IN ({in_clause})
        """)
    except Exception as e:
        logger.debug("Column properties query failed: %s", e)
    try:
        fk_rows = execute_sql(f"""
            SELECT src_column, src_table, dst_column, dst_table, final_confidence,
                   ai_reasoning, ai_confidence, col_similarity, rule_score
            FROM {fk_tbl}
            WHERE (src_table IN ({in_clause}) OR dst_table IN ({in_clause}))
              AND src_table != dst_table
        """)
    except Exception as e:
        logger.warning("Enriched FK query failed (%s), falling back to simple query", e)
        try:
            fk_rows = execute_sql(f"""
                SELECT src_column, src_table, dst_column, dst_table, final_confidence
                FROM {fk_tbl}
                WHERE (src_table IN ({in_clause}) OR dst_table IN ({in_clause}))
                  AND src_table != dst_table
            """)
        except Exception as e:
            logger.debug("FK fallback query also failed: %s", e)

    cols_by_table = {}
    for c in col_rows:
        cols_by_table.setdefault(c["table_name"], []).append(c)
    onto_by_table = {}
    for o in onto_rows:
        raw_cols = o.get("source_columns")
        if isinstance(raw_cols, str):
            try:
                raw_cols = json.loads(raw_cols)
            except Exception:
                raw_cols = None
        onto_by_table.setdefault(o["table_name"], []).append({
            "entity_id": o.get("entity_id"),
            "entity_type": o["entity_type"],
            "entity_name": o["entity_name"],
            "confidence": o["confidence"],
            "entity_role": o.get("entity_role", "primary"),
            "discovery_confidence": o.get("discovery_confidence"),
            "source_columns": raw_cols if isinstance(raw_cols, list) else None,
            "validation_notes": o.get("validation_notes"),
            "validated": o.get("validated"),
            "entity_uri": o.get("entity_uri"),
            "source_ontology": o.get("source_ontology"),
        })
    for tbl_name, ents in onto_by_table.items():
        seen = {}
        for e in ents:
            key = (e["entity_type"], tuple(e["source_columns"] or []))
            if key not in seen or float(e["confidence"] or 0) > float(seen[key]["confidence"] or 0):
                seen[key] = e
        onto_by_table[tbl_name] = list(seen.values())

    # Build column properties lookup
    col_props_by_table: dict = {}
    for cp in col_prop_rows:
        col_props_by_table.setdefault(cp["table_name"], []).append(cp)

    fk_by_table = {}
    for f in fk_rows:
        for tn in set([f.get("src_table"), f.get("dst_table")]):
            if tn in tbl_names:
                fk_by_table.setdefault(tn, []).append(f)

    def _to_bool(v):
        if isinstance(v, bool):
            return v
        if v is None:
            return False
        return str(v).lower() in ("true", "1")

    result = []
    for t in tbl_rows:
        tn = t["table_name"]
        ents = onto_by_table.get(tn, [])
        primary_ents = [e for e in ents if e.get("entity_role") == "primary"]
        if primary_ents:
            primary_entity = primary_ents[0]
        elif ents:
            primary_entity = max(ents, key=lambda e: float(e.get("confidence") or 0))
        else:
            primary_entity = None
        result.append({
            **t,
            "has_pii": _to_bool(t.get("has_pii")),
            "has_phi": _to_bool(t.get("has_phi")),
            "review_status": t.get("review_status", "unreviewed"),
            "columns": cols_by_table.get(tn, []),
            "primary_entity": primary_entity,
            "ontology_entities": ents,
            "column_properties": col_props_by_table.get(tn, []),
            "fk_predictions": fk_by_table.get(tn, []),
        })
    truncated = total_count is not None and total_count > 200
    return {"tables": result, "total_count": total_count or len(result), "truncated": truncated}


class ExportVolumeRequest(BaseModel):
    tables: list[str]
    format: str = "tsv"
    include_columns: bool = True
    metadata_type: Optional[str] = None


@app.post("/api/metadata/export-volume")
def export_to_volume(body: ExportVolumeRequest):
    """Export metadata for selected tables to a volume as TSV or Excel."""
    import io, csv
    from datetime import datetime

    combined = review_combined(ReviewCombinedRequest(tables=body.tables))
    rows = []
    for t in combined.get("tables", []):
        rows.append({
            "level": "table", "table_name": t["table_name"], "column_name": "",
            "data_type": "", "comment": t.get("comment", ""),
            "domain": t.get("domain", ""), "subdomain": t.get("subdomain", ""),
            "has_pii": str(t.get("has_pii", "")), "has_phi": str(t.get("has_phi", "")),
            "classification": "", "classification_type": "",
        })
        if body.include_columns:
            for c in t.get("columns", []):
                rows.append({
                    "level": "column", "table_name": t["table_name"],
                    "column_name": c.get("column_name", ""), "data_type": c.get("data_type", ""),
                    "comment": c.get("comment", ""), "domain": "", "subdomain": "",
                    "has_pii": "", "has_phi": "",
                    "classification": c.get("classification", ""),
                    "classification_type": c.get("classification_type", ""),
                })
    if not rows:
        raise HTTPException(400, "No data to export")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    current_date = datetime.now().strftime("%Y%m%d")
    volume_name = os.environ.get("VOLUME_NAME", "generated_metadata")
    ws = _get_effective_client()
    current_user = "app"
    try:
        current_user = ws.current_user.me().user_name.split("@")[0]
    except Exception:
        pass

    if body.format == "excel":
        import openpyxl
        wb = openpyxl.Workbook()
        ws_sheet = wb.active
        ws_sheet.title = "Metadata"
        headers = list(rows[0].keys())
        ws_sheet.append(headers)
        for r in rows:
            ws_sheet.append([r.get(h, "") for h in headers])
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()
        ext = "xlsx"
    else:
        buf = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()), delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)
        content = buf.getvalue().encode("utf-8")
        ext = "tsv"

    type_suffix = f"_{body.metadata_type}" if body.metadata_type else ""
    vol_path = f"/Volumes/{CATALOG}/{SCHEMA}/{volume_name}/{current_user}/{current_date}/review_export{type_suffix}_{ts}.{ext}"
    try:
        ws.files.upload(vol_path, io.BytesIO(content) if isinstance(content, bytes) else io.BytesIO(content), overwrite=True)
    except Exception as e:
        raise HTTPException(500, detail=f"Failed to write to volume: {e}")
    return {"path": vol_path, "rows": len(rows), "format": ext}


# ---------------------------------------------------------------------------
# Import reviewed metadata
# ---------------------------------------------------------------------------


@app.get("/api/metadata/volume-files")
def list_volume_files():
    """List importable TSV/Excel files in the volume for the current user."""
    volume_name = os.environ.get("VOLUME_NAME", "generated_metadata")
    base = f"/Volumes/{CATALOG}/{SCHEMA}/{volume_name}"
    ws = _get_effective_client()
    results = []

    def _walk(path: str, depth: int = 0):
        if depth > 4:
            return
        try:
            entries = list(ws.files.list_directory_contents(path))
        except Exception:
            return
        for entry in entries:
            ep = entry.path if hasattr(entry, "path") else str(entry)
            name = ep.rsplit("/", 1)[-1] if "/" in ep else ep
            if entry.is_directory if hasattr(entry, "is_directory") else False:
                _walk(ep, depth + 1)
            elif name.endswith((".tsv", ".xlsx", ".xls")):
                results.append({
                    "path": ep,
                    "name": name,
                    "size": getattr(entry, "file_size", None),
                    "last_modified": str(getattr(entry, "last_modified", "")),
                })

    _walk(base)
    return results


def _parse_review_file(content: bytes, filename: str) -> list[dict]:
    """Parse a TSV or Excel review file into a list of row dicts."""
    import csv as _csv
    if filename.endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws_sheet = wb.active
        rows_iter = ws_sheet.iter_rows(values_only=True)
        headers = [str(h or "").strip() for h in next(rows_iter)]
        return [dict(zip(headers, [str(v) if v is not None else "" for v in row])) for row in rows_iter]
    else:
        text = content.decode("utf-8")
        reader = _csv.DictReader(io.StringIO(text), delimiter="\t")
        return [row for row in reader]


def _import_rows_to_kb(rows: list[dict]) -> dict:
    """Split parsed rows by level and upsert into table/column KB tables."""
    tbl_kb = fq("table_knowledge_base")
    col_kb = fq("column_knowledge_base")
    _ensure_labeled_updates_table()
    _ensure_review_updated_at("table_knowledge_base")
    _ensure_review_updated_at("column_knowledge_base")

    tables_updated = 0
    columns_updated = 0
    skipped = 0
    errors = []

    for row in rows:
        level = (row.get("level") or "").strip().lower()
        table_name = (row.get("table_name") or "").strip()
        if not table_name:
            skipped += 1
            continue

        try:
            if level == "table":
                updates = []
                for field, col in [("comment", "comment"), ("domain", "domain"), ("subdomain", "subdomain")]:
                    val = row.get(field)
                    if val is not None and val != "":
                        updates.append(f"{col} = {_safe_sql_str(val)}")
                for bool_field in ("has_pii", "has_phi"):
                    val = (row.get(bool_field) or "").strip().lower()
                    if val in ("true", "false"):
                        updates.append(f"{bool_field} = {val}")
                if updates:
                    updates.append("updated_at = current_timestamp()")
                    updates.append("review_updated_at = current_timestamp()")
                    execute_sql(f"UPDATE {tbl_kb} SET {', '.join(updates)} WHERE table_name = {_safe_sql_str(table_name)}")
                    tables_updated += 1
                else:
                    skipped += 1

            elif level == "column":
                col_name = (row.get("column_name") or "").strip()
                if not col_name:
                    skipped += 1
                    continue
                updates = []
                for field, col in [("comment", "comment"), ("classification", "classification"), ("classification_type", "classification_type")]:
                    val = row.get(field)
                    if val is not None and val != "":
                        updates.append(f"{col} = {_safe_sql_str(val)}")
                if updates:
                    updates.append("updated_at = current_timestamp()")
                    updates.append("review_updated_at = current_timestamp()")
                    where = f"table_name = {_safe_sql_str(table_name)} AND column_name = {_safe_sql_str(col_name)}"
                    execute_sql(f"UPDATE {col_kb} SET {', '.join(updates)} WHERE {where}")
                    columns_updated += 1
                else:
                    skipped += 1
            else:
                skipped += 1
        except Exception as e:
            errors.append(f"{table_name}: {e}")

    return {"tables_updated": tables_updated, "columns_updated": columns_updated, "skipped": skipped, "errors": errors}


class ImportReviewedRequest(BaseModel):
    volume_path: str


@app.post("/api/metadata/import-reviewed")
def import_reviewed_from_volume(body: ImportReviewedRequest):
    """Import a reviewed TSV/Excel from a volume path into KB tables."""
    ws = _get_effective_client()
    vp = body.volume_path.strip()
    if not vp:
        raise HTTPException(400, "volume_path is required")
    try:
        resp = ws.files.download(vp)
        content = resp.contents.read()
    except Exception as e:
        raise HTTPException(404, detail=f"Cannot read volume file: {e}")
    filename = vp.rsplit("/", 1)[-1]
    rows = _parse_review_file(content, filename)
    if not rows:
        raise HTTPException(400, "File is empty or has no parseable rows")
    result = _import_rows_to_kb(rows)
    result["source"] = vp
    result["total_rows"] = len(rows)
    return result


@app.post("/api/metadata/import-reviewed-upload")
async def import_reviewed_upload(file: UploadFile = File(...)):
    """Import a reviewed TSV/Excel via file upload into KB tables.

    Optionally saves the uploaded file to the volume before parsing.
    """
    content = await file.read()
    filename = file.filename or "upload.tsv"
    if not filename.endswith((".tsv", ".xlsx", ".xls")):
        raise HTTPException(400, "File must be .tsv, .xlsx, or .xls")

    volume_name = os.environ.get("VOLUME_NAME", "generated_metadata")
    ws = _get_effective_client()
    current_user = "app"
    try:
        current_user = ws.current_user.me().user_name.split("@")[0]
    except Exception:
        pass
    from datetime import datetime
    current_date = datetime.now().strftime("%Y%m%d")
    vol_path = f"/Volumes/{CATALOG}/{SCHEMA}/{volume_name}/{current_user}/reviewed_outputs/{current_date}/{filename}"
    try:
        ws.files.upload(vol_path, io.BytesIO(content), overwrite=True)
    except Exception as e:
        logger.warning("Could not save uploaded file to volume: %s", e)
        vol_path = None

    rows = _parse_review_file(content, filename)
    if not rows:
        raise HTTPException(400, "File is empty or has no parseable rows")
    result = _import_rows_to_kb(rows)
    result["total_rows"] = len(rows)
    if vol_path:
        result["saved_to"] = vol_path
    return result


# ---------------------------------------------------------------------------
# Ontology endpoints
# ---------------------------------------------------------------------------


@app.get("/api/ontology/discovery-diff")
def get_discovery_diff(
    catalog: Optional[str] = Query(None, description="Catalog name (default: env CATALOG_NAME)"),
    schema: Optional[str] = Query(None, description="Schema name (default: env SCHEMA_NAME)"),
):
    """Return the latest discovery diff report for the given catalog/schema."""
    cat = catalog or CATALOG
    sch = schema or SCHEMA
    if not cat or not sch:
        raise HTTPException(400, "catalog and schema required (or set CATALOG_NAME, SCHEMA_NAME)")
    if not _SAFE_IDENT_RE.match(cat) or not _SAFE_IDENT_RE.match(sch):
        raise HTTPException(400, "Invalid catalog or schema")
    tbl = f"`{cat}`.`{sch}`.`discovery_diff_report`"
    try:
        rows = execute_sql(
            f"SELECT diff_json, bundle_version, previous_version, timestamp "
            f"FROM {tbl} ORDER BY created_at DESC LIMIT 1",
            timeout=15,
        )
    except Exception as e:
        if _NOT_FOUND_RE.search(str(e)):
            raise HTTPException(404, f"discovery_diff_report not found: {e}")
        raise HTTPException(500, str(e))
    if not rows:
        return {
            "bundle_version": None,
            "previous_version": None,
            "timestamp": None,
            "entity_changes": {"added": [], "removed": [], "changed": []},
            "column_changes": {"role_changed": [], "new_columns": [], "removed_columns": []},
            "relationship_changes": {"added": [], "removed": []},
        }
    r = rows[0]
    diff_json = r.get("diff_json")
    if diff_json:
        try:
            return json.loads(diff_json)
        except Exception:
            pass
    return {
        "bundle_version": r.get("bundle_version"),
        "previous_version": r.get("previous_version"),
        "timestamp": r.get("timestamp"),
        "entity_changes": {"added": [], "removed": [], "changed": []},
        "column_changes": {"role_changed": [], "new_columns": [], "removed_columns": []},
        "relationship_changes": {"added": [], "removed": []},
    }


@app.get("/api/ontology/entities")
def get_ontology_entities(limit: int = 200):
    q = f"SELECT * FROM {fq('ontology_entities')} ORDER BY confidence DESC LIMIT {limit}"
    return execute_sql(q)


@app.get("/api/ontology/relationships")
def get_ontology_relationships(limit: int = 500):
    q = f"""
        SELECT relationship_id, src_entity_type, relationship_name,
               dst_entity_type, cardinality, evidence_column,
               evidence_table, source, confidence
        FROM {fq('ontology_relationships')}
        ORDER BY confidence DESC
        LIMIT {min(limit, 2000)}
    """
    try:
        return execute_sql(q)
    except Exception:
        return []


@app.get("/api/ontology/graph-edges")
def get_ontology_graph_edges(edge_type: str = "", limit: int = 500):
    """Return edges from the knowledge graph (graph_edges table), optionally filtered by type."""
    ge_tbl = fq("graph_edges")
    clauses = ["relationship NOT IN ('similar_embedding', 'shares_column_name')"]
    if edge_type and _SAFE_IDENT_RE.match(edge_type):
        clauses.append(f"edge_type = '{_esc_sql(edge_type)}'")
    where = " AND ".join(clauses)
    try:
        return execute_sql(f"""
            SELECT src, dst, relationship, edge_type, weight, ontology_rel
            FROM {ge_tbl}
            WHERE {where}
            ORDER BY weight DESC
            LIMIT {min(limit, 2000)}
        """)
    except Exception:
        return []


@app.get("/api/ontology/summary")
def get_ontology_summary():
    q = f"""
        SELECT entity_type, COUNT(*) as count,
               ROUND(AVG(confidence), 2) as avg_confidence,
               SUM(CASE WHEN validated THEN 1 ELSE 0 END) as validated
        FROM {fq('ontology_entities')}
        GROUP BY entity_type ORDER BY count DESC
    """
    return execute_sql(q)


@app.get("/api/ontology/turtle")
def get_ontology_turtle():
    """Return the last generated Turtle file for the current schema."""
    vol_path = f"/Volumes/{CATALOG}/{SCHEMA}/generated_metadata/ontology_output.ttl"
    ttl_candidates = [
        vol_path,
        os.path.join(os.path.dirname(__file__), "ontology_output.ttl"),
        os.path.join(os.path.dirname(__file__), "..", "ontology_output.ttl"),
        f"/tmp/dbxmetagen_ontology_{SCHEMA}.ttl",
    ]
    for path in ttl_candidates:
        if os.path.isfile(path):
            with open(path, "r") as f:
                content = f.read()
            return Response(content=content, media_type="text/turtle",
                            headers={"Content-Disposition": f"attachment; filename=ontology_{SCHEMA}.ttl"})
    try:
        rows = execute_sql(f"SELECT * FROM read_files('{vol_path}') LIMIT 1")
        if rows:
            content = rows[0].get("value", "")
            return Response(content=content, media_type="text/turtle",
                            headers={"Content-Disposition": f"attachment; filename=ontology_{SCHEMA}.ttl"})
    except Exception:
        pass
    return JSONResponse({"error": "No Turtle file found. Run ontology build with Turtle export enabled."}, status_code=404)


@app.get("/api/ontology/bundle-info")
def get_ontology_bundle_info(bundle: str = ""):
    """Return bundle metadata from the cached bundle list (no re-parsing)."""
    if not bundle:
        return JSONResponse({"error": "bundle parameter required"}, status_code=400)
    all_bundles = _list_bundles_local()
    match = next((b for b in all_bundles if b["key"] == bundle), None)
    if match:
        return match
    return {"bundle": bundle, "format_version": "unknown", "has_tier_indexes": False, "entity_count": 0}


class OntologyEntityReviewBody(BaseModel):
    entity_id: str
    entity_type: Optional[str] = None
    entity_uri: Optional[str] = None
    validated: Optional[bool] = None


_ENTITY_UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",
    re.IGNORECASE,
)


@app.get("/api/ontology/quality-summary")
def get_ontology_quality_summary():
    """Active bundle (env), tier hash from disk, and aggregate confidence from ontology_entities."""
    active = os.environ.get("ONTOLOGY_BUNDLE") or os.environ.get("ontology_bundle") or "general"
    bundles = _list_bundles_local()
    info = next((b for b in bundles if b["key"] == active), None)
    out: dict = {
        "active_bundle": active,
        "bundle_info": info,
        "from_entities": None,
    }
    ent_tbl = fq("ontology_entities")
    try:
        rows = execute_sql(
            f"""
            SELECT
              COUNT(1) AS entity_rows,
              ROUND(AVG(confidence), 4) AS avg_confidence,
              SUM(CASE WHEN COALESCE(confidence, 0) < 0.5 THEN 1 ELSE 0 END) AS below_05,
              SUM(CASE WHEN COALESCE(confidence, 0) < 0.6 THEN 1 ELSE 0 END) AS below_06
            FROM {ent_tbl}
            """,
            timeout=25,
        )
        out["from_entities"] = rows[0] if rows else {}
    except Exception as e:
        out["from_entities_error"] = str(e)
    return out


@app.get("/api/ontology/review-queue")
def get_ontology_review_queue(limit: int = 50, max_confidence: float = 0.6):
    """Low-confidence entity rows for human review."""
    ent_tbl = fq("ontology_entities")
    lim = min(max(1, limit), 500)
    mc = float(max_confidence)
    try:
        return execute_sql(
            f"""
            SELECT *
            FROM {ent_tbl}
            WHERE COALESCE(confidence, 0) <= {mc}
            ORDER BY confidence ASC NULLS FIRST
            LIMIT {lim}
            """,
            timeout=30,
        )
    except Exception as e:
        logger.warning("review-queue: %s", e)
        return []


@app.post("/api/ontology/entity-review")
def post_ontology_entity_review(body: OntologyEntityReviewBody):
    """Update entity_type / entity_uri / validated for a row in ontology_entities."""
    if not _ENTITY_UUID_RE.match((body.entity_id or "").strip()):
        raise HTTPException(400, "entity_id must be a UUID")
    sets = []
    if body.entity_type is not None:
        et = body.entity_type.strip()
        sets.append(f"entity_type = {_safe_sql_str(et)}")
        sets.append(f"entity_name = {_safe_sql_str(et)}")
    if body.entity_uri is not None:
        sets.append(f"entity_uri = {_safe_sql_str(body.entity_uri)}")
    if body.validated is not None:
        sets.append(f"validated = {str(bool(body.validated)).lower()}")
    if not sets:
        raise HTTPException(400, "Provide at least one of entity_type, entity_uri, validated")
    ent_tbl = fq("ontology_entities")
    eid = _esc_sql(body.entity_id.strip())
    sql = f"UPDATE {ent_tbl} SET {', '.join(sets)} WHERE entity_id = '{eid}'"
    try:
        execute_sql(sql, timeout=45)
        return {"ok": True, "entity_id": body.entity_id.strip()}
    except Exception as e:
        raise HTTPException(500, str(e)) from e


# ---------------------------------------------------------------------------
# Ontology graph store (lazy singleton for SPARQL endpoint)
# ---------------------------------------------------------------------------
_ontology_graph_store = None
_ontology_graph_lock = threading.Lock()


def _get_ontology_graph_store():
    """Lazy-load OntologyGraphStore from Turtle files."""
    global _ontology_graph_store
    if _ontology_graph_store is not None:
        return _ontology_graph_store
    with _ontology_graph_lock:
        if _ontology_graph_store is not None:
            return _ontology_graph_store
        try:
            from dbxmetagen.ontology_graph_store import OntologyGraphStore, is_available
            if not is_available():
                logger.warning("pyoxigraph not installed -- SPARQL endpoint disabled")
                return None
        except ImportError:
            logger.warning("ontology_graph_store not importable -- SPARQL endpoint disabled")
            return None

        store = OntologyGraphStore()
        vol_path = f"/Volumes/{CATALOG}/{SCHEMA}/generated_metadata/ontology_output.ttl"
        ttl_candidates = [
            vol_path,
            os.path.join(os.path.dirname(__file__), "ontology_output.ttl"),
            os.path.join(os.path.dirname(__file__), "..", "ontology_output.ttl"),
            f"/tmp/dbxmetagen_ontology_{SCHEMA}.ttl",
        ]
        loaded = False
        for path in ttl_candidates:
            if os.path.isfile(path):
                store.load_turtle(path)
                loaded = True
                break
        if not loaded:
            logger.info("No Turtle file found for SPARQL store -- store empty until build runs")
        _ontology_graph_store = store
        return store


class SparqlRequest(BaseModel):
    query: str


@app.post("/api/ontology/sparql")
def ontology_sparql(req: SparqlRequest):
    """Run a read-only SPARQL SELECT query against the ontology graph."""
    store = _get_ontology_graph_store()
    if store is None:
        return JSONResponse({"error": "SPARQL store not available (pyoxigraph not installed)"}, status_code=503)
    q = req.query.strip()
    if not q.upper().startswith(("SELECT", "ASK", "PREFIX")):
        return JSONResponse({"error": "Only SELECT and ASK queries are supported"}, status_code=400)
    try:
        results = store.sparql(q)
        return {"results": results, "count": len(results), "triple_count": store.triple_count}
    except Exception as e:
        return JSONResponse({"error": str(e)}, status_code=400)


@app.get("/api/ontology/source/{bundle_name}")
def get_ontology_source(bundle_name: str):
    """Serve the source OWL/TTL file for a given bundle."""
    bd = _find_bundle_dir()
    if not bd:
        return JSONResponse({"error": "Bundle dir not found"}, status_code=404)
    bundle_dir = _safe_bundle_path(bd, bundle_name)
    if not bundle_dir:
        return JSONResponse({"error": "Invalid bundle name"}, status_code=400)
    if os.path.isdir(bundle_dir):
        for fname in sorted(os.listdir(bundle_dir)):
            if fname.startswith("source") and fname.endswith((".ttl", ".owl")):
                fpath = os.path.join(bundle_dir, fname)
                with open(fpath, "r", encoding="utf-8") as f:
                    content = f.read()
                media = "text/turtle" if fname.endswith(".ttl") else "application/rdf+xml"
                return Response(content=content, media_type=media,
                                headers={"Content-Disposition": f"attachment; filename={fname}"})
    return JSONResponse({"error": f"No source file found for bundle '{bundle_name}'"}, status_code=404)


@app.get("/api/ontology/source-classes/{bundle_name}")
def get_ontology_source_classes(bundle_name: str, limit: int = 500):
    """Return the class hierarchy from a bundle's tier3 YAML for the class browser."""
    bd = _find_bundle_dir()
    if not bd:
        return JSONResponse({"error": "Bundle dir not found"}, status_code=404)
    safe = _safe_bundle_path(bd, bundle_name)
    if not safe:
        return JSONResponse({"error": "Invalid bundle name"}, status_code=400)
    tier3_path = os.path.join(safe, "entities_tier3.yaml")
    if not os.path.isfile(tier3_path):
        return JSONResponse({"error": f"No tier3 index for bundle '{bundle_name}'"}, status_code=404)
    with open(tier3_path, "r", encoding="utf-8") as f:
        tier3 = yaml.safe_load(f) or {}
    classes = []
    for name, data in list(tier3.items())[:limit]:
        classes.append({
            "name": name,
            "description": data.get("description", ""),
            "uri": data.get("uri", ""),
            "parents": data.get("parents", []),
            "source_ontology": data.get("source_ontology", ""),
            "keywords": data.get("keywords", [])[:5],
            "relationships": list(data.get("relationships", {}).keys())[:10],
            "typical_attributes": data.get("typical_attributes", [])[:10],
        })
    return {"bundle": bundle_name, "classes": classes, "total": len(tier3)}


@app.post("/api/ontology/import")
async def import_ontology(
    file: UploadFile = File(...),
    bundle_name: str = Form("imported"),
):
    """Import a custom OWL/TTL file and generate a bundle YAML + tier indexes."""
    import tempfile
    bd = _find_bundle_dir()
    if not bd:
        return JSONResponse({"error": "Bundle directory not found"}, status_code=500)
    safe_yaml = _safe_bundle_path(bd, f"{bundle_name}.yaml")
    safe_subdir = _safe_bundle_path(bd, bundle_name)
    if not safe_yaml or not safe_subdir:
        return JSONResponse({"error": "Invalid bundle name"}, status_code=400)

    content = await file.read()
    suffix = ".ttl" if file.filename and file.filename.endswith(".ttl") else ".owl"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    try:
        from dbxmetagen.ontology_import import owl_to_bundle_yaml
        output_yaml = safe_yaml
        bundle = owl_to_bundle_yaml(tmp_path, output_path=output_yaml, bundle_name=bundle_name)

        bundle_subdir = safe_subdir
        os.makedirs(bundle_subdir, exist_ok=True)
        source_dest = os.path.join(bundle_subdir, f"source{suffix}")
        with open(source_dest, "wb") as f:
            f.write(content)

        # Generate tier indexes
        try:
            from pathlib import Path

            from dbxmetagen.ontology_bundle_indexes import (
                build_tiers,
                entities_from_bundle,
                load_edge_catalog,
            )

            entities = entities_from_bundle(Path(output_yaml))
            edge_cat = load_edge_catalog(Path(output_yaml))
            counts = build_tiers(entities, Path(bundle_subdir), edge_catalog=edge_cat or None)
        except Exception as tier_err:
            logger.warning("Tier generation failed (bundle still created): %s", tier_err)
            counts = {}

        # Invalidate bundle list cache
        _yaml_cache.clear()

        entity_count = len(bundle.get("ontology", {}).get("entities", {}).get("definitions", {}))
        edge_count = len(bundle.get("ontology", {}).get("edge_catalog", {}))
        return {
            "bundle_name": bundle_name,
            "entity_count": entity_count,
            "edge_count": edge_count,
            "tier_counts": counts,
            "source_stored": True,
        }
    except ImportError:
        return JSONResponse({"error": "rdflib required for import"}, status_code=500)
    except Exception as e:
        logger.exception("Import failed")
        return JSONResponse({"error": str(e)}, status_code=500)
    finally:
        os.unlink(tmp_path)


_DOMAIN_CONFIG_DIR = "configurations"
_BUNDLE_SUBDIR = os.path.join(_DOMAIN_CONFIG_DIR, "ontology_bundles")
_CONFIG_DIR_CANDIDATES = [
    _DOMAIN_CONFIG_DIR,
    os.path.join("..", _DOMAIN_CONFIG_DIR),
    os.path.join(os.path.dirname(__file__), _DOMAIN_CONFIG_DIR),
    os.path.join(os.path.dirname(__file__), "..", "..", "..", _DOMAIN_CONFIG_DIR),
]


def _find_domain_config_dir() -> Optional[str]:
    for d in _CONFIG_DIR_CANDIDATES:
        resolved = os.path.abspath(d)
        exists = os.path.isdir(resolved)
        logger.debug("config-dir candidate: %s (resolved=%s, exists=%s)", d, resolved, exists)
        if exists:
            return d
    logger.warning("No config dir found. cwd=%s __file__=%s", os.getcwd(), __file__)
    return None


def _find_bundle_dir() -> Optional[str]:
    """Locate the ontology_bundles directory."""
    for base in _CONFIG_DIR_CANDIDATES:
        bd = os.path.join(base, "ontology_bundles")
        resolved = os.path.abspath(bd)
        exists = os.path.isdir(resolved)
        logger.debug("bundle-dir candidate: %s (resolved=%s, exists=%s)", bd, resolved, exists)
        if exists:
            return bd
    logger.warning("No bundle dir found. cwd=%s __file__=%s", os.getcwd(), __file__)
    return None


def _safe_bundle_path(bundle_dir: str, bundle_name: str) -> Optional[str]:
    """Return the resolved path inside bundle_dir, or None if it escapes the root."""
    joined = os.path.normpath(os.path.join(bundle_dir, bundle_name))
    if not joined.startswith(os.path.normpath(bundle_dir) + os.sep) and joined != os.path.normpath(bundle_dir):
        return None
    return joined


def _read_bundle_metadata_fast(filepath: str) -> dict | None:
    """Read only the metadata block from a bundle YAML without parsing the full file.

    Uses line-based parsing: reads lines from `metadata:` until the next
    top-level key (un-indented line), handling nested values, comments, and
    blank lines correctly.
    """
    try:
        lines = []
        in_meta = False
        with open(filepath, "r") as f:
            for line in f:
                stripped = line.rstrip("\n")
                if not in_meta:
                    if stripped.startswith("metadata:"):
                        in_meta = True
                        lines.append(stripped)
                    continue
                if stripped == "" or stripped.lstrip().startswith("#"):
                    lines.append(stripped)
                    continue
                if stripped[0] not in (" ", "\t"):
                    break
                lines.append(stripped)
        if not lines:
            return None
        return yaml.safe_load("\n".join(lines)).get("metadata", {})
    except Exception:
        return None


def _count_yaml_list(path: str) -> int:
    """Load a YAML file and return its length if it's a list, else 0."""
    try:
        with open(path, "r") as f:
            data = yaml.safe_load(f)
        return len(data) if isinstance(data, list) else 0
    except Exception:
        return 0


@cached(_yaml_cache, key=lambda: "bundles", lock=_yaml_lock)
def _list_bundles_local() -> list[dict]:
    """Read ontology bundle YAMLs directly (no dbxmetagen import needed). Cached 300s.

    Uses fast metadata-only parsing to avoid loading multi-MB bundle files.
    Falls back to full parse for small files or if fast parse fails.
    Counts entities/edges across all tier files (tier1 + tier2).
    """
    bd = _find_bundle_dir()
    if not bd:
        logger.warning("_list_bundles_local: no bundle dir found, returning empty list")
        return []
    bundles = []
    for fname in sorted(os.listdir(bd)):
        if not fname.endswith(".yaml"):
            continue
        try:
            filepath = os.path.join(bd, fname)
            bundle_key = fname.replace(".yaml", "")
            tier_dir = os.path.join(bd, bundle_key)
            has_tiers = os.path.isdir(tier_dir) and os.path.isfile(os.path.join(tier_dir, "entities_tier1.yaml"))

            file_size = os.path.getsize(filepath)
            meta = None
            entity_count = 0
            edge_count = 0
            domain_count = 0

            if file_size > 100_000:
                meta = _read_bundle_metadata_fast(filepath)

            if meta is None:
                with open(filepath, "r") as f:
                    raw = yaml.safe_load(f)
                meta = raw.get("metadata", {})
                entity_count = len(raw.get("ontology", {}).get("entities", {}).get("definitions", {}))
                domain_count = len(raw.get("domains", {}))

            if meta is not None and entity_count == 0:
                entity_count = meta.get("entity_count", 0)
                edge_count = meta.get("edge_count", 0)
                domain_count = meta.get("domain_count", domain_count)

            if has_tiers:
                tier_entity_total = 0
                tier_edge_total = 0
                for tier_name in ("entities_tier1.yaml", "entities_tier2.yaml"):
                    p = os.path.join(tier_dir, tier_name)
                    if os.path.isfile(p):
                        tier_entity_total += _count_yaml_list(p)
                for tier_name in ("edges_tier1.yaml", "edges_tier2.yaml", "edges_tier3.yaml"):
                    p = os.path.join(tier_dir, tier_name)
                    if os.path.isfile(p):
                        tier_edge_total += _count_yaml_list(p)
                if tier_entity_total > entity_count:
                    entity_count = tier_entity_total
                if tier_edge_total > edge_count:
                    edge_count = tier_edge_total

            bundle_info = {
                "key": bundle_key,
                "name": meta.get("name", bundle_key),
                "industry": meta.get("industry", "general"),
                "description": meta.get("description", ""),
                "standards_alignment": meta.get("standards_alignment", ""),
                "entity_count": entity_count,
                "edge_count": edge_count,
                "domain_count": domain_count,
                "bundle_type": meta.get("bundle_type", "ontology"),
                "tag_key": meta.get("tag_key", ""),
                "format_version": meta.get("format_version", "1.0"),
                "has_tier_indexes": has_tiers,
            }
            source_url = meta.get("source_url")
            if source_url:
                bundle_info["source_url"] = source_url
            if has_tiers:
                try:
                    from pathlib import Path as _Path

                    from dbxmetagen.ontology_provenance import (
                        compute_tier_index_hash,
                        tier_indexes_stale,
                    )

                    bundle_info["tier_index_hash"] = compute_tier_index_hash(_Path(tier_dir))
                    bundle_info["tier_indexes_stale"] = tier_indexes_stale(_Path(filepath), _Path(tier_dir))
                except Exception:
                    bundle_info["tier_indexes_stale"] = True
            else:
                try:
                    from pathlib import Path as _Path

                    from dbxmetagen.ontology_provenance import tier_indexes_stale

                    bundle_info["tier_indexes_stale"] = tier_indexes_stale(_Path(filepath))
                except Exception:
                    bundle_info["tier_indexes_stale"] = True
            bundles.append(bundle_info)
        except Exception as e:
            logger.debug("Could not read bundle %s: %s", fname, e)
    return bundles


def _resolve_bundle_path_local(bundle_name: str) -> str:
    """Resolve a bundle key to its YAML path."""
    filename = f"{bundle_name}.yaml" if not bundle_name.endswith(".yaml") else bundle_name
    bd = _find_bundle_dir()
    if bd:
        path = os.path.join(bd, filename)
        if os.path.exists(path):
            return path
    return os.path.join(_BUNDLE_SUBDIR, filename)


@app.get("/api/ontology/bundles")
def get_ontology_bundles():
    """List available ontology bundles with metadata."""
    return _list_bundles_local()


@app.post("/api/ontology/bundles/{bundle_key}/rebuild-indexes")
def rebuild_bundle_indexes(bundle_key: str):
    """Regenerate tier index files for an existing ontology bundle."""
    bd = _find_bundle_dir()
    if not bd:
        return JSONResponse({"error": "Bundle directory not found"}, status_code=500)
    bundle_yaml = os.path.join(bd, f"{bundle_key}.yaml")
    if not os.path.isfile(bundle_yaml):
        return JSONResponse({"error": f"Bundle '{bundle_key}' not found"}, status_code=404)
    tier_dir = os.path.join(bd, bundle_key)
    os.makedirs(tier_dir, exist_ok=True)
    try:
        from pathlib import Path

        from dbxmetagen.ontology_bundle_indexes import (
            build_tiers,
            entities_from_bundle,
            load_edge_catalog,
        )

        entities = entities_from_bundle(Path(bundle_yaml))
        edge_cat = load_edge_catalog(Path(bundle_yaml))
        counts = build_tiers(entities, Path(tier_dir), edge_catalog=edge_cat or None)
    except Exception as e:
        logger.exception("Failed to rebuild indexes for bundle %s", bundle_key)
        return JSONResponse({"error": str(e)}, status_code=500)
    _yaml_cache.clear()
    return {"bundle_key": bundle_key, "counts": counts, "tier_indexes_stale": False}


@app.get("/api/ontology/edge-catalog")
def get_edge_catalog(
    catalog: Optional[str] = Query(None, description="Catalog name (default: env CATALOG_NAME)"),
    schema: Optional[str] = Query(None, description="Schema name (default: env SCHEMA_NAME)"),
    bundle: str = Query("general", description="Ontology bundle key for edge definitions"),
):
    """Return the edge catalog from the bundle YAML and ontology_relationships counts."""
    cat = catalog or CATALOG
    sch = schema or SCHEMA
    if not cat or not sch:
        return {"edges": []}
    _validate_filter(cat, "catalog")
    _validate_filter(sch, "schema")
    rel_table = f"`{cat}`.`{sch}`.`ontology_relationships`"

    path = _resolve_bundle_path_local(bundle)
    ec = {}
    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                raw = yaml.safe_load(f)
            ec = raw.get("ontology", {}).get("edge_catalog", {}) or {}
        except Exception as e:
            logger.warning("edge-catalog YAML load failed: %s", e)

    rel_counts = {}
    rel_valid_invalid = {}
    try:
        rel_rows = execute_sql(
            f"SELECT relationship_name, COUNT(*) as cnt, "
            f"SUM(CASE WHEN validated = true THEN 1 ELSE 0 END) as valid_cnt "
            f"FROM {rel_table} GROUP BY relationship_name",
            timeout=15,
        )
        for r in rel_rows:
            name = r["relationship_name"]
            cnt = r["cnt"]
            valid = r.get("valid_cnt") or 0
            rel_counts[name] = cnt
            rel_valid_invalid[name] = {"valid": int(valid), "invalid": int(cnt - valid)}
    except Exception:
        try:
            rel_rows = execute_sql(
                f"SELECT relationship_name, COUNT(*) as cnt FROM {rel_table} GROUP BY relationship_name",
                timeout=15,
            )
            for r in rel_rows:
                rel_counts[r["relationship_name"]] = r["cnt"]
                rel_valid_invalid[r["relationship_name"]] = {"valid": r["cnt"], "invalid": 0}
        except Exception:
            pass

    edges = []
    seen = set()
    for name, spec in (ec or {}).items():
        seen.add(name)
        vi = rel_valid_invalid.get(name, {"valid": 0, "invalid": 0})
        cnt = rel_counts.get(name, 0)
        if isinstance(spec, dict):
            edges.append({
                "name": name,
                "inverse": spec.get("inverse"),
                "domain": spec.get("domain"),
                "range": spec.get("range"),
                "symmetric": spec.get("symmetric", False),
                "category": spec.get("category", "structural"),
                "count": cnt,
                "valid": vi["valid"],
                "invalid": vi["invalid"],
            })
        else:
            edges.append({
                "name": name,
                "inverse": None,
                "domain": None,
                "range": None,
                "symmetric": False,
                "category": "structural",
                "count": cnt,
                "valid": vi["valid"],
                "invalid": vi["invalid"],
            })

    for name, cnt in rel_counts.items():
        if name not in seen:
            vi = rel_valid_invalid.get(name, {"valid": cnt, "invalid": 0})
            edges.append({
                "name": name,
                "inverse": None,
                "domain": None,
                "range": None,
                "symmetric": False,
                "category": "structural",
                "count": cnt,
                "valid": vi["valid"],
                "invalid": vi["invalid"],
            })
    return {"edges": edges}


@app.get("/api/ontology/entities-summary")
def get_ontology_entities_summary(
    catalog: Optional[str] = Query(None),
    schema: Optional[str] = Query(None),
):
    """Return per-entity summary: table count, column count, avg confidence, bundle vs heuristic, roles, tables list."""
    ent_tbl = fq("ontology_entities")
    cp_tbl = fq("ontology_column_properties")
    where_ent = "source_tables IS NOT NULL AND SIZE(source_tables) > 0"
    where_cp = "1=1"
    if catalog and schema and _SAFE_IDENT_RE.match(catalog) and _SAFE_IDENT_RE.match(schema):
        prefix = f"{catalog}.{schema}."
        where_ent = f"{where_ent} AND EXISTS(source_tables, t -> t LIKE '{_esc_sql(prefix)}%')"
        where_cp = f"table_name LIKE '{_esc_sql(prefix)}%'"

    entities = []
    try:
        ent_rows = execute_sql(
            f"""
            SELECT entity_type, source_ontology, entity_uri, EXPLODE(source_tables) AS table_name
            FROM {ent_tbl}
            WHERE {where_ent}
            """,
            timeout=30,
        )
        tables_by_entity = {}
        source_onto_by_entity: dict[str, set] = {}
        uri_by_entity: dict[str, set] = {}
        for r in ent_rows:
            et = r.get("entity_type")
            tn = r.get("table_name")
            if et and tn:
                tables_by_entity.setdefault(et, set()).add(tn)
            if et and r.get("source_ontology"):
                source_onto_by_entity.setdefault(et, set()).add(r["source_ontology"])
            if et and r.get("entity_uri"):
                uri_by_entity.setdefault(et, set()).add(r["entity_uri"])
    except Exception as e:
        logger.debug("entities-summary entities failed: %s", e)
        return {"entities": []}

    try:
        cp_cols = execute_sql(f"DESCRIBE TABLE {cp_tbl}", timeout=10)
        has_discovery = any(c.get("col_name") == "discovery_method" for c in cp_cols)
    except Exception:
        has_discovery = False

    try:
        agg_expr = """
            owning_entity_type,
            COUNT(*) AS column_count,
            ROUND(AVG(confidence), 2) AS avg_confidence,
            COUNT(DISTINCT table_name) AS table_count
        """
        if has_discovery:
            agg_expr += """,
            SUM(CASE WHEN COALESCE(discovery_method, '') LIKE '%bundle%' OR discovery_method = 'bundle_match' THEN 1 ELSE 0 END) AS bundle_matches,
            SUM(CASE WHEN NOT (COALESCE(discovery_method, '') LIKE '%bundle%' OR discovery_method = 'bundle_match') THEN 1 ELSE 0 END) AS heuristic_matches
        """
        else:
            agg_expr += """,
            0 AS bundle_matches,
            COUNT(*) AS heuristic_matches
        """
        cp_agg = execute_sql(
            f"""
            SELECT {agg_expr}
            FROM {cp_tbl}
            WHERE owning_entity_type IS NOT NULL AND {where_cp}
            GROUP BY owning_entity_type
            """,
            timeout=30,
        )
    except Exception as e:
        logger.debug("entities-summary column props failed: %s", e)
        cp_agg = []

    try:
        role_rows = execute_sql(
            f"""
            SELECT owning_entity_type, property_role, COUNT(*) AS cnt
            FROM {cp_tbl}
            WHERE owning_entity_type IS NOT NULL AND property_role IS NOT NULL AND {where_cp}
            GROUP BY owning_entity_type, property_role
            """,
            timeout=20,
        )
        roles_by_entity = {}
        for r in role_rows:
            et = r["owning_entity_type"]
            role = r["property_role"] or "attribute"
            cnt = int(r["cnt"] or 0)
            roles_by_entity.setdefault(et, {})[role] = cnt
    except Exception:
        roles_by_entity = {}

    entity_types = set(tables_by_entity.keys())
    for row in cp_agg:
        entity_types.add(row["owning_entity_type"])

    for et in sorted(entity_types):
        tables = sorted(tables_by_entity.get(et, []))
        table_count = len(tables)
        row = next((r for r in cp_agg if r["owning_entity_type"] == et), None)
        col_count = int(row["column_count"]) if row else 0
        avg_conf = float(row["avg_confidence"] or 0) if row else 0
        bundle_m = int(row.get("bundle_matches") or 0) if row else 0
        heur_m = int(row.get("heuristic_matches") or 0) if row else 0
        if table_count == 0 and row:
            table_count = int(row.get("table_count") or 0)
        entities.append({
            "entity_type": et,
            "table_count": table_count,
            "column_count": col_count,
            "avg_confidence": round(avg_conf, 2),
            "bundle_matches": bundle_m,
            "heuristic_matches": heur_m,
            "roles": roles_by_entity.get(et, {}),
            "tables": tables,
            "source_ontology": ", ".join(sorted(source_onto_by_entity.get(et, set()))) or None,
            "entity_uri": next(iter(uri_by_entity.get(et, set())), None),
        })
    return {"entities": entities}


@app.get("/api/ontology/entity-summary")
def get_entity_summary():
    """Return entity types with table/column/relationship counts from ontology tables."""
    ent_tbl = fq("ontology_entities")
    cp_tbl = fq("ontology_column_properties")
    rel_tbl = fq("ontology_relationships")
    entities = []
    try:
        # Entity summary: entity_type, table_count, role (prefer primary)
        summary_rows = execute_sql(
            f"""
            SELECT entity_type, COUNT(DISTINCT t) AS table_count,
                   COALESCE(MAX(CASE WHEN COALESCE(entity_role, 'primary') = 'primary' THEN 'primary' END), 'secondary') AS role
            FROM (
                SELECT entity_type, COALESCE(entity_role, 'primary') AS entity_role, EXPLODE(source_tables) AS t
                FROM {ent_tbl}
                WHERE source_tables IS NOT NULL AND SIZE(source_tables) > 0
            ) sub
            GROUP BY entity_type
            ORDER BY table_count DESC
            """,
            timeout=30,
        )
        entity_by_type = {r["entity_type"]: {"entity_type": r["entity_type"], "table_count": r["table_count"], "role": r.get("role", "primary")} for r in summary_rows}
    except Exception as e:
        logger.debug("entity-summary table count failed: %s", e)
        return {"entities": []}
    try:
        col_counts = execute_sql(
            f"""
            SELECT owning_entity_type, COUNT(*) AS cnt
            FROM {cp_tbl}
            WHERE owning_entity_type IS NOT NULL
            GROUP BY owning_entity_type
            """,
            timeout=15,
        )
        for r in col_counts:
            et = r["owning_entity_type"]
            if et in entity_by_type:
                entity_by_type[et]["column_count"] = r["cnt"]
            else:
                entity_by_type[et] = {"entity_type": et, "table_count": 0, "column_count": r["cnt"]}
    except Exception:
        pass
    try:
        rel_counts = execute_sql(
            f"""
            SELECT src_entity_type AS et FROM {rel_tbl} WHERE src_entity_type IS NOT NULL
            UNION ALL
            SELECT dst_entity_type AS et FROM {rel_tbl} WHERE dst_entity_type IS NOT NULL
            """,
            timeout=15,
        )
        rel_by_type = Counter(r["et"] for r in rel_counts)
        for et, cnt in rel_by_type.items():
            if et in entity_by_type:
                entity_by_type[et]["relationship_count"] = cnt
            else:
                entity_by_type[et] = {"entity_type": et, "table_count": 0, "relationship_count": cnt}
    except Exception:
        pass
    for e in entity_by_type.values():
        e.setdefault("column_count", 0)
        e.setdefault("relationship_count", 0)
        entities.append(e)
    return {"entities": entities}


@app.get("/api/ontology/entity-detail")
def get_entity_detail(entity_type: str):
    """Return tables and column properties for a specific entity type."""
    if not entity_type or not _SAFE_IDENT_RE.match(entity_type.replace(".", "x")):
        return {"tables": [], "properties": []}
    ent_tbl = fq("ontology_entities")
    cp_tbl = fq("ontology_column_properties")
    tables = []
    properties = []
    entity_uri = None
    source_ontology = None
    try:
        rows = execute_sql(
            f"""
            SELECT entity_type, entity_uri, source_ontology, EXPLODE(source_tables) AS table_name
            FROM {ent_tbl}
            WHERE entity_type = '{entity_type.replace("'", "''")}'
              AND source_tables IS NOT NULL AND SIZE(source_tables) > 0
            """,
            timeout=30,
        )
        tables = sorted(set(r["table_name"] for r in rows if r.get("table_name")))
        for r in rows:
            if r.get("entity_uri") and not entity_uri:
                entity_uri = r["entity_uri"]
            if r.get("source_ontology") and not source_ontology:
                source_ontology = r["source_ontology"]
    except Exception as e:
        logger.debug("entity-detail tables failed: %s", e)
    try:
        prop_rows = execute_sql(
            f"""
            SELECT table_name, column_name, property_role, confidence, linked_entity_type
            FROM {cp_tbl}
            WHERE owning_entity_type = '{entity_type.replace("'", "''")}'
            ORDER BY table_name, column_name
            """,
            timeout=15,
        )
        properties = [dict(r) for r in prop_rows]
    except Exception as e:
        logger.debug("entity-detail properties failed: %s", e)
    try:
        # Merge description from bundle if available
        bundles = _list_bundles_local()
        for b in bundles:
            path = _resolve_bundle_path_local(b["key"])
            if os.path.exists(path):
                with open(path, "r") as f:
                    raw = yaml.safe_load(f)
                defs = raw.get("ontology", {}).get("entities", {}).get("definitions", {})
                if entity_type in defs:
                    desc = defs[entity_type].get("description")
                    if desc:
                        return {"tables": tables, "properties": properties, "description": desc,
                                "entity_uri": entity_uri, "source_ontology": source_ontology}
    except Exception:
        pass
    return {"tables": tables, "properties": properties, "entity_uri": entity_uri, "source_ontology": source_ontology}


def _resolve_domain_config_path(key: str) -> str:
    """Resolve a domain config key to a file path for the job parameter."""
    bundles = {b["key"]: b for b in _list_bundles_local()}
    if key in bundles:
        return _resolve_bundle_path_local(key)
    cfg_dir = _find_domain_config_dir()
    if cfg_dir:
        path = os.path.join(cfg_dir, f"{key}.yaml")
        if os.path.exists(path):
            return path
    return key


@cached(_yaml_cache, key=lambda: "domain_configs", lock=_yaml_lock)
def _list_domain_configs_cached() -> list[dict]:
    items = []
    cfg_dir = _find_domain_config_dir()
    if cfg_dir:
        for fname in sorted(os.listdir(cfg_dir)):
            if not fname.startswith("domain_config") or not fname.endswith(".yaml"):
                continue
            file_key = fname.replace(".yaml", "")
            try:
                with open(os.path.join(cfg_dir, fname), "r") as f:
                    raw = yaml.safe_load(f)
                domain_count = len(raw.get("domains", {})) if raw else 0
            except Exception:
                domain_count = 0
            items.append({
                "key": file_key,
                "name": file_key.replace("_", " ").replace("domain config ", "").title() + " (standalone)",
                "source": "file",
                "domain_count": domain_count,
            })
    return items


@app.get("/api/domain-configs")
def list_domain_configs():
    """List standalone domain config YAML files. Cached 300s."""
    return _list_domain_configs_cached()


class OntologyApplyItem(BaseModel):
    entity_type: str
    source_tables: Union[list[str], str]
    source_columns: Optional[list[str]] = None
    entity_role: Optional[str] = None


class OntologyApplyBody(BaseModel):
    selections: list[OntologyApplyItem]


def _build_ontology_tag_ddl(
    selections: Optional[list] = None,
    identifiers: Optional[list[str]] = None,
) -> list[str]:
    """Build ontology tag DDL statements without executing them.

    Returns a list of ALTER TABLE ... SET TAGS SQL strings.
    """
    wh_id = os.environ.get("WAREHOUSE_ID", "")
    if not wh_id:
        raise HTTPException(500, detail="WAREHOUSE_ID not configured")
    ent_tbl = fq("ontology_entities")
    cp_tbl = fq("ontology_column_properties")
    tkb_tbl = fq("table_knowledge_base")
    rel_tbl = fq("ontology_relationships")
    stmts: list[str] = []

    allowed_pairs: Optional[set] = None
    allowed_tables: Optional[set] = None
    if selections:
        allowed_pairs = set()
        allowed_tables = set()
        for sel in selections:
            et = sel.get("entity_type", "") if isinstance(sel, dict) else getattr(sel, "entity_type", "")
            tbls = sel.get("source_tables", []) if isinstance(sel, dict) else getattr(sel, "source_tables", [])
            if isinstance(tbls, str):
                tbls = [tbls]
            for t in tbls:
                allowed_pairs.add((et.strip(), t.strip()))
                allowed_tables.add(t.strip())
    if identifiers:
        id_set = {t.strip() for t in identifiers}
        if allowed_tables is not None:
            allowed_tables &= id_set
        else:
            allowed_tables = id_set

    try:
        ent_q = f"""
            SELECT e.entity_type, e.confidence, e.source_tables, e.entity_role,
                   COALESCE(e.attributes['granularity'], 'table') AS granularity
            FROM {ent_tbl} e
            WHERE e.confidence >= 0.5
        """
        entities = execute_sql(ent_q, warehouse_id=wh_id, timeout=60)
    except Exception as e:
        raise HTTPException(404, detail=f"ontology_entities not found: {e}")

    try:
        tkb = execute_sql(f"SELECT table_name, domain FROM {tkb_tbl}", warehouse_id=wh_id, timeout=30)
        tkb_domain = {r["table_name"]: (r.get("domain") or "") for r in tkb}
    except Exception:
        tkb_domain = {}

    seen_tables: dict[str, dict] = {}
    for e in entities:
        tables = e.get("source_tables") or []
        if isinstance(tables, str):
            try:
                tables = json.loads(tables) if tables.startswith("[") else [tables]
            except Exception:
                tables = [tables]
        et = (e.get("entity_type") or "").strip()
        conf = e.get("confidence")
        role = (e.get("entity_role") or "primary").strip()
        gran = (e.get("granularity") or "table").strip()
        if not et or gran != "table" or role != "primary":
            continue
        for tbl in tables:
            if not tbl or not isinstance(tbl, str):
                continue
            tbl = tbl.strip()
            if allowed_pairs is not None and (et, tbl) not in allowed_pairs:
                continue
            if allowed_tables is not None and tbl not in allowed_tables:
                continue
            domain = tkb_domain.get(tbl, "")
            prev = seen_tables.get(tbl)
            conf_f = float(conf or 0)
            if prev:
                if et not in prev["entity_type"]:
                    prev["entity_type"] = prev["entity_type"] + "," + et
                prev["conf_max"] = max(prev.get("conf_max", 0), conf_f)
                if domain:
                    prev["domain"] = domain
            else:
                seen_tables[tbl] = {"entity_type": et, "conf_max": conf_f, "domain": domain}

    for tbl, vals in seen_tables.items():
        if not _SAFE_IDENT_RE.match(tbl.replace(".", "x")):
            continue
        conf_str = str(round(vals.get("conf_max", 0), 2))
        tags = [f"'ontology_entity_type' = '{_esc_sql(vals['entity_type'])}'"]
        tags.append(f"'ontology_confidence' = '{conf_str}'")
        if vals.get("domain"):
            tags.append(f"'ontology_domain' = '{_esc_sql(vals['domain'])}'")
        stmts.append(f"ALTER TABLE {tbl} SET TAGS ({', '.join(tags)});")

    try:
        cp_q = f"""
            SELECT table_name, column_name, property_role, confidence, linked_entity_type
            FROM {cp_tbl}
            WHERE confidence >= 0.5
        """
        props = execute_sql(cp_q, warehouse_id=wh_id, timeout=60)
    except Exception:
        props = []

    try:
        rels = execute_sql(
            f"SELECT evidence_table, evidence_column, relationship_name FROM {rel_tbl}",
            warehouse_id=wh_id, timeout=30,
        )
        rel_edge = {(r.get("evidence_table"), r.get("evidence_column")): (r.get("relationship_name") or "") for r in rels}
    except Exception:
        rel_edge = {}

    for p in props:
        tbl = (p.get("table_name") or "").strip()
        col = (p.get("column_name") or "").strip()
        role = (p.get("property_role") or "").strip()
        linked = (p.get("linked_entity_type") or "").strip()
        conf = p.get("confidence")
        conf_str = str(round(float(conf), 2)) if conf is not None else "0"
        if not tbl or not col or not _SAFE_IDENT_RE.match(tbl.replace(".", "x")):
            continue
        if allowed_tables is not None and tbl not in allowed_tables:
            continue
        col_safe = col.replace("`", "")
        edge = rel_edge.get((tbl, col), "")
        tags = [f"'ontology_property_role' = '{_esc_sql(role)}'"]
        tags.append(f"'ontology_confidence' = '{conf_str}'")
        if edge:
            tags.append(f"'ontology_edge' = '{_esc_sql(edge)}'")
        if linked:
            tags.append(f"'ontology_linked_entity' = '{_esc_sql(linked)}'")
        stmts.append(f"ALTER TABLE {tbl} ALTER COLUMN `{col_safe}` SET TAGS ({', '.join(tags)});")

    return stmts


def _apply_ontology_tags_from_tables(
    selections: Optional[list] = None,
) -> dict:
    """Read ontology_entities and ontology_column_properties, apply ontology_* UC tags to tables/columns.

    Delegates DDL building to _build_ontology_tag_ddl, then executes via
    _execute_stmts_batched (column-level ALTERs are batched per table).
    """
    stmts = _build_ontology_tag_ddl(selections=selections)

    def _extract_table(sql):
        return sql.split("ALTER TABLE", 1)[-1].split("SET TAGS")[0].split("ALTER COLUMN")[0].strip()

    table_stmts = [s for s in stmts if "ALTER COLUMN" not in s.upper()]
    col_stmts = [s for s in stmts if "ALTER COLUMN" in s.upper()]

    t_applied, t_errors = _execute_stmts_batched(table_stmts, batch=False, timeout=30)
    c_applied, c_errors = _execute_stmts_batched(col_stmts, batch=True, timeout=30)

    t_err_set = {e["statement"][:80] for e in t_errors}
    table_results = [
        {"table": _extract_table(s), "ok": s.rstrip(";").strip()[:80] not in t_err_set}
        for s in table_stmts if s.rstrip(";").strip()
    ]
    col_results = (
        [{"table": t, "ok": True} for t in ["batched"] * c_applied]
        + [{"table": e.get("statement", "")[:60], "ok": False, "error": e["error"]} for e in c_errors]
    )

    t_ok = sum(1 for r in table_results if r["ok"])
    c_ok = c_applied
    summary = {
        "tables_tagged": t_ok,
        "tables_failed": len(table_results) - t_ok,
        "columns_tagged": c_ok,
        "columns_failed": len(c_errors),
    }
    return {
        "summary": summary,
        "table_results": table_results,
        "column_results": col_results,
        "results": table_results,
    }


@app.post("/api/ontology/apply-tags")
def ontology_apply_tags(body: Optional[OntologyApplyBody] = Body(default=None)):
    """Apply ontology_* namespaced UC tags from ontology_entities and ontology_column_properties.
    Reads from ontology tables and applies: ontology_entity_type, ontology_domain, ontology_confidence
    at table level; ontology_property_role, ontology_edge, ontology_linked_entity, ontology_confidence
    at column level. Returns a summary of tags applied."""
    sels = None
    if body and body.selections:
        sels = [s.model_dump() for s in body.selections]
    return _apply_ontology_tags_from_tables(selections=sels)


@app.post("/api/ontology/apply-all-tags")
def ontology_apply_all_tags():
    """Alias for apply-tags: read ontology tables and apply ontology.* namespaced UC tags."""
    return _apply_ontology_tags_from_tables()


@app.get("/api/ontology/export")
def export_ontology_jsonld(
    catalog: str = Query(..., description="Catalog name"),
    schema: str = Query(..., description="Schema name"),
    format: str = Query("jsonld", description="jsonld or jsonld_download"),
):
    """Export discovered ontology as JSON-LD from ontology_entities, ontology_column_properties, ontology_relationships."""
    if not _SAFE_IDENT_RE.match(catalog) or not _SAFE_IDENT_RE.match(schema):
        raise HTTPException(400, detail="Invalid catalog or schema")
    wh_id = os.environ.get("WAREHOUSE_ID", "")
    if not wh_id:
        raise HTTPException(500, detail="WAREHOUSE_ID not configured")
    base = f"`{catalog}`.`{schema}`"

    entities_rows = []
    rels_rows = []
    try:
        entities_rows = execute_sql(f"SELECT * FROM {base}.`ontology_entities`", warehouse_id=wh_id, timeout=60)
    except HTTPException as he:
        if he.status_code == 404:
            pass
        else:
            raise
    except Exception:
        pass

    try:
        rels_rows = execute_sql(f"SELECT * FROM {base}.`ontology_relationships`", warehouse_id=wh_id, timeout=60)
    except HTTPException as he:
        if he.status_code == 404:
            pass
        else:
            raise
    except Exception:
        pass

    col_props_rows = []
    try:
        col_props_rows = execute_sql(
            f"SELECT * FROM {base}.`ontology_column_properties`",
            warehouse_id=wh_id, timeout=60,
        )
    except Exception:
        pass

    _SCHEMA_ORG_TYPE_MAP = {
        "Person": "schema:Person",
        "Organization": "schema:Organization",
        "Product": "schema:Product",
        "Event": "schema:Event",
        "Location": "schema:Place",
        "Patient": "schema:Patient",
        "Document": "schema:DigitalDocument",
    }

    context = {
        "schema": "https://schema.org/",
        "ontology": "urn:dbxmetagen:ontology:",
        "entity_type": "ontology:entityType",
        "property_role": "ontology:propertyRole",
        "confidence": "ontology:confidence",
        "source_tables": "ontology:sourceTables",
    }
    graph = []
    for e in entities_rows:
        et = e.get("entity_type") or ""
        if not et:
            continue
        src_tables = e.get("source_tables") or []
        if isinstance(src_tables, str):
            src_tables = [src_tables] if src_tables else []
        schema_type = _SCHEMA_ORG_TYPE_MAP.get(et, "schema:Thing")
        node = {
            "@id": f"ontology:Entity/{et}",
            "@type": schema_type,
            "entity_type": et,
            "schema:name": et,
            "confidence": e.get("confidence"),
            "source_tables": src_tables,
        }
        graph.append(node)

    for r in rels_rows:
        src = r.get("src_entity_type") or ""
        dst = r.get("dst_entity_type") or ""
        name = r.get("relationship_name") or "references"
        rel_id = f"{src}_{name}_{dst}"
        graph.append({
            "@id": f"ontology:Relationship/{rel_id}",
            "@type": "ontology:Relationship",
            "ontology:from": {"@id": f"ontology:Entity/{src}"},
            "ontology:to": {"@id": f"ontology:Entity/{dst}"},
            "ontology:relationshipName": name,
        })

    for cp in col_props_rows:
        entity = cp.get("owning_entity_type") or ""
        col = cp.get("column_name") or ""
        tbl = cp.get("table_name") or ""
        if not entity or not col:
            continue
        prop_id = f"{tbl}.{col}".replace("`", "")
        graph.append({
            "@id": f"ontology:Property/{prop_id}",
            "@type": "ontology:ColumnProperty",
            "ontology:owningEntity": {"@id": f"ontology:Entity/{entity}"},
            "ontology:columnName": col,
            "ontology:tableName": tbl,
            "property_role": cp.get("property_role"),
            "confidence": cp.get("confidence"),
            "ontology:discoveryMethod": cp.get("discovery_method"),
        })

    result = {"@context": context, "@graph": graph}
    if format == "jsonld_download":
        body = json.dumps(result, indent=2)
        return StreamingResponse(
            iter([body]),
            media_type="application/ld+json",
            headers={"Content-Disposition": "attachment; filename=ontology.jsonld"},
        )
    return JSONResponse(content=result)


@cached(_yaml_cache, key=lambda: "entity_type_options", lock=_yaml_lock)
def _entity_type_options_cached() -> list[str]:
    bd = _find_bundle_dir()
    if not bd:
        return []
    types = set()
    for fname in os.listdir(bd):
        if not fname.endswith(".yaml"):
            continue
        try:
            with open(os.path.join(bd, fname), "r") as f:
                raw = yaml.safe_load(f)
            defs = raw.get("ontology", {}).get("entities", {}).get("definitions", {})
            types.update(defs.keys())
        except Exception:
            pass
    return sorted(types)


@app.get("/api/ontology/entity-type-options")
def get_entity_type_options():
    """Return deduplicated entity type names from all ontology bundle YAMLs. Cached 300s."""
    return _entity_type_options_cached()


class UpdateEntityTypeBody(BaseModel):
    entity_id: str
    new_entity_type: str


@app.post("/api/ontology/update-entity-type")
def update_entity_type(body: UpdateEntityTypeBody):
    """Update the entity_type for a specific ontology entity row."""
    eid = body.entity_id.replace("'", "''")
    new_val = body.new_entity_type.strip().replace("'", "''")
    if not new_val:
        raise HTTPException(400, detail="new_entity_type must not be empty")
    execute_sql(
        f"UPDATE {fq('ontology_entities')} SET entity_type = '{new_val}' "
        f"WHERE entity_id = '{eid}'"
    )
    return {"updated": True, "entity_id": body.entity_id, "new_entity_type": body.new_entity_type}


class SetRecommendedEntityBody(BaseModel):
    table_name: str
    entity_type: str
    entity_role: str = "primary"


@app.post("/api/ontology/set-recommended-entity")
def set_recommended_entity(body: SetRecommendedEntityBody):
    """Insert a steward override entity and optionally demote previous primary."""
    import uuid as _u
    ent_tbl = fq("ontology_entities")
    tbl_clean = body.table_name.strip()
    et = body.entity_type.strip()
    if not tbl_clean or not et:
        raise HTTPException(400, "table_name and entity_type required")
    if body.entity_role == "primary":
        try:
            execute_sql(f"""
                UPDATE {ent_tbl}
                SET entity_role = 'referenced', updated_at = current_timestamp()
                WHERE entity_role = 'primary'
                  AND EXISTS(source_tables, t -> t = '{tbl_clean}')
            """, timeout=30)
        except Exception as e:
            logger.warning("Failed to demote previous primary for %s: %s", tbl_clean, e)
    eid = str(_u.uuid4())
    try:
        execute_sql(f"""
            INSERT INTO {ent_tbl}
            (entity_id, entity_name, entity_type, source_tables, source_columns,
             confidence, discovery_confidence, entity_role, is_canonical, auto_discovered,
             validated, validation_notes, created_at, updated_at)
            VALUES ('{eid}', '{et}', '{et}', ARRAY('{tbl_clean}'), ARRAY(),
                    1.0, 1.0, '{body.entity_role}', false, false,
                    true, 'Steward override', current_timestamp(), current_timestamp())
        """, timeout=30)
    except Exception as e:
        raise HTTPException(500, f"Insert failed: {e}")
    return {"ok": True, "entity_id": eid, "entity_type": et}


class UpdateColumnPropertyBody(BaseModel):
    property_id: str
    property_role: str
    linked_entity_type: Optional[str] = None


@app.post("/api/ontology/update-column-property")
def update_column_property(body: UpdateColumnPropertyBody):
    """Upsert property_role (and optionally linked_entity_type) on a column property row."""
    cp_tbl = fq("ontology_column_properties")
    pid = body.property_id.replace("'", "''")
    role = body.property_role.replace("'", "''")
    linked = body.linked_entity_type
    sets = [f"property_role = '{role}'", "updated_at = current_timestamp()"]
    if linked is not None:
        sets.append(f"linked_entity_type = '{linked.replace(chr(39), chr(39)*2)}'")
    try:
        execute_sql(f"UPDATE {cp_tbl} SET {', '.join(sets)} WHERE property_id = '{pid}'", timeout=30)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


class ApplyPropertyTagsBody(BaseModel):
    items: list[dict]


@app.post("/api/ontology/apply-property-tags")
def apply_property_tags(body: ApplyPropertyTagsBody):
    """Apply property_role tags to columns via ALTER TABLE ALTER COLUMN SET TAGS (batched per table)."""
    stmts: list[str] = []
    stmt_meta: list[dict] = []
    for item in body.items:
        tbl = (item.get("table_name") or "").strip()
        col = (item.get("column_name") or "").strip()
        role = (item.get("property_role") or "").strip()
        if not (tbl and col and role):
            continue
        col_safe = col.replace("`", "")
        linked = (item.get("linked_entity_type") or "").strip()
        tags = [f"'property_role' = '{role}'"]
        if linked:
            tags.append(f"'linked_entity_type' = '{linked}'")
        sql = f"ALTER TABLE {tbl} ALTER COLUMN `{col_safe}` SET TAGS ({', '.join(tags)})"
        stmts.append(sql)
        stmt_meta.append({"table": tbl, "column": col, "sql": sql})

    applied, errors = _execute_stmts_batched(stmts, batch=True, timeout=30)
    err_prefixes = {e["statement"][:80] for e in errors}
    results = []
    for meta in stmt_meta:
        prefix = meta["sql"].rstrip(";").strip()[:80]
        if prefix in err_prefixes:
            err = next((e for e in errors if e["statement"][:80] == prefix), {})
            results.append({"table": meta["table"], "column": meta["column"], "ok": False, "sql": meta["sql"], "error": err.get("error", "batch failed")})
        else:
            results.append({"table": meta["table"], "column": meta["column"], "ok": True, "sql": meta["sql"]})

    # --- Knowledge base write-back ---
    col_kb = fq("column_knowledge_base")
    role_lookup = {(item.get("table_name", "").strip(), item.get("column_name", "").strip()): item.get("property_role", "").strip() for item in body.items}
    for r in results:
        if not r.get("ok"):
            continue
        role_val = role_lookup.get((r["table"], r["column"]), "")
        if not role_val:
            continue
        try:
            _ensure_column(col_kb, "property_role")
            col_safe = r["column"].replace("'", "''")
            execute_sql(
                f"UPDATE {col_kb} SET property_role = '{role_val.replace(chr(39), chr(39)*2)}', "
                f"updated_at = current_timestamp() "
                f"WHERE table_name = '{r['table']}' AND column_name = '{col_safe}'",
                timeout=15,
            )
        except Exception as e:
            logger.warning("KB write-back (property_role) failed for %s.%s: %s", r.get("table"), r.get("column"), e)

    return {"results": results}


@app.get("/api/ontology/override-stats")
def get_override_stats(
    catalog: Optional[str] = Query(None),
    schema: Optional[str] = Query(None),
    bundle: str = Query("general", description="Ontology bundle key"),
):
    """Track steward overrides and suggest bundle refinements from override patterns."""
    cp_tbl = fq("ontology_column_properties")
    ent_tbl = fq("ontology_entities")
    where_cp = "1=1"
    where_ent = "entity_role = 'primary' AND source_tables IS NOT NULL AND SIZE(source_tables) > 0"
    if catalog and schema and _SAFE_IDENT_RE.match(catalog) and _SAFE_IDENT_RE.match(schema):
        prefix = f"{catalog}.{schema}."
        where_ent = f"{where_ent} AND EXISTS(source_tables, t -> t LIKE '{_esc_sql(prefix)}%')"
        where_cp = f"table_name LIKE '{_esc_sql(prefix)}%'"

    try:
        cp_cols = execute_sql(f"DESCRIBE TABLE {cp_tbl}", timeout=10)
        has_dm = any(c.get("col_name") == "discovery_method" for c in cp_cols)
    except Exception:
        has_dm = False

    cp_rows = []
    try:
        cols = "table_name, column_name, property_role, owning_entity_type, property_name" + (", discovery_method" if has_dm else "")
        cp_rows = execute_sql(f"SELECT {cols} FROM {cp_tbl} WHERE {where_cp}", timeout=30)
    except Exception as e:
        logger.debug("override-stats cp failed: %s", e)

    path = _resolve_bundle_path_local(bundle)
    bundle_role_by_entity_col = {}
    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                raw = yaml.safe_load(f)
            defs = raw.get("ontology", {}).get("entities", {}).get("definitions", {})
            for ename, espec in (defs or {}).items():
                for pname, pval in (espec.get("properties") or {}).items():
                    role = pval.get("role", "")
                    for attr in (pval.get("typical_attributes") or []):
                        bundle_role_by_entity_col[(ename, str(attr).lower())] = role
        except Exception as e:
            logger.debug("override-stats bundle load failed: %s", e)

    overrides = []
    for r in cp_rows:
        dm = r.get("discovery_method") or ""
        if "bundle" not in dm.lower() and dm != "bundle_match":
            continue
        et = r.get("owning_entity_type")
        col = (r.get("column_name") or "").lower()
        current_role = (r.get("property_role") or "").strip()
        bundle_role = bundle_role_by_entity_col.get((et, col)) if et else None
        if bundle_role and current_role and current_role != bundle_role:
            overrides.append({
                "entity_type": et,
                "column_name": r.get("column_name"),
                "original_role": bundle_role,
                "overridden_to": current_role,
                "property_name": r.get("property_name"),
            })

    patterns = []
    pattern_key_counts = {}
    for o in overrides:
        col = o["column_name"] or ""
        pattern = "*_date" if col.endswith("_date") else ("*_id" if col.endswith("_id") else col)
        key = (o["entity_type"], pattern, o["original_role"], o["overridden_to"])
        pattern_key_counts[key] = pattern_key_counts.get(key, 0) + 1

    for (et, pat, orig, over), cnt in pattern_key_counts.items():
        patterns.append({
            "entity_type": et,
            "column_pattern": pat,
            "original_role": orig,
            "overridden_to": over,
            "count": cnt,
            "suggestion": f"Consider adding '{pat}' to {et}.properties with role '{over}'",
        })

    suggested = []
    prop_overrides = {}
    for o in overrides:
        prop = o.get("property_name") or o["column_name"]
        key = (o["entity_type"], prop, o["overridden_to"])
        prop_overrides[key] = prop_overrides.get(key, 0) + 1
    for (entity, prop, role), cnt in sorted(prop_overrides.items(), key=lambda x: -x[1]):
        suggested.append({
            "entity": entity,
            "property": prop,
            "suggested_role": role,
            "evidence_count": cnt,
            "suggestion_id": f"{entity}|{prop}|{role}",
        })

    return {
        "override_count": len(overrides),
        "patterns": patterns,
        "suggested_bundle_updates": suggested,
    }


class ApplySuggestionsBody(BaseModel):
    suggestion_ids: list[str]


@app.post("/api/ontology/apply-suggestions")
def apply_suggestions(body: ApplySuggestionsBody):
    """Apply suggested property_role updates to ontology_column_properties."""
    cp_tbl = fq("ontology_column_properties")
    applied = 0
    for sid in body.suggestion_ids or []:
        parts = sid.split("|")
        if len(parts) != 3:
            continue
        entity, property_name, role = parts[0], parts[1], parts[2]
        entity_esc = entity.replace("'", "''")
        prop_esc = property_name.replace("'", "''")
        role_esc = role.replace("'", "''")
        try:
            execute_sql(
                f"UPDATE {cp_tbl} SET property_role = '{role_esc}', updated_at = current_timestamp() "
                f"WHERE owning_entity_type = '{entity_esc}' AND (property_name = '{prop_esc}' OR column_name = '{prop_esc}')",
                timeout=30,
            )
            applied += 1
        except Exception as e:
            logger.warning("apply-suggestions failed for %s: %s", sid, e)
    return {"applied": applied, "suggestion_ids": body.suggestion_ids}


class SetReviewStatusBody(BaseModel):
    table_name: str
    review_status: str


@app.post("/api/ontology/set-review-status")
def set_review_status(body: SetReviewStatusBody):
    """Set review_status on a table in table_knowledge_base."""
    tbl_kb = fq("table_knowledge_base")
    valid = {"unreviewed", "in_review", "approved"}
    if body.review_status not in valid:
        raise HTTPException(400, f"review_status must be one of {valid}")
    tname = body.table_name.strip()
    status = body.review_status
    try:
        cols = execute_sql(f"DESCRIBE TABLE {tbl_kb}", timeout=15)
        if not any(r.get("col_name") == "review_status" for r in cols):
            execute_sql(f"ALTER TABLE {tbl_kb} ADD COLUMN review_status STRING", timeout=15)
    except Exception:
        pass
    try:
        execute_sql(f"""
            UPDATE {tbl_kb}
            SET review_status = '{status}', updated_at = current_timestamp()
            WHERE table_name = '{tname}'
        """, timeout=30)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(500, str(e))


class FKApplyPredictionItem(BaseModel):
    src_table: str
    src_column: str
    dst_table: str
    dst_column: str


class FKApplyPredictionsBody(BaseModel):
    predictions: list[FKApplyPredictionItem]


@app.post("/api/analytics/fk-apply-from-predictions")
def fk_apply_from_predictions(body: FKApplyPredictionsBody):
    """Generate and execute FK constraint DDL from prediction data."""
    results = []
    for p in body.predictions:
        src_short = p.src_column.split(".")[-1] if "." in p.src_column else p.src_column
        dst_short = p.dst_column.split(".")[-1] if "." in p.dst_column else p.dst_column
        constraint = f"fk_{src_short}_{dst_short}"
        ddl = f"ALTER TABLE {p.src_table} ADD CONSTRAINT {constraint} FOREIGN KEY ({src_short}) REFERENCES {p.dst_table}({dst_short})"
        try:
            execute_sql(ddl, timeout=60)
            results.append({"ddl": ddl, "ok": True})
        except Exception as e:
            err = str(e)
            if "PERMISSION_DENIED" in err and "MANAGE" in err:
                err += " [Hint: Try 'Apply as Tags' instead -- it only requires APPLY_TAG permission.]"
            results.append({"ddl": ddl, "ok": False, "error": err})
    return {"results": results}


@app.post("/api/analytics/fk-generate-sql")
def fk_generate_sql(body: FKApplyPredictionsBody):
    """Generate FK constraint DDL statements without executing them."""
    statements = []
    for p in body.predictions:
        src_short = p.src_column.split(".")[-1] if "." in p.src_column else p.src_column
        dst_short = p.dst_column.split(".")[-1] if "." in p.dst_column else p.dst_column
        constraint = f"fk_{src_short}_{dst_short}"
        statements.append(
            f"ALTER TABLE {p.src_table} ADD CONSTRAINT {constraint} "
            f"FOREIGN KEY ({src_short}) REFERENCES {p.dst_table}({dst_short});"
        )
    return {"sql": "\n".join(statements), "count": len(statements)}


@app.post("/api/analytics/fk-apply-as-tags")
def fk_apply_as_tags(body: FKApplyPredictionsBody):
    """Apply FK relationships as column tags (requires APPLY_TAG, not MANAGE).

    Sets a tag like: ALTER TABLE <src_table> ALTER COLUMN <col> SET TAGS ('fk_references' = '<dst_table>.<col>')
    Batches column-level ALTERs per table via _execute_stmts_batched.
    """
    stmts: list[str] = []
    for p in body.predictions:
        src_col = p.src_column.split(".")[-1] if "." in p.src_column else p.src_column
        dst_col = p.dst_column.split(".")[-1] if "." in p.dst_column else p.dst_column
        tag_val = f"{p.dst_table}.{dst_col}"
        stmts.append(f"ALTER TABLE {p.src_table} ALTER COLUMN `{src_col}` SET TAGS ('fk_references' = '{tag_val}')")

    applied, errors = _execute_stmts_batched(stmts, batch=True, timeout=60)
    err_prefixes = {e["statement"][:80] for e in errors}
    results = []
    for sql in stmts:
        prefix = sql.rstrip(";").strip()[:80]
        if prefix in err_prefixes:
            err = next((e for e in errors if e["statement"][:80] == prefix), {})
            results.append({"sql": sql, "ok": False, "error": err.get("error", "batch failed")})
        else:
            results.append({"sql": sql, "ok": True})
    return {"results": results}


@app.get("/api/ontology/metrics")
def get_ontology_metrics():
    """Return computed ontology health metrics from ontology_metrics table."""
    q = f"""
        SELECT metric_name, description, sql_definition as value, updated_at
        FROM {fq('ontology_metrics')}
        WHERE aggregation_type = 'COMPUTED'
        ORDER BY metric_name
    """
    try:
        return execute_sql(q)
    except HTTPException as e:
        if e.status_code == 404:
            return []
        raise


# ---------------------------------------------------------------------------
# Analytics endpoints
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# FK Predictions
# ---------------------------------------------------------------------------


@app.get("/api/analytics/fk-predictions")
def get_fk_predictions(limit: int = 200):
    """Return predicted foreign key relationships."""
    q = f"SELECT * FROM {fq('fk_predictions')} WHERE src_table != dst_table ORDER BY final_confidence DESC LIMIT {limit}"
    return execute_sql(q)


@app.get("/api/analytics/fk-ddl")
def get_fk_ddl():
    """Return generated FK DDL statements."""
    try:
        q = f"SELECT * FROM {fq('fk_ddl_statements')} ORDER BY confidence DESC"
        return execute_sql(q)
    except HTTPException:
        return []


class FKApplyBody(BaseModel):
    statements: list[str]


@app.post("/api/analytics/fk-apply")
def fk_apply(body: FKApplyBody):
    """Execute selected FK DDL statements. Run FK prediction job first to populate fk_ddl_statements."""
    results = []
    for stmt in (body.statements or []):
        s = (stmt or "").strip()
        if not s or not s.upper().startswith("ALTER TABLE"):
            results.append({"ok": False, "error": "Not an ALTER TABLE statement", "statement": s})
            continue
        try:
            execute_sql(s, timeout=60)
            results.append({"ok": True, "statement": s})
        except Exception as e:
            err = str(e)
            if "PERMISSION_DENIED" in err and "MANAGE" in err:
                err += " [Hint: Try 'Apply as Tags' instead -- it only requires APPLY_TAG permission.]"
            results.append({"ok": False, "error": err, "statement": s})
    return {"results": results}


class FKDeleteBody(BaseModel):
    predictions: list[dict]


@app.post("/api/analytics/fk-delete")
def delete_fk_predictions(body: FKDeleteBody):
    """Delete FK predictions and cascade to fk_ddl_statements and graph_edges."""
    deleted = 0
    errors = []
    preds_tbl = fq("fk_predictions")
    ddl_tbl = fq("fk_ddl_statements")
    edges_tbl = fq("graph_edges")
    for p in (body.predictions or []):
        src_col = _esc_sql(p.get("src_column", ""))
        dst_col = _esc_sql(p.get("dst_column", ""))
        src_tbl = _esc_sql(p.get("src_table", ""))
        dst_tbl = _esc_sql(p.get("dst_table", ""))
        if not src_col or not dst_col:
            errors.append({"prediction": p, "error": "Missing src_column or dst_column"})
            continue
        try:
            execute_sql(
                f"DELETE FROM {preds_tbl} WHERE src_column = '{src_col}' AND dst_column = '{dst_col}'"
                f" AND src_table = '{src_tbl}' AND dst_table = '{dst_tbl}'"
            )
            deleted += 1
        except Exception as e:
            errors.append({"prediction": p, "error": f"fk_predictions: {e}"})
            continue
        try:
            execute_sql(
                f"DELETE FROM {ddl_tbl} WHERE src_column = '{src_col}' AND dst_column = '{dst_col}'"
                f" AND src_table = '{src_tbl}' AND dst_table = '{dst_tbl}'"
            )
        except Exception:
            pass
        src_fq = _esc_sql(p.get("src_table", "") + "." + p.get("src_column", ""))
        dst_fq = _esc_sql(p.get("dst_table", "") + "." + p.get("dst_column", ""))
        try:
            execute_sql(
                f"DELETE FROM {edges_tbl} WHERE source_system = 'fk_predictions' "
                f"AND src = '{src_tbl}' AND dst = '{dst_tbl}'"
            )
            execute_sql(
                f"DELETE FROM {edges_tbl} WHERE source_system = 'fk_predictions' "
                f"AND src = '{src_fq}' AND dst = '{dst_fq}'"
            )
        except Exception:
            pass
        try:
            execute_sql(
                f"DELETE FROM {edges_tbl} WHERE relationship = 'predicted_fk' "
                f"AND ((src = '{src_fq}' AND dst = '{dst_fq}') "
                f"OR (src = '{dst_fq}' AND dst = '{src_fq}'))"
            )
        except Exception:
            pass
    invalidate_query_caches()
    return {"deleted": deleted, "errors": errors}


# ---------------------------------------------------------------------------
# Visualization composite endpoints
# ---------------------------------------------------------------------------


@app.get("/api/viz/fk-map")
def viz_fk_map():
    """Composite data for FK Map visualization. Cached 60s."""
    with _coverage_lock:
        if "fk_map" in _coverage_cache:
            return _coverage_cache["fk_map"]
    tables = graph_query(
        "SELECT id, node_type, domain, security_level, comment "
        "FROM public.graph_nodes WHERE node_type='table' ORDER BY id"
    )
    fk_edges = execute_sql(
        f"SELECT src_table, dst_table, src_column, dst_column, final_confidence "
        f"FROM {fq('fk_predictions')} ORDER BY final_confidence DESC LIMIT 500"
    )
    clusters = execute_sql(
        f"SELECT id, cluster FROM {fq('node_cluster_assignments')} "
        f"WHERE node_type='table' ORDER BY cluster, id"
    )
    result = {"tables": tables, "fk_edges": fk_edges, "clusters": clusters}
    with _coverage_lock:
        _coverage_cache["fk_map"] = result
    return result


# ---------------------------------------------------------------------------
# Unprofiled tables (information_schema coverage)
# ---------------------------------------------------------------------------


@app.get("/api/coverage/summary")
def get_coverage_summary(catalog: Optional[str] = None):
    """Coverage summary: profiled vs unprofiled tables. Cached 60s."""
    cat = catalog or CATALOG
    if not re.fullmatch(r"[a-zA-Z0-9_\-]+", cat):
        raise HTTPException(status_code=400, detail="Invalid catalog name")
    cache_key = f"summary:{cat}"
    with _coverage_lock:
        if cache_key in _coverage_cache:
            return _coverage_cache[cache_key]
    _ALL_TABLE_TYPES = "('MANAGED','EXTERNAL','VIEW','STREAMING_TABLE','MATERIALIZED_VIEW','FOREIGN')"
    q = f"""
        SELECT t.table_catalog, t.table_schema,
               COUNT(*) as total_tables,
               COUNT(kb.table_name) as profiled_tables,
               COUNT(*) - COUNT(kb.table_name) as unprofiled_tables
        FROM system.information_schema.tables t
        LEFT JOIN {fq('table_knowledge_base')} kb
          ON CONCAT(t.table_catalog, '.', t.table_schema, '.', t.table_name) = kb.table_name
        WHERE t.table_catalog = '{cat}'
          AND t.table_schema NOT IN ('information_schema', '__internal')
          AND t.table_type IN {_ALL_TABLE_TYPES}
          AND NOT t.table_name RLIKE '^(__|event_log_[0-9a-f]{{8}}_)'
        GROUP BY t.table_catalog, t.table_schema
        ORDER BY unprofiled_tables DESC
    """
    try:
        result = execute_sql(q)
    except HTTPException as e:
        if e.status_code != 404:
            raise
        q_simple = f"""
            SELECT table_catalog, table_schema, COUNT(*) as total_tables,
                   0 as profiled_tables, COUNT(*) as unprofiled_tables
            FROM system.information_schema.tables
            WHERE table_catalog = '{cat}'
              AND table_schema NOT IN ('information_schema', '__internal')
              AND table_type IN {_ALL_TABLE_TYPES}
              AND NOT table_name RLIKE '^(__|event_log_[0-9a-f]{{8}}_)'
            GROUP BY table_catalog, table_schema
            ORDER BY total_tables DESC
        """
        result = execute_sql(q_simple)
    with _coverage_lock:
        _coverage_cache[cache_key] = result
    return result


@app.get("/api/coverage/type-breakdown")
def get_coverage_type_breakdown(catalog: Optional[str] = None):
    """Count tables per table_type. Cached 60s."""
    cat = catalog or CATALOG
    cache_key = f"type_breakdown:{cat}"
    with _coverage_lock:
        if cache_key in _coverage_cache:
            return _coverage_cache[cache_key]
    q = f"""
        SELECT table_type, COUNT(*) as count
        FROM system.information_schema.tables
        WHERE table_catalog = '{cat}'
          AND table_schema NOT IN ('information_schema', '__internal')
          AND NOT table_name RLIKE '^(__|event_log_[0-9a-f]{{8}}_)'
        GROUP BY table_type
        ORDER BY count DESC
    """
    result = execute_sql(q)
    with _coverage_lock:
        _coverage_cache[cache_key] = result
    return result


@app.get("/api/coverage/metadata-summary")
def get_coverage_metadata_summary(catalog: Optional[str] = None, schema: Optional[str] = None):
    """Metadata completeness rates. Cached 60s."""
    cache_key = f"meta_summary:{catalog}:{schema}"
    with _coverage_lock:
        if cache_key in _coverage_cache:
            return _coverage_cache[cache_key]
    schema_filter = ""
    if catalog and schema:
        schema_filter = f" WHERE table_name LIKE '{catalog}.{schema}.%'"
    elif catalog:
        schema_filter = f" WHERE table_name LIKE '{catalog}.%'"
    result = {}
    try:
        rows = execute_sql(f"""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN comment IS NOT NULL AND comment != '' THEN 1 ELSE 0 END) as with_comments,
                SUM(CASE WHEN has_pii = true OR has_phi = true THEN 1 ELSE 0 END) as with_pii,
                SUM(CASE WHEN domain IS NOT NULL AND domain != '' THEN 1 ELSE 0 END) as with_domain
            FROM {fq('table_knowledge_base')}{schema_filter}
        """)
        result = rows[0] if rows else {}
    except Exception:
        result = {"total": 0, "with_comments": 0, "with_pii": 0, "with_domain": 0}
    if catalog and schema:
        onto_filter = f" WHERE t.table_name LIKE '{catalog}.{schema}.%'"
    elif catalog:
        onto_filter = f" WHERE t.table_name LIKE '{catalog}.%'"
    else:
        onto_filter = ""
    try:
        onto = execute_sql(f"SELECT COUNT(DISTINCT t.table_name) as with_ontology FROM (SELECT EXPLODE(source_tables) as table_name FROM {fq('ontology_entities')}) t{onto_filter}")
        result["with_ontology"] = onto[0]["with_ontology"] if onto else 0
    except Exception:
        result["with_ontology"] = 0
    fk_conf_filter = " WHERE final_confidence >= 0.5"
    if catalog and schema:
        fk_conf_filter += f" AND (src_table LIKE '{catalog}.{schema}.%' OR dst_table LIKE '{catalog}.{schema}.%')"
    elif catalog:
        fk_conf_filter += f" AND (src_table LIKE '{catalog}.%' OR dst_table LIKE '{catalog}.%')"
    try:
        fks = execute_sql(f"""SELECT COUNT(DISTINCT t) as with_fk FROM (
            SELECT src_table AS t FROM {fq('fk_predictions')}{fk_conf_filter}
            UNION
            SELECT dst_table AS t FROM {fq('fk_predictions')}{fk_conf_filter}
        )""")
        result["with_fk"] = fks[0]["with_fk"] if fks else 0
    except Exception:
        result["with_fk"] = 0
    with _coverage_lock:
        _coverage_cache[cache_key] = result
    return result


@app.get("/api/coverage/tables")
def get_coverage_tables(catalog: Optional[str] = None, schema: Optional[str] = None, kb_only: bool = False):
    """List individual tables and whether they've been profiled."""
    cat = catalog or CATALOG
    conditions = [f"t.table_catalog = '{cat}'"]
    if schema:
        conditions.append(f"t.table_schema = '{schema}'")
    else:
        conditions.append("t.table_schema NOT IN ('information_schema', '__internal')")
    conditions.append("t.table_type IN ('MANAGED', 'EXTERNAL', 'VIEW', 'STREAMING_TABLE', 'MATERIALIZED_VIEW', 'FOREIGN')")
    conditions.append("NOT t.table_name RLIKE '^(__|event_log_[0-9a-f]{8}_)'")
    where = " AND ".join(conditions)
    join_type = "INNER" if kb_only else "LEFT"
    q = f"""
        SELECT t.table_catalog, t.table_schema, t.table_name, t.table_type,
               CASE WHEN kb.table_name IS NOT NULL THEN true ELSE false END as is_profiled
        FROM system.information_schema.tables t
        {join_type} JOIN {fq('table_knowledge_base')} kb
          ON CONCAT(t.table_catalog, '.', t.table_schema, '.', t.table_name) = kb.table_name
        WHERE {where}
        ORDER BY t.table_schema, t.table_name
    """
    try:
        return execute_sql(q)
    except HTTPException:
        if kb_only:
            return []
        q_simple = f"""
            SELECT table_catalog, table_schema, table_name, table_type,
                   false as is_profiled
            FROM system.information_schema.tables
            WHERE {where.replace('t.', '')}
            ORDER BY table_schema, table_name
        """
        return execute_sql(q_simple)


@app.get("/api/coverage/holistic")
def get_coverage_holistic(catalog: Optional[str] = None):
    """Single endpoint returning all metadata-type coverage counts."""
    cat = catalog or CATALOG
    result = {
        "total_tables": 0, "profiled": 0, "with_comments": 0,
        "with_pii": 0, "with_domain": 0, "with_ontology": 0,
        "with_fk": 0, "metric_views": 0, "metric_view_statuses": {},
        "vs_documents": 0, "vs_by_type": {},
        "avg_confidence": None, "entity_type_count": 0, "fk_count": 0,
    }
    _ALL_TYPES = "('MANAGED','EXTERNAL','VIEW','STREAMING_TABLE','MATERIALIZED_VIEW','FOREIGN')"
    try:
        rows = execute_sql(f"""
            SELECT COUNT(*) as total_tables,
                   COUNT(kb.table_name) as profiled,
                   SUM(CASE WHEN kb.comment IS NOT NULL AND kb.comment != '' THEN 1 ELSE 0 END) as with_comments,
                   SUM(CASE WHEN kb.has_pii = true OR kb.has_phi = true THEN 1 ELSE 0 END) as with_pii,
                   SUM(CASE WHEN kb.domain IS NOT NULL AND kb.domain != '' THEN 1 ELSE 0 END) as with_domain
            FROM (
                SELECT DISTINCT table_catalog, table_schema, table_name
                FROM system.information_schema.tables
                WHERE table_catalog = '{cat}'
                  AND table_schema NOT IN ('information_schema','__internal')
                  AND table_type IN {_ALL_TYPES}
                  AND NOT table_name RLIKE '^(__|event_log_[0-9a-f]{{8}}_)'
            ) t
            LEFT JOIN {fq('table_knowledge_base')} kb
              ON CONCAT(t.table_catalog, '.', t.table_schema, '.', t.table_name) = kb.table_name
        """)
        if rows:
            r = rows[0]
            result["total_tables"] = int(r.get("total_tables") or 0)
            result["profiled"] = int(r.get("profiled") or 0)
            result["with_comments"] = int(r.get("with_comments") or 0)
            result["with_pii"] = int(r.get("with_pii") or 0)
            result["with_domain"] = int(r.get("with_domain") or 0)
    except Exception as e:
        logger.warning("holistic: main coverage query failed: %s", e)
    cat_like = f"{cat}.%"
    try:
        onto = execute_sql(f"""
            SELECT COUNT(DISTINCT entity_type) as type_cnt,
                   AVG(confidence) as avg_conf,
                   COUNT(DISTINCT t.tbl) as tbl_cnt
            FROM {fq('ontology_entities')}
            LATERAL VIEW EXPLODE(source_tables) t AS tbl
            WHERE t.tbl LIKE '{cat_like}'
        """)
        if onto:
            result["entity_type_count"] = int(onto[0].get("type_cnt") or 0)
            result["avg_confidence"] = round(float(onto[0].get("avg_conf") or 0), 3) if onto[0].get("avg_conf") else None
            result["with_ontology"] = int(onto[0].get("tbl_cnt") or 0)
    except Exception as e:
        logger.warning("holistic: ontology query failed: %s", e)
    try:
        fks = execute_sql(f"SELECT COUNT(*) as cnt FROM {fq('fk_predictions')} WHERE final_confidence >= 0.5 AND (src_table LIKE '{cat_like}' OR dst_table LIKE '{cat_like}')")
        result["fk_count"] = int(fks[0]["cnt"]) if fks else 0
    except Exception as e:
        logger.warning("holistic: fk_count query failed: %s", e)
    try:
        fk_tbls = execute_sql(f"""SELECT COUNT(DISTINCT t) as cnt FROM (
            SELECT src_table AS t FROM {fq('fk_predictions')} WHERE final_confidence >= 0.5 AND src_table LIKE '{cat_like}'
            UNION
            SELECT dst_table AS t FROM {fq('fk_predictions')} WHERE final_confidence >= 0.5 AND dst_table LIKE '{cat_like}'
        )""")
        result["with_fk"] = int(fk_tbls[0]["cnt"]) if fk_tbls else 0
    except Exception as e:
        logger.warning("holistic: fk_tables query failed: %s", e)
    try:
        mvs = execute_sql(f"SELECT status, COUNT(*) as cnt FROM {fq('metric_view_definitions')} WHERE source_table LIKE '{cat_like}' GROUP BY status")
        result["metric_view_statuses"] = {r["status"]: int(r["cnt"]) for r in mvs} if mvs else {}
        result["metric_views"] = sum(result["metric_view_statuses"].values())
    except Exception as e:
        logger.warning("holistic: metric_views query failed: %s", e)
    try:
        docs = execute_sql(f"SELECT doc_type, COUNT(*) AS cnt FROM {fq('metadata_documents')} WHERE table_name LIKE '{cat_like}' GROUP BY doc_type")
        result["vs_by_type"] = {r["doc_type"]: int(r["cnt"]) for r in docs} if docs else {}
        result["vs_documents"] = sum(result["vs_by_type"].values())
    except Exception as e:
        logger.warning("holistic: vs_documents query failed: %s", e)
    return result


@app.get("/api/coverage/review-summary")
def get_coverage_review_summary(catalog: Optional[str] = None):
    """Count tables by review_status in table_knowledge_base."""
    cat = catalog or CATALOG
    try:
        rows = execute_sql(f"""
            SELECT COALESCE(review_status, 'unreviewed') AS status, COUNT(*) AS cnt
            FROM {fq('table_knowledge_base')}
            WHERE table_name LIKE '{cat}.%'
            GROUP BY 1
        """)
        return rows or []
    except Exception:
        return []


# ---------------------------------------------------------------------------
# GraphRAG endpoint (delegates to agent)
# ---------------------------------------------------------------------------


@app.post("/api/graph/query")
async def graph_rag_query(req: GraphQueryRequest):
    """Answer a natural-language question by traversing the knowledge graph.

    Delegates to the deterministic GraphRAG pipeline in deep_analysis.
    """
    try:
        from agent.deep_analysis import run_deep_analysis
    except ImportError as e:
        raise HTTPException(503, detail=f"Agent not available: {e}")
    try:
        result = run_deep_analysis(req.question, mode="graphrag")
        return result
    except Exception as exc:
        msg = str(exc)
        if "REQUEST_LIMIT_EXCEEDED" in msg or "429" in msg or "RateLimitError" in msg:
            raise HTTPException(
                429,
                detail="Model rate limit exceeded. Try again in a minute or switch to a different model.",
            ) from exc
        logger.error("GraphRAG agent error: %s", exc)
        raise HTTPException(500, detail=f"Agent error: {msg}") from exc


# ---------------------------------------------------------------------------
# Graph Explorer endpoints
# ---------------------------------------------------------------------------


@app.get("/api/graph/traverse")
def graph_traverse_endpoint(
    start_node: str,
    max_hops: int = 2,
    direction: str = "both",
    relationship: Optional[str] = None,
    edge_type: Optional[str] = None,
    hide_contains: bool = True,
):
    """BFS traversal with optional progressive disclosure (collapse column edges)."""
    result = multi_hop_traverse(
        start_node=start_node,
        max_hops=min(max_hops, 4),
        relationship=relationship,
        edge_type=edge_type,
        direction=direction,
    )
    if hide_contains:
        contains_count: dict[str, int] = {}
        other_edges = []
        for e in result["edges"]:
            if e.get("relationship") == "contains":
                contains_count[e["src"]] = contains_count.get(e["src"], 0) + 1
            else:
                other_edges.append(e)
        result["edges"] = other_edges
        result["collapsed_columns"] = contains_count
    return result


@app.get("/api/graph/nodes")
def graph_nodes_endpoint(
    node_type: Optional[str] = None,
    domain: Optional[str] = None,
    search: Optional[str] = None,
    limit: int = 100,
):
    """Search graph nodes for the explorer table picker."""
    conditions = []
    if node_type:
        conditions.append(f"node_type = {_safe_sql_str(node_type)}")
    if domain:
        conditions.append(f"domain = {_safe_sql_str(domain)}")
    if search:
        safe_search = search.replace("'", "''").replace("%", "\\%")
        conditions.append(f"(id LIKE '%{safe_search}%' OR display_name LIKE '%{safe_search}%')")
    where = "WHERE " + " AND ".join(conditions) if conditions else ""
    return graph_query(
        f"SELECT id, node_type, domain, display_name, short_description, sensitivity "
        f"FROM public.graph_nodes {where} ORDER BY id LIMIT {limit}"
    )


# ---------------------------------------------------------------------------
# Catalog / Schema / Table discovery (cascading selectors)
# ---------------------------------------------------------------------------


@app.get("/api/catalogs")
def list_catalogs():
    try:
        q = (
            "SELECT catalog_name FROM system.information_schema.catalogs "
            "WHERE catalog_name NOT IN ('system', '__databricks_internal') "
            "ORDER BY catalog_name"
        )
        rows = execute_sql(q)
        return [r["catalog_name"] for r in rows]
    except Exception as e:
        logger.warning("list_catalogs failed (permissions?): %s", e)
        raise HTTPException(status_code=403, detail=f"Cannot list catalogs: {e}")


@app.get("/api/schemas")
def list_schemas(catalog: str):
    try:
        q = (
            f"SELECT schema_name FROM `{catalog}`.information_schema.schemata "
            f"WHERE schema_name NOT IN ('information_schema', '__internal') "
            f"ORDER BY schema_name"
        )
        rows = execute_sql(q)
        return [r["schema_name"] for r in rows]
    except Exception as e:
        logger.warning("list_schemas(%s) failed: %s", catalog, e)
        raise HTTPException(status_code=403, detail=f"Cannot list schemas in {catalog}: {e}")


@app.get("/api/tables")
def list_tables(catalog: str, schema: str):
    try:
        q = (
            f"SELECT table_name FROM `{catalog}`.information_schema.tables "
            f"WHERE table_schema = '{schema}' "
            f"AND table_type IN ('MANAGED','EXTERNAL','VIEW','STREAMING_TABLE','MATERIALIZED_VIEW','FOREIGN') "
            f"AND NOT table_name RLIKE '^(__|event_log_[0-9a-f]{{8}}_)' "
            f"ORDER BY table_name"
        )
        rows = execute_sql(q)
        return [r["table_name"] for r in rows]
    except Exception as e:
        logger.warning("list_tables(%s.%s) failed: %s", catalog, schema, e)
        raise HTTPException(status_code=403, detail=f"Cannot list tables in {catalog}.{schema}: {e}")


# ---------------------------------------------------------------------------
# Semantic Layer endpoints
# ---------------------------------------------------------------------------

_sl_tables_ensured = False


def _ensure_semantic_layer_tables():
    global _sl_tables_ensured
    if _sl_tables_ensured:
        return
    _TABLE_DDLS = [
        f"""CREATE TABLE IF NOT EXISTS {fq('semantic_layer_questions')} (
            question_id STRING NOT NULL, question_text STRING, status STRING,
            created_at TIMESTAMP, processed_at TIMESTAMP
        ) COMMENT 'Business questions for semantic layer generation'""",
        f"""CREATE TABLE IF NOT EXISTS {fq('metric_view_definitions')} (
            definition_id STRING NOT NULL, metric_view_name STRING, source_table STRING,
            json_definition STRING, source_questions STRING, status STRING,
            validation_errors STRING, genie_space_id STRING, created_at TIMESTAMP,
            applied_at TIMESTAMP, version INT, parent_definition_id STRING,
            project_id STRING
        ) COMMENT 'Generated metric view definitions with version history'""",
        f"""CREATE TABLE IF NOT EXISTS {fq('semantic_layer_profiles')} (
            profile_id STRING NOT NULL, profile_name STRING, questions STRING,
            table_patterns STRING, created_at TIMESTAMP, updated_at TIMESTAMP,
            business_context STRING
        ) COMMENT 'Named question profiles for semantic layer'""",
        f"""CREATE TABLE IF NOT EXISTS {fq('semantic_layer_projects')} (
            project_id STRING NOT NULL, project_name STRING, description STRING,
            created_at TIMESTAMP, selected_tables STRING
        ) COMMENT 'Named projects for grouping metric view definitions'""",
    ]
    for ddl in _TABLE_DDLS:
        table_name = ddl.split("IF NOT EXISTS")[-1].split("(")[0].strip() if "IF NOT EXISTS" in ddl else "unknown"
        logger.info("Ensuring semantic layer table: %s", table_name)
        try:
            execute_sql(ddl)
        except Exception as e:
            logger.error("Failed to create table %s: %s", table_name, e)
            raise
    for col_ddl in [
        "version INT, parent_definition_id STRING",
        "project_id STRING",
        "complexity_score INT, complexity_level STRING",
        "deployed_catalog STRING, deployed_schema STRING",
    ]:
        try:
            execute_sql(f"ALTER TABLE {fq('metric_view_definitions')} ADD COLUMNS ({col_ddl})")
        except Exception:
            pass
    try:
        execute_sql(f"ALTER TABLE {fq('semantic_layer_projects')} ADD COLUMNS (selected_tables STRING)")
    except Exception:
        pass
    try:
        execute_sql(f"ALTER TABLE {fq('semantic_layer_profiles')} ADD COLUMNS (business_context STRING)")
    except Exception:
        pass
    _sl_tables_ensured = True
    logger.info("Semantic layer tables ensured")


@app.get("/api/semantic-layer/questions")
def list_semantic_questions():
    _ensure_semantic_layer_tables()
    q = f"SELECT question_id, question_text, status, created_at, processed_at FROM {fq('semantic_layer_questions')} ORDER BY created_at DESC"
    try:
        return execute_sql(q)
    except HTTPException as e:
        if e.status_code == 404:
            return []
        raise


@app.post("/api/semantic-layer/questions")
def add_semantic_questions(req: SemanticLayerQuestionsRequest):
    _ensure_semantic_layer_tables()
    from datetime import datetime as _dt

    now = _dt.utcnow().isoformat()
    rows = []
    for q_text in req.questions:
        q_text = q_text.strip()
        if not q_text:
            continue
        qid = str(_uuid.uuid4())
        escaped = q_text.replace("'", "''")
        rows.append(f"('{qid}', '{escaped}', 'pending', '{now}', NULL)")
    if not rows:
        raise HTTPException(400, detail="No valid questions provided")
    values = ", ".join(rows)
    try:
        execute_sql(f"INSERT INTO {fq('semantic_layer_questions')} VALUES {values}")
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Failed to insert semantic layer questions: %s", e)
        raise HTTPException(500, detail=f"Failed to save questions: {e}")
    return {"added": len(rows)}


@app.get("/api/semantic-layer/definitions")
def list_semantic_definitions(project_id: Optional[str] = None):
    _ensure_semantic_layer_tables()
    where = "WHERE status != 'superseded'"
    if project_id:
        where += f" AND project_id = '{project_id}'"
    q = (
        f"SELECT definition_id, metric_view_name, source_table, status, "
        f"validation_errors, genie_space_id, created_at, applied_at, "
        f"COALESCE(version, 1) as version, parent_definition_id, project_id, "
        f"complexity_score, complexity_level, deployed_catalog, deployed_schema "
        f"FROM ("
        f"  SELECT *, ROW_NUMBER() OVER ("
        f"    PARTITION BY metric_view_name, source_table "
        f"    ORDER BY CASE status "
        f"      WHEN 'applied' THEN 0 WHEN 'validated' THEN 1 "
        f"      WHEN 'created' THEN 2 ELSE 3 END, "
        f"    created_at DESC"
        f"  ) AS _rn "
        f"  FROM {fq('metric_view_definitions')} "
        f"  {where}"
        f") WHERE _rn = 1 "
        f"ORDER BY created_at DESC"
    )
    try:
        return execute_sql(q)
    except HTTPException as e:
        if e.status_code == 404:
            return []
        raise


@app.get("/api/semantic-layer/definitions/{definition_id}/json")
def get_semantic_definition_json(definition_id: str):
    q = (
        f"SELECT json_definition FROM {fq('metric_view_definitions')} "
        f"WHERE definition_id = '{definition_id}'"
    )
    rows = execute_sql(q)
    if not rows:
        raise HTTPException(404, detail="Definition not found")
    return {"json_definition": rows[0].get("json_definition", "")}


@app.delete("/api/semantic-layer/definitions/{definition_id}")
def delete_semantic_definition(
    definition_id: str,
    drop_view: bool = False,
    catalog: Optional[str] = None,
    schema: Optional[str] = None,
):
    _ensure_semantic_layer_tables()
    if drop_view and catalog and schema:
        rows = execute_sql(
            f"SELECT metric_view_name, json_definition, status FROM {fq('metric_view_definitions')} "
            f"WHERE definition_id = '{definition_id}'"
        )
        if rows and rows[0].get("status") == "applied":
            defn = rows[0].get("json_definition", "{}")
            if isinstance(defn, str):
                defn = json.loads(defn) if defn.strip() else {}
            mv_name = defn.get("name") or rows[0].get("metric_view_name", "")
            if mv_name:
                try:
                    execute_sql(f"DROP VIEW IF EXISTS `{catalog}`.`{schema}`.`{mv_name}`")
                except Exception:
                    pass
    execute_sql(
        f"DELETE FROM {fq('metric_view_definitions')} WHERE definition_id = '{definition_id}'"
    )
    return {"deleted": True, "definition_id": definition_id}


# --- Profiles ---


@app.get("/api/semantic-layer/profiles")
def list_profiles():
    _ensure_semantic_layer_tables()
    q = f"SELECT profile_id, profile_name, questions, table_patterns, created_at, updated_at, business_context FROM {fq('semantic_layer_profiles')} ORDER BY updated_at DESC"
    try:
        return execute_sql(q)
    except HTTPException as e:
        if e.status_code == 404:
            return []
        raise


@app.post("/api/semantic-layer/profiles")
def save_profile(req: SemanticProfileRequest):
    _ensure_semantic_layer_tables()
    from datetime import datetime as _dt

    now = _dt.utcnow().isoformat()
    qs_json = json.dumps(req.questions).replace("'", "''")
    tp_json = json.dumps(req.table_patterns).replace("'", "''")
    name_esc = req.profile_name.replace("'", "''")
    biz_ctx_esc = (req.business_context or "").replace("'", "''")

    existing = execute_sql(
        f"SELECT profile_id FROM {fq('semantic_layer_profiles')} WHERE profile_name = '{name_esc}'"
    )
    if existing:
        pid = existing[0]["profile_id"]
        execute_sql(
            f"UPDATE {fq('semantic_layer_profiles')} "
            f"SET questions = '{qs_json}', table_patterns = '{tp_json}', "
            f"business_context = '{biz_ctx_esc}', updated_at = '{now}' "
            f"WHERE profile_id = '{pid}'"
        )
        return {"profile_id": pid, "updated": True}
    pid = str(_uuid.uuid4())
    execute_sql(
        f"INSERT INTO {fq('semantic_layer_profiles')} "
        f"(profile_id, profile_name, questions, table_patterns, created_at, updated_at, business_context) "
        f"VALUES ('{pid}', '{name_esc}', '{qs_json}', '{tp_json}', '{now}', '{now}', '{biz_ctx_esc}')"
    )
    return {"profile_id": pid, "updated": False}


@app.delete("/api/semantic-layer/profiles/{profile_id}")
def delete_profile(profile_id: str):
    _ensure_semantic_layer_tables()
    pid_esc = profile_id.replace("'", "''")
    execute_sql(
        f"DELETE FROM {fq('semantic_layer_profiles')} WHERE profile_id = '{pid_esc}'"
    )
    return {"deleted": True}


# --- Projects ---


@app.get("/api/semantic-layer/projects")
def list_projects():
    _ensure_semantic_layer_tables()
    try:
        return execute_sql(
            f"SELECT project_id, project_name, description, created_at, selected_tables "
            f"FROM {fq('semantic_layer_projects')} ORDER BY created_at DESC"
        )
    except HTTPException as e:
        if e.status_code == 404:
            return []
        raise


@app.get("/api/semantic/metric-views")
def list_metric_views(status: Optional[str] = None):
    """List metric views from the definitions table.

    Defaults to status='created'. Pass status='all' to return every non-superseded
    metric view (latest version only), or a comma-separated list like
    'applied,validated,created'.
    """
    _ensure_semantic_layer_tables()
    if status and status.lower() == "all":
        status_filter = "status != 'superseded'"
    elif status:
        vals = ", ".join(f"'{s.strip()}'" for s in status.split(",") if s.strip())
        status_filter = f"status IN ({vals})" if vals else "status = 'created'"
    else:
        status_filter = "status = 'created'"
    q = (
        f"SELECT definition_id, metric_view_name, source_table, status, "
        f"genie_space_id, created_at, deployed_catalog, deployed_schema "
        f"FROM ("
        f"  SELECT *, ROW_NUMBER() OVER ("
        f"    PARTITION BY metric_view_name, source_table "
        f"    ORDER BY CASE status "
        f"      WHEN 'applied' THEN 0 WHEN 'validated' THEN 1 "
        f"      WHEN 'created' THEN 2 ELSE 3 END, "
        f"    created_at DESC"
        f"  ) AS _rn "
        f"  FROM {fq('metric_view_definitions')} "
        f"  WHERE {status_filter}"
        f") WHERE _rn = 1 "
        f"ORDER BY source_table, metric_view_name"
    )
    try:
        return execute_sql(q)
    except HTTPException as e:
        if e.status_code == 404:
            return []
        raise


@app.post("/api/semantic-layer/projects")
def create_project(req: SemanticProjectRequest):
    _ensure_semantic_layer_tables()
    from datetime import datetime as _dt

    pid = str(_uuid.uuid4())
    now = _dt.utcnow().isoformat()
    name_esc = req.project_name.replace("'", "''")
    desc_esc = req.description.replace("'", "''")
    execute_sql(
        f"INSERT INTO {fq('semantic_layer_projects')} VALUES "
        f"('{pid}', '{name_esc}', '{desc_esc}', '{now}', NULL)"
    )
    return {"project_id": pid, "project_name": req.project_name}


@app.delete("/api/semantic-layer/projects/{project_id}")
def delete_project(project_id: str):
    _ensure_semantic_layer_tables()
    execute_sql(
        f"DELETE FROM {fq('semantic_layer_projects')} WHERE project_id = '{project_id}'"
    )
    execute_sql(
        f"UPDATE {fq('metric_view_definitions')} SET project_id = NULL "
        f"WHERE project_id = '{project_id}'"
    )
    return {"deleted": True}


class ProjectTablesUpdate(BaseModel):
    selected_tables: list[str]


@app.patch("/api/semantic-layer/projects/{project_id}/tables")
def update_project_tables(project_id: str, req: ProjectTablesUpdate):
    _ensure_semantic_layer_tables()
    tables_json = json.dumps(req.selected_tables).replace("'", "''")
    execute_sql(
        f"UPDATE {fq('semantic_layer_projects')} SET selected_tables = '{tables_json}' "
        f"WHERE project_id = '{project_id}'"
    )
    return {"project_id": project_id, "selected_tables": req.selected_tables}


# --- In-app metric view generation ---

_sl_tasks: dict[str, dict] = {}


def _sl_vs_enrich(questions: list[str], selected_tables: set[str]) -> str:
    """Phase 1a: Vector Search per question to discover relevant tables/columns."""
    try:
        from agent.metadata_tools import _get_vs_index, VS_INDEX_SUFFIX
        vs_index_name = f"{CATALOG}.{SCHEMA}.{VS_INDEX_SUFFIX}"
        index = _get_vs_index(vs_index_name)
    except Exception:
        return ""

    seen = set()
    lines: list[str] = []
    for q in questions[:8]:
        try:
            results = index.similarity_search(
                query_text=q,
                columns=["doc_type", "content", "table_name", "entity_type"],
                num_results=5,
                query_type="HYBRID",
            )
            cols = results.get("manifest", {}).get("columns", [])
            col_names = [c.get("name", f"col{i}") for i, c in enumerate(cols)]
            for row in results.get("result", {}).get("data_array", []):
                match = dict(zip(col_names, row)) if col_names else {}
                tname = match.get("table_name", "")
                doc_type = match.get("doc_type", "")
                content = (match.get("content") or "")[:200]
                key = f"{doc_type}:{tname}:{content[:60]}"
                if key in seen or not content:
                    continue
                seen.add(key)
                is_new = tname and tname not in selected_tables
                tag = " [NOT SELECTED - consider adding]" if is_new else ""
                lines.append(f"  [{doc_type}] {tname}{tag}: {content}")
        except Exception:
            continue

    if not lines:
        return ""
    return "\nSEMANTIC SEARCH DISCOVERIES (relevant to business questions):\n" + "\n".join(lines[:25])


def _sl_graph_enrich(fq_tables: list[str]) -> str:
    """Phase 1b: 1-2 hop graph traversal from selected tables."""
    edges: list[str] = []
    for tname in fq_tables[:10]:
        tname_esc = tname.replace("'", "''")
        try:
            rows = graph_query(
                f"SELECT e.src, e.dst, e.relationship, e.edge_type, e.weight, e.join_expression "
                f"FROM public.graph_edges e "
                f"WHERE (e.src = '{tname_esc}' OR e.dst = '{tname_esc}') "
                f"AND e.edge_type IN ('references','contains','instance_of','same_domain','derives_from') "
                f"LIMIT 20"
            )
            for r in rows:
                expr = r.get("join_expression") or ""
                expr_str = f" JOIN: {expr}" if expr else ""
                line = f"  {r['src']} --[{r.get('relationship', r.get('edge_type', ''))}]--> {r['dst']}{expr_str}"
                if line not in edges:
                    edges.append(line)
        except Exception:
            continue

    # 2-hop: find paths through intermediate nodes
    if edges and len(fq_tables) > 1:
        table_set = set(fq_tables)
        try:
            in_clause = ", ".join(f"'{t.replace(chr(39), chr(39)+chr(39))}'" for t in fq_tables)
            hop2 = graph_query(
                f"SELECT DISTINCT e1.src as t1, e1.dst as mid, e2.dst as t2, "
                f"e1.relationship as r1, e2.relationship as r2, e2.join_expression "
                f"FROM public.graph_edges e1 "
                f"JOIN public.graph_edges e2 ON e1.dst = e2.src "
                f"WHERE e1.src IN ({in_clause}) AND e2.dst IN ({in_clause}) "
                f"AND e1.src != e2.dst "
                f"AND e1.edge_type IN ('references','contains','instance_of') "
                f"AND e2.edge_type IN ('references','contains','instance_of') "
                f"LIMIT 10"
            )
            for h in hop2:
                line = f"  {h['t1']} --[{h['r1']}]--> {h['mid']} --[{h['r2']}]--> {h['t2']}"
                if line not in edges:
                    edges.append(line)
        except Exception:
            pass

    if not edges:
        return ""
    return "\nGRAPH RELATIONSHIPS (structural join paths and entity connections):\n" + "\n".join(edges[:30])


def _sl_extra_sql_context(in_clause: str) -> str:
    """Phase 1c: ontology_relationships, column_properties, existing MVs, profiling."""
    parts: list[str] = []

    # Ontology relationships
    try:
        rel_rows = execute_sql(
            f"SELECT source_entity, target_entity, relationship_type, description "
            f"FROM {fq('ontology_relationships')} LIMIT 50"
        )
        if rel_rows:
            parts.append("\nONTOLOGY ENTITY RELATIONSHIPS:")
            for r in rel_rows:
                desc = f" ({r['description']})" if r.get("description") else ""
                parts.append(f"  {r['source_entity']} --[{r['relationship_type']}]--> {r['target_entity']}{desc}")
    except Exception:
        pass

    # Column properties
    try:
        cp_rows = execute_sql(
            f"SELECT table_name, column_name, property_name, property_value "
            f"FROM {fq('ontology_column_properties')} WHERE table_name IN ({in_clause}) LIMIT 100"
        )
        if cp_rows:
            parts.append("\nCOLUMN PROPERTY ANNOTATIONS:")
            by_col: dict[str, list[str]] = {}
            for cp in cp_rows:
                key = f"{cp['table_name']}.{cp['column_name']}"
                by_col.setdefault(key, []).append(f"{cp['property_name']}={cp['property_value']}")
            for col_key, props in list(by_col.items())[:40]:
                parts.append(f"  {col_key}: {', '.join(props)}")
    except Exception:
        pass

    # Existing metric view definitions (for deduplication)
    try:
        mv_rows = execute_sql(
            f"SELECT metric_view_name, source_table, status "
            f"FROM {fq('metric_view_definitions')} "
            f"WHERE status NOT IN ('superseded', 'deleted') LIMIT 30"
        )
        if mv_rows:
            parts.append("\nEXISTING METRIC VIEWS (avoid duplicating these):")
            for mv in mv_rows:
                parts.append(f"  {mv['metric_view_name']} (source: {mv['source_table']}, status: {mv['status']})")
    except Exception:
        pass

    # Profiling summaries
    try:
        prof_rows = execute_sql(
            f"SELECT table_name, column_name, distinct_count, null_count "
            f"FROM {fq('profiling_results')} WHERE table_name IN ({in_clause}) "
            f"AND (distinct_count IS NOT NULL OR null_count IS NOT NULL) LIMIT 100"
        )
        if prof_rows:
            parts.append("\nPROFILING SUMMARIES (cardinality/nulls -- use for dimension vs measure decisions):")
            for p in prof_rows:
                dc = p.get("distinct_count", "?")
                nc = p.get("null_count", "?")
                parts.append(f"  {p['table_name']}.{p['column_name']}: distinct={dc}, nulls={nc}")
    except Exception:
        pass

    return "\n".join(parts)


def _build_sl_context(
    tables: list[str], cat: str, sch: str, questions: list[str] | None = None,
    business_context: str | None = None, profile_id: str | None = None,
) -> str:
    """Build enriched context from KB tables, Vector Search, graph, and ontology. Cached 120s."""
    q_key = ",".join(sorted(questions)) if questions else ""
    cache_key = f"{cat}.{sch}:" + ",".join(sorted(tables)) + ":" + q_key + ":" + (profile_id or "")
    with _sl_context_lock:
        if cache_key in _sl_context_cache:
            return _sl_context_cache[cache_key]

    fq_tables = []
    for t in tables:
        if "." in t:
            fq_tables.append(t)
        else:
            fq_tables.append(f"{cat}.{sch}.{t}")
    in_clause = ", ".join(f"'{t}'" for t in fq_tables)
    selected_set = set(fq_tables)

    # --- Core SQL context (original) ---
    parts: list[str] = []
    if business_context and business_context.strip():
        parts.append(
            f"BUSINESS CONTEXT (provided by the user -- this defines the semantic frame for all analysis):\n{business_context.strip()}"
        )
    table_rows = execute_sql(
        f"SELECT table_name, comment, domain, subdomain, has_pii, has_phi FROM {fq('table_knowledge_base')} "
        f"WHERE table_name IN ({in_clause})"
    )
    col_rows = execute_sql(
        f"SELECT table_name, column_name, data_type, comment, classification "
        f"FROM {fq('column_knowledge_base')} WHERE table_name IN ({in_clause})"
    )
    col_by_table: dict[str, list] = {}
    for c in col_rows:
        col_by_table.setdefault(c["table_name"], []).append(c)

    fk_rows = []
    try:
        fk_rows = execute_sql(
            f"SELECT src_table, dst_table, src_column, dst_column, final_confidence "
            f"FROM {fq('fk_predictions')} WHERE is_fk = 'true' AND final_confidence >= 0.85 "
            f"AND (src_table IN ({in_clause}) OR dst_table IN ({in_clause}))"
        )
    except HTTPException:
        pass

    ont_rows = []
    try:
        ont_rows = execute_sql(
            f"SELECT entity_type, source_tables, description FROM {fq('ontology_entities')} WHERE confidence >= 0.4"
        )
    except HTTPException:
        pass
    entity_map: dict[str, dict] = {}
    for o in ont_rows:
        src_tables = o.get("source_tables") or ""
        if isinstance(src_tables, str):
            try:
                src_tables = json.loads(src_tables)
            except (json.JSONDecodeError, TypeError):
                src_tables = [src_tables] if src_tables else []
        for t in src_tables:
            entity_map[t] = {"type": o["entity_type"], "description": o.get("description", "")}

    # Ontology relationships for cross-entity context
    ont_rels = []
    try:
        ont_rels = execute_sql(
            f"SELECT src_entity_type, dst_entity_type, relationship_name, cardinality "
            f"FROM {fq('ontology_relationships')} WHERE confidence >= 0.4 LIMIT 50"
        )
    except (HTTPException, Exception):
        pass

    pii_tables = {t["table_name"] for t in table_rows if t.get("has_pii")}
    phi_tables = {t["table_name"] for t in table_rows if t.get("has_phi")}

    for t in table_rows:
        tname = t["table_name"]
        ent_info = entity_map.get(tname, {})
        ent_type = ent_info.get("type", "")
        ent_desc = ent_info.get("description", "")
        ent_str = f" Entity: {ent_type}" if ent_type else ""
        if ent_desc:
            ent_str += f" -- {ent_desc}"
        pii_tag = ""
        if tname in pii_tables:
            pii_tag += " [PII]"
        if tname in phi_tables:
            pii_tag += " [PHI]"
        line = f"Table: {tname} (Comment: \"{t.get('comment', '')}\" Domain: {t.get('domain', '')} / {t.get('subdomain', '')}){ent_str}{pii_tag}"
        cols = col_by_table.get(tname, [])
        if len(cols) > 80:
            fk_cols = {fk["src_column"] for fk in fk_rows if fk["src_table"] == tname}
            fk_cols |= {fk["dst_column"] for fk in fk_rows if fk["dst_table"] == tname}
            prioritized = sorted(cols, key=lambda c: (
                c["column_name"] not in fk_cols,
                not bool(c.get("comment")),
                c.get("column_name", ""),
            ))
            cols = prioritized[:80]
        is_pii_table = tname in pii_tables or tname in phi_tables
        col_strs = []
        for c in cols:
            cls = (c.get("classification") or "").lower()
            col_tag = ""
            if is_pii_table and any(k in cls for k in ("pii", "phi", "personal", "sensitive", "name", "email", "ssn", "phone", "address", "dob")):
                col_tag = " [PII]"
            col_strs.append(f"  - {c['column_name']} {c.get('data_type', '')} : {c.get('comment', '')}{col_tag}")
        parts.append(
            line + "\n  Columns:\n" + "\n".join(col_strs) if col_strs else line
        )

    if fk_rows:
        parts.append("\nFOREIGN KEY RELATIONSHIPS:")
        for fk in fk_rows:
            parts.append(
                f"  {fk['src_table']}.{fk['src_column']} -> {fk['dst_table']}.{fk['dst_column']} (confidence {fk['final_confidence']})"
            )

    if ont_rels:
        parts.append("\nENTITY RELATIONSHIPS (use for cross-entity join and metric design):")
        for r in ont_rels:
            card = r.get("cardinality", "")
            card_str = f" ({card})" if card else ""
            parts.append(f"  {r.get('src_entity_type', '')} --{r.get('relationship_name', '')}--> {r.get('dst_entity_type', '')}{card_str}")

    # Ontology metric suggestions
    metric_rows = []
    try:
        metric_rows = execute_sql(
            f"SELECT metric_name, description, entity_id, aggregation_type, source_field, filter_condition "
            f"FROM {fq('ontology_metrics')}"
        )
    except (HTTPException, Exception):
        pass
    if metric_rows:
        parts.append(
            "\nONTOLOGY METRIC SUGGESTIONS (use as hints for measures/dimensions):"
        )
        for m in metric_rows:
            line = f"  - {m.get('metric_name', '')}: {m.get('description', '')}"
            if m.get("aggregation_type") and m.get("source_field"):
                line += f"  -> {m['aggregation_type']}({m['source_field']})"
            if m.get("filter_condition"):
                line += f" WHERE {m['filter_condition']}"
            parts.append(line)

    # Fallback to information_schema when KB is empty -- supports multi-schema
    if not table_rows:
        tables_by_location: dict[tuple[str, str], list[str]] = {}
        for t in fq_tables:
            t_parts = t.split(".")
            if len(t_parts) == 3:
                tables_by_location.setdefault((t_parts[0], t_parts[1]), []).append(t_parts[2])
            else:
                tables_by_location.setdefault((cat, sch), []).append(t_parts[-1])
        for (t_cat, t_sch), t_names in tables_by_location.items():
            short_clause = ", ".join(f"'{t}'" for t in t_names)
            info_cols = execute_sql(
                f"SELECT table_name, column_name, data_type "
                f"FROM `{t_cat}`.information_schema.columns "
                f"WHERE table_schema = '{t_sch}' AND table_name IN ({short_clause})"
            )
            col_by_tbl: dict[str, list] = {}
            for c in info_cols:
                col_by_tbl.setdefault(c["table_name"], []).append(c)
            for tname, cols in col_by_tbl.items():
                col_strs = [
                    f"  - {c['column_name']} {c.get('data_type', '')}" for c in cols
                ]
                parts.append(
                    f"Table: {t_cat}.{t_sch}.{tname}\n  Columns:\n" + "\n".join(col_strs)
                )

    # --- Phase 1 enrichment: VS, Graph, Extended SQL (parallel) ---
    enrichment_parts: list[str] = []
    with ThreadPoolExecutor(max_workers=3) as pool:
        futures = {}
        if questions:
            futures["vs"] = pool.submit(_sl_vs_enrich, questions, selected_set)
        futures["graph"] = pool.submit(_sl_graph_enrich, fq_tables)
        futures["sql_ext"] = pool.submit(_sl_extra_sql_context, in_clause)

        for key, fut in futures.items():
            try:
                result_str = fut.result(timeout=30)
                if result_str:
                    enrichment_parts.append(result_str)
            except Exception as exc:
                logger.warning("SL context enrichment '%s' failed: %s", key, exc)

    # KPI library enrichment (skip gracefully if table doesn't exist yet)
    try:
        kpi_where = f" WHERE profile_id = '{profile_id.replace(chr(39), chr(39)*2)}'" if profile_id else ""
        kpi_rows = execute_sql(
            f"SELECT name, description, formula, domain, target_tables, validation_status FROM {fq('kpi_definitions')}{kpi_where}"
        )
        if kpi_rows:
            # Filter KPIs to those whose target_tables overlap with selected tables
            relevant = []
            for k in kpi_rows:
                kt = k.get("target_tables") or []
                if isinstance(kt, str):
                    try:
                        kt = json.loads(kt)
                    except Exception:
                        kt = [kt]
                kt_set = {t.lower() for t in kt} | {t.split(".")[-1].lower() for t in kt}
                if not kt or kt_set & {t.lower() for t in fq_tables} | {t.split(".")[-1].lower() for t in fq_tables}:
                    relevant.append(k)
            if relevant:
                kpi_block = (
                    "\nREQUIRED KPIs -- you MUST implement each as a measure in an appropriate metric view.\n"
                    "If a KPI cannot be implemented with the available columns, skip it and note why in the metric view comment."
                )
                for i, k in enumerate(relevant, 1):
                    v_status = f" [{k.get('validation_status', '')}]" if k.get("validation_status") else ""
                    kpi_block += f"\n  {i}. {k['name']} ({k.get('domain', '')}){v_status} : {k.get('description', '')} | Formula: {k.get('formula', 'N/A')}"
                parts.append(kpi_block)
    except Exception:
        pass

    result = "\n".join(parts) + "\n".join(enrichment_parts)
    with _sl_context_lock:
        _sl_context_cache[cache_key] = result
    return result


_FEW_SHOT_BY_DOMAIN = {
    "sales": """\
INPUT tables:
  sales.orders columns: [order_id BIGINT, customer_id BIGINT, order_date DATE, total_amount DECIMAL(10,2), region STRING, status STRING, is_returned BOOLEAN]
  sales.customers columns: [id BIGINT, name STRING, segment STRING, signup_date DATE]
  FK: orders.customer_id -> customers.id (confidence 0.95)
INPUT questions:
  1. What is total revenue by region?  2. How many orders per month?  3. What is the fulfillment rate by segment?
OUTPUT:
[
  {"name": "order_performance_metrics", "source": "sales.orders",
   "comment": "Order performance including revenue, fulfillment rates, and return analysis",
   "filter": "status IS NOT NULL",
   "dimensions": [
     {"name": "Order Month", "expr": "DATE_TRUNC('MONTH', order_date)", "comment": "Month of order placement"},
     {"name": "Region", "expr": "region", "comment": "Sales region"},
     {"name": "Customer Segment", "expr": "segment", "comment": "Customer segment from joined customers table"},
     {"name": "Customer Tier", "expr": "CASE WHEN segment IN ('Enterprise', 'Strategic') THEN 'Top Tier' WHEN segment = 'Mid-Market' THEN 'Growth' ELSE 'Standard' END", "comment": "Customer tier grouping"}],
   "measures": [
     {"name": "Total Revenue", "expr": "SUM(total_amount)", "comment": "Sum of all order values"},
     {"name": "Avg Order Value", "expr": "AVG(total_amount)", "comment": "Average order amount"},
     {"name": "Revenue per Customer", "expr": "SUM(total_amount) / NULLIF(COUNT(DISTINCT customer_id), 0)", "comment": "Average revenue per unique customer"},
     {"name": "Fulfillment Rate", "expr": "SUM(CASE WHEN status = 'fulfilled' THEN 1 ELSE 0 END) * 1.0 / NULLIF(COUNT(*), 0)", "comment": "Fraction of orders fulfilled"},
     {"name": "Fulfilled Revenue", "expr": "SUM(total_amount) FILTER (WHERE status = 'fulfilled')", "comment": "Revenue from fulfilled orders only"},
     {"name": "30-Day Rolling Avg Revenue", "expr": "AVG(SUM(total_amount))", "window": {"order_by": "order_date", "range": "INTERVAL 30 DAYS PRECEDING"}, "comment": "Rolling 30-day average of daily revenue"}],
   "joins": [{"name": "customers", "source": "sales.customers", "on": "source.customer_id = customers.id"}]}
]""",
    "healthcare": """\
INPUT tables:
  clinical.encounters columns: [encounter_id BIGINT, patient_id BIGINT, provider_id BIGINT, admit_date DATE, discharge_date DATE, encounter_type STRING, department STRING, total_charges DECIMAL(12,2), status STRING]
  clinical.patients columns: [patient_id BIGINT, birth_date DATE, gender STRING, zip_code STRING, insurance_type STRING]
  FK: encounters.patient_id -> patients.patient_id (confidence 0.92)
INPUT questions:
  1. What is the average length of stay by department?  2. What is the readmission rate within 30 days?  3. How does patient volume trend by month?
OUTPUT:
[
  {"name": "encounter_throughput_metrics", "source": "clinical.encounters",
   "comment": "Encounter volume, throughput, and clinical outcome metrics",
   "filter": "status != 'cancelled'",
   "dimensions": [
     {"name": "Admit Month", "expr": "DATE_TRUNC('MONTH', admit_date)", "comment": "Month of admission"},
     {"name": "Department", "expr": "department", "comment": "Clinical department"},
     {"name": "Encounter Type", "expr": "encounter_type", "comment": "Inpatient, outpatient, ED, etc."},
     {"name": "Insurance Type", "expr": "insurance_type", "comment": "Patient insurance from joined patients table"}],
   "measures": [
     {"name": "Encounter Count", "expr": "COUNT(*)", "comment": "Total encounters"},
     {"name": "Unique Patients", "expr": "COUNT(DISTINCT patient_id)", "comment": "Distinct patient count"},
     {"name": "Avg Length of Stay", "expr": "AVG(DATEDIFF(discharge_date, admit_date))", "comment": "Average days from admit to discharge"},
     {"name": "Encounters per Patient", "expr": "COUNT(*) * 1.0 / NULLIF(COUNT(DISTINCT patient_id), 0)", "comment": "Average visits per patient"},
     {"name": "Total Charges", "expr": "SUM(total_charges)", "comment": "Sum of encounter charges"},
     {"name": "Charge per Encounter", "expr": "SUM(total_charges) / NULLIF(COUNT(*), 0)", "comment": "Average charge per encounter"}],
   "joins": [{"name": "patients", "source": "clinical.patients", "on": "source.patient_id = patients.patient_id"}]}
]""",
    "finance": """\
INPUT tables:
  finance.transactions columns: [txn_id BIGINT, account_id BIGINT, txn_date DATE, amount DECIMAL(12,2), txn_type STRING, category STRING, is_fraud BOOLEAN]
  finance.accounts columns: [account_id BIGINT, customer_name STRING, account_type STRING, opened_date DATE, region STRING]
  FK: transactions.account_id -> accounts.account_id (confidence 0.94)
INPUT questions:
  1. What is the total transaction volume by category?  2. What is the fraud rate by account type?  3. How has monthly deposit growth trended?
OUTPUT:
[
  {"name": "transaction_risk_metrics", "source": "finance.transactions",
   "comment": "Transaction volume, fraud rates, and financial flow analysis",
   "dimensions": [
     {"name": "Transaction Month", "expr": "DATE_TRUNC('MONTH', txn_date)", "comment": "Month of transaction"},
     {"name": "Category", "expr": "category", "comment": "Transaction category"},
     {"name": "Account Type", "expr": "account_type", "comment": "Account classification from joined accounts"}],
   "measures": [
     {"name": "Transaction Count", "expr": "COUNT(*)", "comment": "Total transactions"},
     {"name": "Total Amount", "expr": "SUM(amount)", "comment": "Sum of transaction amounts"},
     {"name": "Avg Transaction Size", "expr": "AVG(amount)", "comment": "Average transaction amount"},
     {"name": "Fraud Rate", "expr": "SUM(CASE WHEN is_fraud = TRUE THEN 1 ELSE 0 END) * 1.0 / NULLIF(COUNT(*), 0)", "comment": "Fraction of flagged transactions"},
     {"name": "Deposit Volume", "expr": "SUM(amount) FILTER (WHERE txn_type = 'deposit')", "comment": "Total deposit inflows"}],
   "joins": [{"name": "accounts", "source": "finance.accounts", "on": "source.account_id = accounts.account_id"}]}
]""",
    "project_management": """\
INPUT tables:
  pm.tasks columns: [task_id BIGINT, project_id BIGINT, assignee_id BIGINT, title STRING, status STRING, priority STRING, created_date DATE, due_date DATE, completed_date DATE, estimated_hours DECIMAL(6,1), actual_hours DECIMAL(6,1)]
  pm.projects columns: [project_id BIGINT, name STRING, department STRING, start_date DATE, target_date DATE, status STRING]
  FK: tasks.project_id -> projects.project_id (confidence 0.95)
INPUT questions:
  1. Which projects have the most overdue tasks?  2. What is the on-time completion rate?  3. How does effort estimation accuracy vary by project?
OUTPUT:
[
  {"name": "task_delivery_metrics", "source": "pm.tasks",
   "comment": "Task delivery, effort accuracy, and project health metrics",
   "filter": "status IS NOT NULL",
   "dimensions": [
     {"name": "Created Month", "expr": "DATE_TRUNC('MONTH', created_date)", "comment": "Month task was created"},
     {"name": "Priority", "expr": "priority", "comment": "Task priority level"},
     {"name": "Project Name", "expr": "name", "comment": "Project name from joined projects table"},
     {"name": "Department", "expr": "department", "comment": "Department from joined projects table"}],
   "measures": [
     {"name": "Task Count", "expr": "COUNT(*)", "comment": "Total tasks"},
     {"name": "Completed Tasks", "expr": "SUM(CASE WHEN status = 'done' THEN 1 ELSE 0 END)", "comment": "Number of completed tasks"},
     {"name": "On-Time Rate", "expr": "SUM(CASE WHEN completed_date <= due_date THEN 1 ELSE 0 END) * 1.0 / NULLIF(SUM(CASE WHEN status = 'done' THEN 1 ELSE 0 END), 0)", "comment": "Fraction of tasks completed by due date"},
     {"name": "Avg Cycle Time Days", "expr": "AVG(DATEDIFF(completed_date, created_date))", "comment": "Average days from creation to completion"},
     {"name": "Effort Accuracy", "expr": "AVG(actual_hours / NULLIF(estimated_hours, 0))", "comment": "Ratio of actual to estimated hours (1.0 = perfect)"},
     {"name": "Overdue Tasks", "expr": "SUM(CASE WHEN due_date < CURRENT_DATE() AND status != 'done' THEN 1 ELSE 0 END)", "comment": "Tasks past due date still open"}],
   "joins": [{"name": "projects", "source": "pm.projects", "on": "source.project_id = projects.project_id"}]}
]""",
}


def _select_few_shot(context: str) -> str:
    """Pick the best few-shot example based on domain keywords in the context."""
    ctx_lower = context.lower()
    scores: dict[str, int] = {}
    domain_keywords = {
        "healthcare": ["patient", "encounter", "provider", "clinical", "diagnosis", "admit", "discharge", "readmission", "icd", "npi"],
        "finance": ["transaction", "account", "ledger", "balance", "deposit", "withdrawal", "fraud", "loan", "interest", "portfolio"],
        "sales": ["order", "customer", "revenue", "product", "invoice", "shipment", "discount", "cart", "purchase"],
        "project_management": ["project", "task", "milestone", "sprint", "resource", "assignment", "issue", "ticket", "backlog", "epic", "story", "incident"],
    }
    for domain, keywords in domain_keywords.items():
        scores[domain] = sum(1 for kw in keywords if kw in ctx_lower)
    best = max(scores, key=scores.get) if max(scores.values()) > 0 else "sales"
    return _FEW_SHOT_BY_DOMAIN[best]


def _load_reference_rules() -> str:
    """Load anti-patterns and validation checklist from metric_view_reference.json."""
    ref_path = os.path.join(os.path.dirname(__file__), "..", "..", "..", "configurations", "agent_references", "metric_view_reference.json")
    try:
        with open(ref_path) as f:
            ref = json.load(f)
    except Exception:
        return ""
    parts = []
    if ref.get("anti_patterns"):
        parts.append("ANTI-PATTERNS (NEVER do these):")
        for ap in ref["anti_patterns"]:
            parts.append(f"  - {ap}")
    if ref.get("validation_checklist"):
        parts.append("SELF-CHECK before outputting:")
        for vc in ref["validation_checklist"]:
            parts.append(f"  - {vc}")
    return "\n".join(parts)


_REFERENCE_RULES_BLOCK = _load_reference_rules()


def _build_prompt(questions: list[str], context: str) -> str:
    q_block = "\n".join(f"  {i+1}. {q}" for i, q in enumerate(questions))
    few_shot = _select_few_shot(context)
    return f"""You are a data modeler building a semantic layer for Databricks Unity Catalog.

TASK: Generate metric view definitions (as a JSON array) that enable answering the business questions below.

ANALYTICAL QUALITY (HIGHEST PRIORITY):
- Every metric view MUST include at least one RATIO measure (x / NULLIF(y, 0)) and one computed dimension (CASE, DATE_TRUNC)
- Include RATE measures (conditional_count * 1.0 / NULLIF(total, 0)) for any entity with status/outcome columns
- Every KPI listed in the REQUIRED KPIs section MUST appear as a measure in at least one metric view. Map KPI formulas directly to measure expressions. If a KPI cannot be implemented, note why in the view comment
- Organize views around analytical themes, not just one-per-table. Multiple views from the same source are encouraged if they address different analytical angles
- When Entity types are annotated: People -> counts, rates, segmentation; Transactions -> volumes, values, time-based rates; Resources -> utilization, efficiency ratios
- Use COLUMN PROPERTY ANNOTATIONS: is_temporal -> date dimensions; is_categorical -> grouping dims; is_identifier -> count-distinct measures
- Use PROFILING SUMMARIES: low-cardinality (< 50 distinct) -> dimensions; high-cardinality -> measure inputs or filters
- Skip non-quantitative questions (document search, free-text lookups) silently

JOIN AND RELATIONSHIP RULES:
- Include joins for FK relationships where the join is relevant to the metric being computed and confidence is high.
- Not all FKs need to be used -- join only where the FK supports meaningful cross-table measures or dimensions.
- In join "on" clauses, always reference the source table as "source": "on": "source.fk_col = joined_alias.pk_col"
- CRITICAL: Each join's "on" clause must ONLY reference "source" on one side. Do NOT chain joins (e.g., "alias_a.col = alias_b.col"). If you need data from a table that only connects through another join, skip it. Star-schema only.
- COLUMN REFERENCING: Use "source.col" for source table columns, "join_alias.col" for joined table columns.
- Include joins when FK relationships OR GRAPH RELATIONSHIPS show a valid path
- Cross-table metrics are encouraged when FKs exist: join fact tables to dimension tables for breakdowns and ratios
- EXISTING METRIC VIEWS: do NOT duplicate -- build on existing coverage

STRUCTURE:
- Every view needs at least one measure, one dimension, a top-level "comment", and comments on each dimension/measure
- Names must be unique and descriptive (e.g. staffing_efficiency_metrics, ed_throughput_analysis)
- Use "filter" for persistent WHERE clauses; use measure-level FILTER for conditional aggregation
- Output ONLY a valid JSON array, no explanation

SQL SYNTAX REMINDERS:
- DATE_TRUNC('MONTH', col) -- always single-quote the interval
- Single-quote ALL string literals in comparisons, CASE results, IN lists, CONCAT separators
- Standard aggregates: SUM, COUNT, AVG, MIN, MAX, COUNT(DISTINCT ...)
- FILTER syntax: SUM(col) FILTER (WHERE condition)
- WINDOW MEASURES for rolling/cumulative KPIs: put aggregate in "expr" and add a "window" object:
  {{"name": "30-Day Rolling Revenue", "expr": "AVG(SUM(amount))", "window": {{"order_by": "date_col", "range": "INTERVAL 30 DAYS PRECEDING"}}}}
  Use range for time-based windows, rows for row-count windows (e.g. "UNBOUNDED PRECEDING" for cumulative)
- For MoM/YoY growth, use FILTER on date ranges rather than window functions

AGGREGATION CORRECTNESS:
- Ratios must use NULLIF in denominator: SUM(a) / NULLIF(SUM(b), 0), NEVER SUM(a) / SUM(b)
- AVG of a pre-aggregated value is usually wrong. For "average revenue per customer", use SUM(revenue) / NULLIF(COUNT(DISTINCT customer_id), 0), not AVG(revenue)
- Percentages: SUM(CASE WHEN cond THEN 1 ELSE 0 END) * 100.0 / NULLIF(COUNT(*), 0)

{_REFERENCE_RULES_BLOCK}

EXAMPLE:
{few_shot}

CATALOG METADATA:
{context}

BUSINESS QUESTIONS:
{q_block}

OUTPUT (JSON array only):"""


def _parse_ai_json(response: str) -> list[dict]:
    text = response.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    start = text.find("[")
    end = text.rfind("]")
    if start == -1 or end == -1:
        return []
    try:
        return json.loads(text[start : end + 1])
    except json.JSONDecodeError:
        results = []
        depth, obj_start = 0, None
        for i, ch in enumerate(text[start : end + 1]):
            if ch == "{":
                if depth == 0:
                    obj_start = i + start
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0 and obj_start is not None:
                    try:
                        results.append(json.loads(text[obj_start : i + start + 1]))
                    except json.JSONDecodeError:
                        pass
                    obj_start = None
        return results


_SQL_KEYWORDS = {
    "SUM",
    "COUNT",
    "AVG",
    "MIN",
    "MAX",
    "DATE_TRUNC",
    "DISTINCT",
    "MONTH",
    "QUARTER",
    "YEAR",
    "WEEK",
    "DAY",
    "HOUR",
    "MINUTE",
    "SECOND",
    "CAST",
    "AS",
    "STRING",
    "INT",
    "BIGINT",
    "DOUBLE",
    "FLOAT",
    "DECIMAL",
    "DATE",
    "TIMESTAMP",
    "BOOLEAN",
    "COALESCE",
    "IF",
    "CASE",
    "WHEN",
    "THEN",
    "ELSE",
    "END",
    "AND",
    "OR",
    "NOT",
    "NULL",
    "TRUE",
    "FALSE",
    "CONCAT",
    "UPPER",
    "LOWER",
    "TRIM",
    "FILTER",
    "WHERE",
    "BETWEEN",
    "IN",
    "LIKE",
    "IS",
    "FROM",
    "TO",
    "DATEDIFF",
    "TIMESTAMPDIFF",
    "DATE_ADD",
    "DATE_SUB",
    "ADD_MONTHS",
    "ROUND",
    "ABS",
    "CEIL",
    "CEILING",
    "FLOOR",
    "POWER",
    "SQRT",
    "MOD",
    "LENGTH",
    "SUBSTRING",
    "REPLACE",
    "REGEXP_REPLACE",
    "REGEXP_EXTRACT",
    "SPLIT",
    "ARRAY",
    "MAP",
    "STRUCT",
    "NAMED_STRUCT",
    "EXPLODE",
    "COLLECT_LIST",
    "COLLECT_SET",
    "APPROX_COUNT_DISTINCT",
    "PERCENTILE",
    "PERCENTILE_APPROX",
    "STDDEV",
    "VARIANCE",
    "FIRST",
    "LAST",
    "NVL",
    "IFNULL",
    "NULLIF",
    "CURRENT_DATE",
    "CURRENT_TIMESTAMP",
    "MONTHS_BETWEEN",
    "TO_DATE",
    "TO_TIMESTAMP",
    "DATE_FORMAT",
    "UNIX_TIMESTAMP",
}


def _extract_column_refs(expr: str) -> list[tuple[str | None, str]]:
    """Extract column references as (alias_or_None, column_name) tuples.

    ``account.industry`` -> ``("account", "industry")``
    ``status``           -> ``(None, "status")``
    """
    cleaned = re.sub(r"'[^']*'", "", expr)
    cleaned = re.sub(r'"[^"]*"', "", cleaned)
    func_tokens = {
        m.group(1).upper() for m in re.finditer(r"\b([a-zA-Z_]\w*)\s*\(", cleaned)
    }
    refs: list[tuple[str | None, str]] = []
    seen: set[str] = set()
    qualified_cols: set[str] = set()
    for m in re.finditer(r"\b([a-zA-Z_]\w*)\.([a-zA-Z_]\w*)\b", cleaned):
        alias, col = m.group(1), m.group(2)
        if alias.upper() not in _SQL_KEYWORDS and col.upper() not in _SQL_KEYWORDS:
            key = f"{alias}.{col}"
            if key not in seen:
                refs.append((alias, col))
                seen.add(key)
                qualified_cols.add(alias)
                qualified_cols.add(col)
    for tok in re.findall(r"\b([a-zA-Z_]\w*)\b", cleaned):
        if (
            tok.upper() not in _SQL_KEYWORDS
            and tok.upper() not in func_tokens
            and tok not in qualified_cols
            and tok not in seen
        ):
            refs.append((None, tok))
            seen.add(tok)
    return refs


def _autofix_dimension_columns(defn: dict, table_cols: dict[str, set[str]]) -> dict:
    """Attempt to fuzzy-match hallucinated column names to actual columns.

    table_cols maps alias -> set of column names. 'source' is the primary table.
    Join aliases are also included when available.
    """
    from difflib import get_close_matches

    # Build per-alias lookup and a combined lookup across all aliases
    alias_cols: dict[str, dict[str, str]] = {}
    all_cols_lower: dict[str, str] = {}
    for alias, cols in table_cols.items():
        alias_cols[alias.lower()] = {c.lower(): c for c in cols}
        for c in cols:
            all_cols_lower[c.lower()] = c

    # Collect known alias names (source + join aliases)
    known_aliases = set(alias_cols.keys())

    for item_type in ("dimensions", "measures"):
        for item in defn.get(item_type, []):
            expr = item.get("expr", "")
            if not expr:
                continue
            for m in re.finditer(r"\b([a-zA-Z_]\w*)\.([a-zA-Z_]\w*)\b", expr):
                prefix = m.group(1)
                col = m.group(2)
                prefix_lower = prefix.lower()
                if prefix_lower not in known_aliases:
                    continue
                # Check if column exists for this alias
                alias_map = alias_cols.get(prefix_lower, all_cols_lower)
                if col.lower() in alias_map:
                    continue
                candidates = list(alias_map.keys())
                dim_name = item.get("name", "")
                snake = re.sub(r"\s+", "_", dim_name).lower()
                check_vals = [col.lower(), snake]
                for cv in check_vals:
                    matches = get_close_matches(cv, candidates, n=1, cutoff=0.6)
                    if matches:
                        actual = alias_map[matches[0]]
                        expr = expr.replace(f"{prefix}.{col}", f"{prefix}.{actual}")
                        item["expr"] = expr
                        break
    return defn


def _validate_definition_structure(defn: dict) -> list[str]:
    """Structural validation: check source table exists and columns are valid."""
    errors = []
    source = defn.get("source", "")
    if not source:
        errors.append("Missing source table")
        return errors
    if not defn.get("dimensions") and not defn.get("measures"):
        errors.append("Definition must have at least one dimension or measure")

    parts = source.split(".")
    if len(parts) != 3:
        return errors

    cat, sch, tbl = parts
    source_cols: set[str] = set()
    try:
        rows = execute_sql(
            f"SELECT column_name FROM `{cat}`.information_schema.columns "
            f"WHERE table_catalog = '{cat}' AND table_schema = '{sch}' AND table_name = '{tbl}'"
        )
        source_cols = {r["column_name"].lower() for r in rows}
    except Exception:
        errors.append(f"Could not verify table {source}")
        return errors

    if not source_cols:
        errors.append(f"Table {source} not found in information_schema")
        return errors

    alias_cols: dict[str, set[str]] = {"source": source_cols, tbl: source_cols}
    known_aliases = {"source", tbl}
    for j in defn.get("joins", []):
        j_source = j.get("source", "")
        j_alias = j.get("name", j_source.split(".")[-1])
        j_parts = j_source.split(".")
        known_aliases.add(j_alias)
        if len(j_parts) == 3:
            known_aliases.add(j_parts[2])
            try:
                j_rows = execute_sql(
                    f"SELECT column_name FROM `{j_parts[0]}`.information_schema.columns "
                    f"WHERE table_catalog = '{j_parts[0]}' AND table_schema = '{j_parts[1]}' AND table_name = '{j_parts[2]}'"
                )
                j_cols = {r["column_name"].lower() for r in j_rows}
                alias_cols[j_alias] = j_cols
                alias_cols[j_parts[2]] = j_cols
            except Exception:
                pass

    all_cols = set()
    for cs in alias_cols.values():
        all_cols |= cs

    for item_type in ("dimensions", "measures"):
        for item in defn.get(item_type, []):
            col_refs = _extract_column_refs(item.get("expr", ""))
            for alias, col in col_refs:
                if alias:
                    target = alias_cols.get(alias)
                    if target is None:
                        errors.append(
                            f"{item_type} {item.get('name', '')}: column {alias} not found in {source}"
                        )
                    elif col.lower() not in target:
                        errors.append(
                            f"{item_type} {item.get('name', '')}: column {col} not found in alias {alias}"
                        )
                else:
                    if col.lower() not in all_cols and col.lower() not in known_aliases:
                        errors.append(
                            f"{item_type} {item.get('name', '')}: column {col} not found in {source}"
                        )

    # Detect chained joins (on clause references another join alias instead of source)
    join_aliases = {j.get("name", "").lower() for j in defn.get("joins", []) if j.get("name")}
    ref_pat = re.compile(r"\b([A-Za-z_]\w*)\.\w+")
    for j in defn.get("joins", []):
        on = j.get("on", "")
        own_alias = j.get("name", "").lower()
        refs_in_on = {m.group(1).lower() for m in ref_pat.finditer(on)}
        refs_in_on.discard("source")
        chained = refs_in_on & join_aliases - {own_alias}
        if chained:
            errors.append(
                f"Join '{j.get('name', '?')}' references another join alias {chained} -- only 'source' references are supported"
            )

    return errors


_DATE_TRUNC_INTERVALS = {
    "YEAR",
    "QUARTER",
    "MONTH",
    "WEEK",
    "DAY",
    "HOUR",
    "MINUTE",
    "SECOND",
}
_SQL_RESERVED = {
    "THEN",
    "ELSE",
    "END",
    "AND",
    "OR",
    "NOT",
    "NULL",
    "TRUE",
    "FALSE",
    "CASE",
    "WHEN",
    "IN",
    "IS",
    "LIKE",
    "BETWEEN",
    "SELECT",
    "FROM",
    "WHERE",
    "FILTER",
    "DISTINCT",
    "SUM",
    "AVG",
    "COUNT",
    "MIN",
    "MAX",
    "DATE_TRUNC",
    "IF",
    "COALESCE",
    "NULLIF",
    "OVER",
    "PARTITION",
    "BY",
    "ORDER",
    "ASC",
    "DESC",
}


def _fix_unquoted_literals(expr: str) -> str:
    """Quote bare words/phrases used as string literals in comparisons."""

    def _replacer(m):
        op = m.group(1)
        value = m.group(2).strip()
        trail = m.group(3)
        if not value:
            return m.group(0)
        if value.startswith("'") or value.startswith('"'):
            return m.group(0)
        if re.match(r"^-?\d+(\.\d+)?$", value):
            return m.group(0)
        if "." in value and " " not in value:
            return m.group(0)
        if "(" in value:
            return m.group(0)
        if value.upper() in _SQL_RESERVED or value.upper() in _DATE_TRUNC_INTERVALS:
            return m.group(0)
        return f"{op}'{value}'{trail}"

    return re.sub(
        r"([=!<>]+\s*)(.*?)(\s+(?:THEN|ELSE|END|AND|OR|WHEN)\b|\s*[,)]|$)",
        _replacer,
        expr,
        flags=re.IGNORECASE,
    )


def _fix_then_else_literals(expr: str) -> str:
    """Quote bare text after THEN/ELSE that isn't already quoted or a number/column/keyword."""
    def _replacer(m):
        kw = m.group(1)
        body = m.group(2).strip()
        if not body:
            return m.group(0)
        if body.startswith("'") or body.startswith('"'):
            return m.group(0)
        if re.match(r"^-?\d+(\.\d+)?$", body):
            return m.group(0)
        if "." in body and " " not in body:
            return m.group(0)
        if "(" in body:
            return m.group(0)
        tokens = body.split()
        if len(tokens) == 1 and tokens[0].upper() in _SQL_RESERVED:
            return m.group(0)
        if len(tokens) == 1 and re.match(r"^[A-Za-z_]\w*$", tokens[0]):
            if tokens[0].upper() not in _SQL_RESERVED:
                return f"{kw} '{body}'"
            return m.group(0)
        return f"{kw} '{body}'"

    return re.sub(
        r"\b(THEN|ELSE)\s+(.*?)(?=\s+(?:WHEN|ELSE|END)\b)",
        _replacer,
        expr,
        flags=re.IGNORECASE,
    )


def _fix_in_clause_literals(expr: str) -> str:
    """Quote bare words inside IN (...) clauses."""

    def _fix_in_body(m):
        prefix = m.group(1)
        body = m.group(2)
        tokens = [t.strip() for t in body.split(",")]
        fixed = []
        for tok in tokens:
            if not tok:
                fixed.append(tok)
            elif tok.startswith("'") or tok.startswith('"'):
                fixed.append(tok)
            elif re.match(r"^-?\d+(\.\d+)?$", tok):
                fixed.append(tok)
            elif tok.upper() in _SQL_RESERVED:
                fixed.append(tok)
            else:
                fixed.append(f"'{tok}'")
        return f"{prefix}{', '.join(fixed)})"

    return re.sub(
        r"(\bIN\s*\()([^)]+)\)",
        _fix_in_body,
        expr,
        flags=re.IGNORECASE,
    )


def _fix_concat_separators(expr: str) -> str:
    """Quote bare separator tokens between commas (e.g. -Q, /, : in CONCAT)."""
    def _repl(m):
        tok = m.group(1).strip()
        if re.match(r"^-?\d+(\.\d+)?$", tok):
            return m.group(0)
        return f", '{tok}',"
    return re.sub(
        r",\s*([^\w\s'\"`(][^'\"`(,)]{0,4})\s*,",
        _repl,
        expr,
    )


def _fix_like_patterns(expr: str) -> str:
    """Quote bare LIKE/NOT LIKE patterns: ``col LIKE HW%`` -> ``col LIKE 'HW%'``."""
    def _repl(m):
        prefix = m.group(1)
        pat = m.group(2).strip()
        if pat.startswith("'") or pat.startswith('"'):
            return m.group(0)
        return f"{prefix}'{pat}'"
    return re.sub(r"(LIKE\s+)([^'\"\s(]+)", _repl, expr, flags=re.IGNORECASE)


def _fix_position_bare_char(expr: str) -> str:
    """Quote bare non-alnum char in POSITION(X IN ...): POSITION(- IN col) -> POSITION('-' IN col)."""
    def _repl(m):
        ch = m.group(1).strip()
        if ch.startswith("'") or ch.startswith('"'):
            return m.group(0)
        return f"POSITION('{ch}' IN{m.group(2)}"
    return re.sub(r"POSITION\(\s*([^\w\s'\"]+)\s+(IN\b)", _repl, expr, flags=re.IGNORECASE)


def _fix_double_commas(expr: str) -> str:
    """Collapse empty arguments: CONCAT(a, , b) -> CONCAT(a, b)."""
    while ", ," in expr:
        expr = expr.replace(", ,", ",")
    while ",," in expr:
        expr = expr.replace(",,", ",")
    return expr


def _fix_none_literal(expr: str) -> str:
    """Replace Python None leaked into SQL with NULL."""
    return re.sub(r"\bNone\b", "NULL", expr)


def _fix_concat_bare_first_arg(expr: str) -> str:
    """Quote bare single-word non-column first arg in CONCAT: CONCAT(Q, ...) -> CONCAT('Q', ...)."""
    def _repl(m):
        fn = m.group(1)
        arg = m.group(2).strip()
        if arg.startswith("'") or arg.startswith('"'):
            return m.group(0)
        if "." in arg or arg.upper() in _SQL_RESERVED or re.match(r"^-?\d", arg):
            return m.group(0)
        if len(arg) <= 3 and arg.isalpha():
            return f"{fn}'{arg}',"
        return m.group(0)
    return re.sub(r"(CONCAT\(\s*)([^',\s]+)\s*,", _repl, expr, flags=re.IGNORECASE)


def _autofix_expr(expr: str) -> str:
    """Fix common AI expression mistakes before validation."""

    # Fix unquoted DATE_TRUNC intervals: DATE_TRUNC(WEEK, col) -> DATE_TRUNC('WEEK', col)
    def _fix_date_trunc(m):
        interval = m.group(1)
        rest = m.group(2)
        if interval.upper() in _DATE_TRUNC_INTERVALS:
            return f"DATE_TRUNC('{interval}'{rest}"
        return m.group(0)

    expr = re.sub(
        r"DATE_TRUNC\(\s*([A-Za-z]+)(,)", _fix_date_trunc, expr, flags=re.IGNORECASE
    )

    for iv in _DATE_TRUNC_INTERVALS:
        pat = re.compile(rf"\b{iv}\s*\(([^)]+)\)", re.IGNORECASE)
        match = pat.search(expr)
        if match and expr.strip().upper().startswith(iv.upper()):
            expr = pat.sub(rf"DATE_TRUNC('{iv}', \1)", expr)

    expr = _fix_none_literal(expr)
    expr = _fix_double_commas(expr)
    expr = _fix_position_bare_char(expr)
    expr = _fix_concat_bare_first_arg(expr)
    expr = _fix_unquoted_literals(expr)
    expr = _fix_then_else_literals(expr)
    expr = _fix_in_clause_literals(expr)
    expr = _fix_concat_separators(expr)
    expr = _fix_like_patterns(expr)
    return expr


def _build_from_clause(source_table: str, joins: list[dict] | None = None) -> str:
    """Build ``FROM source AS source [JOIN ...]`` clause for expression dry-runs."""
    clause = f"{source_table} AS source"
    for j in (joins or []):
        j_alias = j.get("name", j.get("source", "").split(".")[-1])
        j_src = j.get("source", "")
        on = j.get("on", "1=1")
        if j_src:
            clause += f" LEFT JOIN {j_src} AS {j_alias} ON {on}"
    return clause


def _validate_expr(expr: str, source_table: str, joins: list[dict] | None = None) -> tuple:
    """Test a SQL expression with optional joins. Returns (error_or_None, possibly_fixed_expr)."""
    from_clause = _build_from_clause(source_table, joins)
    try:
        execute_sql(f"SELECT {expr} FROM {from_clause} LIMIT 0")
        return None, expr
    except Exception as e:
        err_str = str(e)
        m = re.search(r"UNRESOLVED_COLUMN.*?name `(\w+)`", err_str)
        if m:
            bare = m.group(1)
            fixed = re.sub(
                rf"([=!<>]\s{{0,4}}){re.escape(bare)}(?=[\s),$]|$)",
                rf"\1'{bare}'",
                expr,
            )
            if fixed != expr:
                try:
                    execute_sql(f"SELECT {fixed} FROM {from_clause} LIMIT 0")
                    return None, fixed
                except Exception:
                    pass
        return err_str, expr


def _score_definition_complexity(defn: dict) -> dict:
    """Score a metric view definition's analytical richness.

    Returns {"complexity_score": int, "complexity_level": str}.
    """
    score = 0
    _COMPUTED_DIM_PATTERNS = re.compile(r"CASE\b|DATE_TRUNC\b|CONCAT\b", re.IGNORECASE)
    for dim in defn.get("dimensions", []):
        if _COMPUTED_DIM_PATTERNS.search(dim.get("expr", "")):
            score += 1
    for meas in defn.get("measures", []):
        expr = meas.get("expr", "")
        if re.search(r"/\s*NULLIF\b", expr, re.IGNORECASE):
            score += 2
        elif re.search(r"FILTER\b|DISTINCT\b|CASE\b", expr, re.IGNORECASE):
            score += 1
    if defn.get("joins"):
        score += 2
    if defn.get("filter"):
        score += 1
    if score < 3:
        level = "trivial"
    elif score <= 5:
        level = "moderate"
    else:
        level = "rich"
    return {"complexity_score": score, "complexity_level": level}


def _sl_self_repair(defn: dict, errors: list[str], model: str) -> dict | None:
    """Phase 3: LLM-powered repair for a failed metric view definition. Returns fixed dict or None."""
    source = defn.get("source", "")
    col_context = ""
    if source:
        try:
            source_esc = source.replace(chr(39), chr(39)+chr(39))
            short_name = source.split(".")[-1].replace(chr(39), chr(39)+chr(39))
            cols = execute_sql(
                f"SELECT column_name, data_type FROM {fq('column_knowledge_base')} "
                f"WHERE table_name = '{source_esc}' OR table_name LIKE '%{short_name}'"
            )
            if cols:
                col_context = "Available columns: " + ", ".join(
                    f"{c['column_name']} ({c.get('data_type', '')})" for c in cols
                )
        except Exception:
            pass

    prompt = f"""Fix this metric view definition. It failed validation with these errors:

ERRORS:
{chr(10).join(f'  - {e}' for e in errors)}

CURRENT DEFINITION:
{json.dumps(defn, indent=2)}

{col_context}

Fix ONLY the broken expressions. Keep all valid parts unchanged.
Use standard SQL: SUM, COUNT, AVG, MIN, MAX, DATE_TRUNC('MONTH', col).
Always single-quote string literals. Only reference columns that exist.

Return ONLY the fixed JSON definition (single object, not array)."""

    try:
        escaped = prompt.replace("'", "''")
        rows = execute_sql(
            f"SELECT AI_QUERY('{model}', '{escaped}') as response", timeout=120
        )
        response = rows[0]["response"] if rows else ""
        fixed = _parse_single_json(response)
        fixed.setdefault("source", source)
        fixed.setdefault("name", defn.get("name", ""))
        return fixed
    except Exception as exc:
        logger.warning("Self-repair AI call failed: %s", exc)
        return None


def _normalize_joins(defn: dict) -> dict:
    """Rewrite join 'on' clauses to use source.col instead of raw tablename.col for the source table."""
    source_short = (defn.get("source") or "").split(".")[-1]
    if not source_short:
        return defn
    for join in defn.get("joins", []):
        on = join.get("on", "")
        if source_short + "." in on:
            join["on"] = re.sub(
                rf"\b{re.escape(source_short)}\.", "source.", on
            )
    return defn


def _fix_join_alias_refs(defn: dict) -> dict:
    """Rewrite expressions that use invalid join aliases to the closest valid one."""
    joins = defn.get("joins", [])
    valid_aliases = {"source"}
    source_short = (defn.get("source") or "").split(".")[-1]
    alias_from_table: dict[str, str] = {}
    if source_short:
        alias_from_table[source_short.lower()] = "source"
    for j in joins:
        alias = j.get("name", "")
        if alias:
            valid_aliases.add(alias)
            j_short = (j.get("source") or "").split(".")[-1]
            if j_short:
                alias_from_table[j_short.lower()] = alias

    if not joins:
        return defn

    _ref_pat = re.compile(r"\b([A-Za-z_]\w*)\.([A-Za-z_]\w*)\b")

    def _fix_expr(expr: str) -> str:
        def _repl(m):
            alias, col = m.group(1), m.group(2)
            if alias in valid_aliases:
                return m.group(0)
            mapped = alias_from_table.get(alias.lower())
            if mapped:
                return f"{mapped}.{col}"
            for va in valid_aliases:
                if alias.lower() in va.lower() or va.lower() in alias.lower():
                    return f"{va}.{col}"
            return m.group(0)
        return _ref_pat.sub(_repl, expr)

    for section in ("dimensions", "measures"):
        for item in defn.get(section, []):
            if "expr" in item:
                item["expr"] = _fix_expr(item["expr"])
    filt = defn.get("filter")
    if isinstance(filt, str):
        defn["filter"] = _fix_expr(filt)
    return defn


def _flatten_chained_joins(defn: dict) -> dict:
    """Drop joins whose 'on' clause references another join alias instead of 'source'.

    Also removes dimensions/measures that depend on the dropped alias.
    """
    joins = defn.get("joins", [])
    if not joins:
        return defn

    join_aliases = {j.get("name", "").lower() for j in joins if j.get("name")}
    ref_pat = re.compile(r"\b([A-Za-z_]\w*)\.\w+")

    dropped_aliases: set[str] = set()
    kept_joins: list[dict] = []
    for j in joins:
        on = j.get("on", "")
        refs_in_on = {m.group(1).lower() for m in ref_pat.finditer(on)}
        refs_in_on.discard("source")
        # If 'on' references another join alias (not source, not self), it's chained
        own_alias = j.get("name", "").lower()
        chained_refs = refs_in_on & join_aliases - {own_alias}
        if chained_refs:
            logger.warning(
                "Dropping chained join '%s': on clause references other join alias(es) %s",
                j.get("name", "?"), chained_refs,
            )
            dropped_aliases.add(own_alias)
        else:
            kept_joins.append(j)

    if not dropped_aliases:
        return defn

    defn["joins"] = kept_joins

    # Remove dimensions/measures that reference dropped aliases
    for section in ("dimensions", "measures"):
        items = defn.get(section, [])
        cleaned = []
        for item in items:
            expr = item.get("expr", "")
            expr_refs = {m.group(1).lower() for m in ref_pat.finditer(expr)}
            if expr_refs & dropped_aliases:
                logger.warning(
                    "Dropping %s '%s': references dropped join alias",
                    section[:-1], item.get("name", "?"),
                )
            else:
                cleaned.append(item)
        defn[section] = cleaned

    return defn


def _build_plan_prompt(questions: list[str], context: str) -> str:
    q_block = "\n".join(f"  {i+1}. {q}" for i, q in enumerate(questions))
    return f"""You are a data modeler planning a semantic layer for Databricks Unity Catalog.

TASK: Output a PLAN only (no SQL). Reply with a single JSON object: {{ "views": [ ... ] }}.

For each metric view in "views", include:
- "name": unique snake_case name (e.g. order_performance_metrics)
- "source": fully qualified source table (catalog.schema.table)
- "comment": one sentence purpose
- "joins": array of {{ "name": "<alias>", "source": "catalog.schema.table", "on": "source.<fk_col> = <alias>.<pk_col>" }}
  Include joins where FK relationships have high confidence and are relevant to the business questions. Prefer simpler views with fewer joins over complex views that may fail. A view with 0-2 well-chosen joins is better than one with 5 forced joins.
- "dimensions": array of {{ "name": "Display Name", "comment": "what it is" }} (no expr)
- "measures": array of {{ "name": "Display Name", "comment": "what it measures" }} (no expr)
- "question_indices": array of 0-based question indices this view answers

Create measures that match the business questions (ratios, rates, KPIs); avoid generic row count unless a question explicitly asks for it. Each view must have at least one dimension and one measure. Cross-table breakdowns using joined dimension tables are strongly preferred.

CATALOG METADATA:
{context}

BUSINESS QUESTIONS:
{q_block}

OUTPUT (single JSON object with "views" key only, no explanation):"""


def _build_generate_prompt_for_plan(plan_view: dict, questions: list[str], context: str) -> str:
    q_refs = plan_view.get("question_indices", [])
    q_block = "\n".join(f"  {i+1}. {questions[i]}" for i in q_refs if 0 <= i < len(questions))
    plan_str = json.dumps(plan_view, indent=2)
    return f"""You are a data modeler. Output exactly ONE JSON object for a single metric view (not an array).

PLANNED VIEW (names only; you must add "expr" for each dimension and measure):
{plan_str}

RULES:
- Output a single object with keys: name, source, comment, filter (optional), dimensions, measures, joins.
- dimensions: array of {{ "name", "expr", "comment" }}. expr must be valid Databricks SQL using ONLY columns from the metadata below.
- measures: array of {{ "name", "expr", "comment" }}. Use SUM, COUNT, AVG, FILTER, etc. String literals single-quoted.
- For window measures (rolling averages, cumulative): use "window" sub-object: {{"order_by": "col", "range": "INTERVAL N DAYS PRECEDING"}}.
- joins: You MUST implement ALL joins from the plan exactly. Use: on: source.<fk_column> = <join_name>.<pk_column>. Keep same join names as plan. If the plan includes joins, they are REQUIRED in your output. Add dimensions/measures that reference joined table columns.
- COLUMN REFERENCING: Use "source.col" for source table columns, "join_alias.col" for joined table columns. Example: "expr": "source.order_date" or "expr": "account.industry". NEVER use bare column names when joins are present -- always qualify with the alias.
- Ratios must use NULLIF in denominator. AVG of pre-aggregated values is usually wrong.
- Only use column names that appear in the metadata.

{_REFERENCE_RULES_BLOCK}

CATALOG METADATA:
{context}

QUESTIONS this view answers:
{q_block}

OUTPUT (one JSON object only, no array, no explanation):"""


def _parse_single_json_safe(response: str) -> dict:
    """Like _parse_single_json but returns {} instead of raising."""
    text = response.strip()
    text = re.sub(r"^```(?:json)?\s*", "", text)
    text = re.sub(r"\s*```$", "", text)
    start = text.find("{")
    end = text.rfind("}") + 1
    if start == -1 or end <= start:
        return {}
    try:
        return json.loads(text[start:end])
    except json.JSONDecodeError:
        return {}


def _inject_fk_joins(plan_views: list[dict], tables: list[str], cat: str, sch: str):
    """Add high-confidence FK joins to plan views, capped at 3 per view."""
    fq_tables = [t if "." in t else f"{cat}.{sch}.{t}" for t in tables]
    in_clause = ", ".join(f"'{t}'" for t in fq_tables)
    try:
        fk_rows = execute_sql(
            f"SELECT src_table, dst_table, src_column, dst_column "
            f"FROM {fq('fk_predictions')} WHERE is_fk = 'true' AND final_confidence >= 0.85 "
            f"AND (src_table IN ({in_clause}) OR dst_table IN ({in_clause}))"
        )
    except Exception:
        return plan_views
    if not fk_rows:
        return plan_views

    fk_by_table: dict[str, list[dict]] = {}
    for fk in fk_rows:
        fk_by_table.setdefault(fk["src_table"], []).append(fk)
        fk_by_table.setdefault(fk["dst_table"], []).append(fk)

    for pv in plan_views:
        src = pv.get("source", "")
        if not src:
            continue
        fks = fk_by_table.get(src, [])
        if not fks:
            continue
        existing_join_sources = {j.get("source", "") for j in pv.get("joins", [])}
        added = 0
        for fk in fks:
            if added >= 3:
                break
            if fk["src_table"] == src:
                join_table, fk_col, pk_col = fk["dst_table"], fk["src_column"], fk["dst_column"]
            else:
                join_table, fk_col, pk_col = fk["src_table"], fk["dst_column"], fk["src_column"]
            if join_table in existing_join_sources:
                continue
            alias = join_table.split(".")[-1]
            fk_col = fk_col.split(".")[-1]
            pk_col = pk_col.split(".")[-1]
            pv.setdefault("joins", []).append({
                "name": alias,
                "source": join_table,
                "on": f"source.{fk_col} = {alias}.{pk_col}",
            })
            existing_join_sources.add(join_table)
            added += 1
    return plan_views


def _defn_to_yaml(defn: dict) -> str:
    """Convert a metric view JSON definition to YAML body for CREATE VIEW WITH METRICS."""
    mv: dict = {"version": "1.1", "source": defn["source"]}
    if defn.get("comment"):
        mv["comment"] = defn["comment"]
    if defn.get("filter"):
        mv["filter"] = defn["filter"]
    mv["dimensions"] = [
        {k: v for k, v in {"name": d["name"], "expr": d["expr"], "comment": d.get("comment")}.items() if v}
        for d in defn.get("dimensions", [])
    ]
    measures_out = []
    for m in defn.get("measures", []):
        entry = {k: v for k, v in {"name": m["name"], "expr": m["expr"], "comment": m.get("comment")}.items() if v}
        if m.get("window"):
            entry["window"] = m["window"]
        measures_out.append(entry)
    mv["measures"] = measures_out
    if defn.get("joins"):
        mv["joins"] = defn["joins"]
    return yaml.dump(mv, default_flow_style=False, sort_keys=False)


def _yaml_dry_run(defn: dict, cat: str, sch: str) -> Optional[str]:
    """Attempt CREATE VIEW WITH METRICS LANGUAGE YAML; return error string or None."""
    try:
        yaml_body = _defn_to_yaml(defn)
        mv_name = defn.get("name", "dry_run_test")
        dry_name = f"`{cat}`.`{sch}`.`_mv_dryrun_{mv_name}`"
        execute_sql(
            f"CREATE OR REPLACE VIEW {dry_name}\nWITH METRICS LANGUAGE YAML AS $$\n{yaml_body}$$",
            timeout=30,
        )
        execute_sql(f"DROP VIEW IF EXISTS {dry_name}", timeout=15)
        return None
    except Exception as e:
        err = str(e)
        if "DROP VIEW" in err:
            return None
        return f"YAML dry-run failed: {err}"


def _run_sl_generation(
    task_id: str,
    tables: list[str],
    questions: list[str],
    cat: str,
    sch: str,
    model: str,
    project_id: str = None,
    mode: str = "replace",
    business_context: str = None,
    profile_id: str = None,
):
    """Background thread for in-app metric view generation (two-phase)."""
    from datetime import datetime as _dt

    task = _sl_tasks[task_id]
    try:
        _ensure_semantic_layer_tables()

        # Persist questions for traceability
        if questions:
            now_ts = _dt.utcnow().isoformat()
            q_values = []
            for q in questions:
                q_esc = q.strip().replace("'", "''")
                if q_esc:
                    q_values.append(f"('{_uuid.uuid4()}', '{q_esc}', 'pending', '{now_ts}', NULL)")
            if q_values:
                try:
                    execute_sql(f"INSERT INTO {fq('semantic_layer_questions')} VALUES {', '.join(q_values)}", timeout=30)
                except Exception as exc:
                    logger.warning("Failed to persist questions: %s", exc)

        if mode == "replace_all" and project_id:
            try:
                execute_sql(
                    f"UPDATE {fq('metric_view_definitions')} SET status = 'superseded' "
                    f"WHERE project_id = '{project_id}' AND status != 'superseded'"
                )
            except Exception:
                pass

        task["stage"] = "building_context"
        context = _build_sl_context(tables, cat, sch, questions=questions, business_context=business_context, profile_id=profile_id)
        if not context.strip():
            task.update({"status": "error", "error": "No metadata found for selected tables. Run metadata generation first."})
            return

        # Phase 1: Plan
        task["stage"] = "planning"
        plan_prompt = _build_plan_prompt(questions, context)
        escaped = plan_prompt.replace("'", "''")
        rows = execute_sql(f"SELECT AI_QUERY('{model}', '{escaped}') as response", timeout=180)
        plan_response = rows[0]["response"] if rows else ""
        plan_views = []
        if plan_response:
            try:
                plan_data = _parse_single_json_safe(plan_response)
                plan_views = plan_data.get("views") or []
            except Exception:
                pass

        if plan_views:
            plan_views = _inject_fk_joins(plan_views, tables, cat, sch)

        # Phase 2: Generate each view individually, or fall back to single-shot
        response = ""
        definitions = []
        if plan_views:
            task["stage"] = "generating"
            task["planned"] = len(plan_views)
            for idx, pv in enumerate(plan_views):
                task["generating_view"] = idx + 1
                gen_prompt = _build_generate_prompt_for_plan(pv, questions, context)
                escaped = gen_prompt.replace("'", "''")
                try:
                    gen_rows = execute_sql(f"SELECT AI_QUERY('{model}', '{escaped}') as response", timeout=120)
                    gen_resp = gen_rows[0]["response"] if gen_rows else ""
                    defn = _parse_single_json_safe(gen_resp)
                    if defn and defn.get("source"):
                        defn.setdefault("name", pv.get("name", f"metric_view_{idx}"))
                        definitions.append(defn)
                except Exception as exc:
                    logger.warning("Phase 2 generation failed for '%s': %s", pv.get("name", ""), exc)
        else:
            # Fallback to single-shot
            task["stage"] = "calling_ai"
            prompt = _build_prompt(questions, context)
            escaped = prompt.replace("'", "''")
            rows = execute_sql(f"SELECT AI_QUERY('{model}', '{escaped}') as response", timeout=180)
            response = rows[0]["response"] if rows else ""
            definitions = _parse_ai_json(response)
        if not definitions:
            snippet = (response[:200] + "...") if len(response) > 200 else response
            task.update(
                {
                    "status": "error",
                    "error": f"AI returned no valid metric view definitions. Response preview: {snippet}",
                }
            )
            return

        task.update({"stage": "validating", "generated": len(definitions)})
        now = _dt.utcnow().isoformat()
        stats = {"generated": 0, "validated": 0, "failed": 0, "repaired": 0}
        per_definition_results: list[dict] = []

        for defn in definitions:
            defn_id = str(_uuid.uuid4())
            mv_name = defn.get("name", f"metric_view_{defn_id[:8]}")
            source = defn.get("source", "")

            # Auto-fix common AI expression mistakes before validation
            for item_type in ("dimensions", "measures"):
                for item in defn.get(item_type, []):
                    if item.get("expr"):
                        item["expr"] = _autofix_expr(item["expr"])
            if defn.get("filter"):
                defn["filter"] = _autofix_expr(defn["filter"])
            defn = _normalize_joins(defn)
            defn = _fix_join_alias_refs(defn)
            defn = _flatten_chained_joins(defn)

            # Fuzzy-match hallucinated column names to actual columns
            src = defn.get("source", "")
            if src:
                src_parts = src.split(".")
                if len(src_parts) == 3:
                    try:
                        _cols_rows = execute_sql(
                            f"SELECT column_name FROM `{src_parts[0]}`.information_schema.columns "
                            f"WHERE table_catalog = '{src_parts[0]}' AND table_schema = '{src_parts[1]}' AND table_name = '{src_parts[2]}'"
                        )
                        _tbl_cols: dict[str, set[str]] = {
                            "source": {r["column_name"].lower() for r in _cols_rows}
                        }
                        # Also fetch columns for join alias tables
                        for j in defn.get("joins", []):
                            j_src = j.get("source", "")
                            j_name = j.get("name", "")
                            j_parts = j_src.split(".")
                            if j_name and len(j_parts) == 3:
                                try:
                                    j_rows = execute_sql(
                                        f"SELECT column_name FROM `{j_parts[0]}`.information_schema.columns "
                                        f"WHERE table_catalog = '{j_parts[0]}' AND table_schema = '{j_parts[1]}' AND table_name = '{j_parts[2]}'"
                                    )
                                    _tbl_cols[j_name.lower()] = {r["column_name"].lower() for r in j_rows}
                                except Exception:
                                    pass
                        defn = _autofix_dimension_columns(defn, _tbl_cols)
                    except Exception:
                        pass

            json_str = json.dumps(defn).replace("'", "''")

            # In "replace" mode, supersede existing definitions with the same name
            if mode == "replace":
                mv_esc = mv_name.replace("'", "''")
                proj_clause = f" AND project_id = '{project_id}'" if project_id else ""
                try:
                    execute_sql(
                        f"UPDATE {fq('metric_view_definitions')} SET status = 'superseded' "
                        f"WHERE metric_view_name = '{mv_esc}' "
                        f"AND status != 'superseded'{proj_clause}"
                    )
                except Exception:
                    pass

            def _validate_defn(d: dict) -> list[str]:
                errs = _validate_definition_structure(d)
                if not errs:
                    d_joins = d.get("joins", [])
                    for itype in ("dimensions", "measures"):
                        for item in d.get(itype, []):
                            expr = item.get("expr", "")
                            if d.get("source") and expr:
                                err, fixed_expr = _validate_expr(expr, d["source"], d_joins)
                                if err:
                                    errs.append(f"{itype} '{item.get('name', '')}': {err}")
                                elif fixed_expr != expr:
                                    item["expr"] = fixed_expr
                return errs

            errors = _validate_defn(defn)

            # YAML dry-run: catches metric-YAML-specific failures that pass SELECT validation
            if not errors and defn.get("source"):
                yaml_err = _yaml_dry_run(defn, cat, sch)
                if yaml_err:
                    errors.append(yaml_err)
                    logger.info("YAML dry-run failed for '%s': %s", mv_name, yaml_err[:200])

            # Phase 3: Self-repair -- up to 2 LLM retries for failed definitions
            for repair_round in range(1, 3):
                if not errors:
                    break
                logger.info("Definition '%s' repair round %d (%d errors)", mv_name, repair_round, len(errors))
                repaired = _sl_self_repair(defn, errors, model)
                if not repaired:
                    break
                for itype in ("dimensions", "measures"):
                    for item in repaired.get(itype, []):
                        if item.get("expr"):
                            item["expr"] = _autofix_expr(item["expr"])
                if repaired.get("filter"):
                    repaired["filter"] = _autofix_expr(repaired["filter"])
                repair_errors = _validate_defn(repaired)
                if not repair_errors or len(repair_errors) < len(errors):
                    defn = repaired
                    errors = repair_errors
                    mv_name = defn.get("name", mv_name)
                    source = defn.get("source", source)
                    if not errors:
                        stats["repaired"] += 1
                    logger.info("Self-repair round %d %s for '%s' (%d remaining errors)",
                                repair_round, "succeeded" if not errors else "improved", mv_name, len(errors))
                else:
                    break

            cx = _score_definition_complexity(defn)

            if source and source.count('.') < 2:
                match = [t for t in tables if t.endswith('.' + source) or t.split('.')[-1] == source.split('.')[-1]]
                if match:
                    source = match[0]

            json_str = json.dumps(defn).replace("'", "''")
            status = "validated" if not errors else "failed"
            error_str = "; ".join(errors).replace("'", "''") if errors else ""
            proj_val = f"'{project_id}'" if project_id else "NULL"
            execute_sql(
                f"INSERT INTO {fq('metric_view_definitions')} VALUES "
                f"('{defn_id}', '{mv_name}', '{source}', '{json_str}', '', "
                f"'{status}', '{error_str}', NULL, '{now}', NULL, 1, NULL, {proj_val}, "
                f"{cx['complexity_score']}, '{cx['complexity_level']}', NULL, NULL)"
            )
            stats[status] += 1
            stats["generated"] += 1
            per_definition_results.append({
                "name": mv_name,
                "source": source,
                "status": status,
                "validation_errors": errors if errors else None,
                "complexity": cx.get("complexity_level"),
            })

        # Post-generation: LLM coverage check
        coverage = None
        if stats["validated"] > 0 and questions:
            task["stage"] = "checking_coverage"
            try:
                view_summaries = []
                for defn in definitions:
                    measures = [m.get("name", "") for m in defn.get("measures", [])]
                    view_summaries.append(f"- {defn.get('name', '')}: measures=[{', '.join(measures)}]")
                views_block = "\n".join(view_summaries)
                q_block = "\n".join(f"  {i+1}. {q}" for i, q in enumerate(questions))
                cov_prompt = f"""You are evaluating metric view coverage. For each business question, determine if it is COVERED (answerable by the generated metric views) or NOT_COVERED.

GENERATED METRIC VIEWS:
{views_block}

BUSINESS QUESTIONS:
{q_block}

Return ONLY a JSON object: {{"covered": [<1-based question indices>], "not_covered": [<1-based question indices>]}}"""
                cov_escaped = cov_prompt.replace("'", "''")
                cov_rows = execute_sql(f"SELECT AI_QUERY('{model}', '{cov_escaped}') as response", timeout=60)
                cov_resp = cov_rows[0]["response"] if cov_rows else ""
                coverage = _parse_single_json_safe(cov_resp)
                if coverage:
                    stats["coverage"] = coverage
            except Exception as exc:
                logger.warning("Coverage check failed: %s", exc)

        stats["definitions"] = per_definition_results
        task.update({"status": "done", "stage": "done", "result": stats})

    except Exception as e:
        logger.error("Semantic layer generation error: %s", e, exc_info=True)
        task.update({"status": "error", "error": str(e)})


@app.post("/api/semantic-layer/generate")
def start_sl_generation(req: SemanticGenerateRequest):
    """Start in-app metric view generation as a background task."""
    wh = os.environ.get("WAREHOUSE_ID", "")
    if not wh:
        raise HTTPException(500, detail="WAREHOUSE_ID not configured")
    if not req.tables:
        raise HTTPException(400, detail="No tables selected")
    if not req.questions:
        raise HTTPException(400, detail="No questions provided")

    cat = req.catalog_name or CATALOG
    sch = req.schema_name or SCHEMA
    task_id = str(_uuid.uuid4())[:12]
    _sl_tasks[task_id] = {
        "status": "running",
        "stage": "starting",
        "created": time.time(),
    }

    threading.Thread(
        target=_run_sl_generation,
        args=(
            task_id,
            req.tables,
            req.questions,
            cat,
            sch,
            req.model_endpoint,
            req.project_id,
            req.mode,
            req.business_context,
            req.profile_id,
        ),
        daemon=True,
    ).start()

    cutoff = time.time() - 1800
    for tid in list(_sl_tasks):
        if _sl_tasks.get(tid, {}).get("created", 0) < cutoff:
            _sl_tasks.pop(tid, None)

    return {"task_id": task_id}


@app.get("/api/semantic-layer/generate/{task_id}")
def poll_sl_generation(task_id: str):
    task = _sl_tasks.get(task_id)
    if not task:
        raise HTTPException(404, detail="Task not found")
    return task


# ---------------------------------------------------------------------------
# Metric-view per-definition actions (retry / improve / create)
# ---------------------------------------------------------------------------

_DEFAULT_MODEL = _LLM_MODEL


def _fetch_definition(definition_id: str) -> dict:
    """Load a single metric_view_definitions row by ID."""
    rows = execute_sql(
        f"SELECT * FROM {fq('metric_view_definitions')} "
        f"WHERE definition_id = '{definition_id}'"
    )
    if not rows:
        raise HTTPException(404, detail="Definition not found")
    return rows[0]


def _cat_sch_from_source(source: str) -> tuple[str, str]:
    """Extract catalog/schema from an FQ source table, falling back to app defaults."""
    parts = source.split(".") if source else []
    if len(parts) >= 3:
        return parts[0], parts[1]
    return CATALOG, SCHEMA


def _parse_single_json(text: str) -> dict:
    """Extract a single JSON object from an AI response."""
    text = re.sub(r"^```(?:json)?\s*", "", text.strip())
    text = re.sub(r"\s*```$", "", text)
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1:
        raise ValueError("No JSON object found in AI response")
    return json.loads(text[start : end + 1])


def _validate_definition(defn: dict, source: str) -> tuple[str, str]:
    """Two-tier validation: structural then expression. Returns (status, errors_str)."""
    for item_type in ("dimensions", "measures"):
        for item in defn.get(item_type, []):
            if item.get("expr"):
                item["expr"] = _autofix_expr(item["expr"])
    if defn.get("filter"):
        defn["filter"] = _autofix_expr(defn["filter"])
    defn = _normalize_joins(defn)
    defn = _fix_join_alias_refs(defn)
    errors = _validate_definition_structure(defn)
    if not errors:
        defn_joins = defn.get("joins", [])
        for item_type in ("dimensions", "measures"):
            for item in defn.get(item_type, []):
                expr = item.get("expr", "")
                if source and expr:
                    err, fixed = _validate_expr(expr, source, defn_joins)
                    if err:
                        errors.append(f"{item_type} {item.get('name', '')}: {err}")
                    elif fixed != expr:
                        item["expr"] = fixed
    status = "validated" if not errors else "failed"
    return status, "; ".join(errors)


def _update_definition_row(definition_id: str, defn: dict, status: str, errors: str):
    """Create a new version row and mark the old one as superseded."""
    from datetime import datetime as _dt

    json_str = json.dumps(defn).replace("'", "''")
    error_esc = errors.replace("'", "''")

    rows = execute_sql(
        f"SELECT version, metric_view_name, source_table, source_questions, project_id "
        f"FROM {fq('metric_view_definitions')} WHERE definition_id = '{definition_id}'"
    )
    old_version = int(rows[0].get("version") or 1) if rows else 1
    mv_name = rows[0].get("metric_view_name", "") if rows else ""
    source_table = rows[0].get("source_table", defn.get("source", "")) if rows else ""
    source_qs = rows[0].get("source_questions", "") if rows else ""
    proj_id = rows[0].get("project_id") if rows else None

    execute_sql(
        f"UPDATE {fq('metric_view_definitions')} SET status = 'superseded' "
        f"WHERE definition_id = '{definition_id}'"
    )

    new_id = str(_uuid.uuid4())
    new_version = old_version + 1
    now = _dt.utcnow().isoformat()
    proj_val = f"'{proj_id}'" if proj_id else "NULL"
    cx = _score_definition_complexity(defn)
    execute_sql(
        f"INSERT INTO {fq('metric_view_definitions')} VALUES ("
        f"'{new_id}', '{mv_name}', '{source_table}', '{json_str}', '{source_qs}', "
        f"'{status}', '{error_esc}', NULL, '{now}', NULL, {new_version}, '{definition_id}', {proj_val}, "
        f"{cx['complexity_score']}, '{cx['complexity_level']}', NULL, NULL)"
    )
    return new_id


def _yaml_esc(s: str) -> str:
    """Escape a string for embedding inside YAML double-quotes."""
    return s.replace("\\", "\\\\").replace('"', '\\"')


def _definition_to_yaml(defn: dict) -> str:
    """Convert a JSON definition to the YAML body for CREATE VIEW WITH METRICS.

    Uses manual construction instead of yaml.dump to guarantee SQL single
    quotes inside CASE/FILTER expressions survive the round-trip.
    """
    source = defn.get("source", "")
    if not source:
        raise ValueError("Metric view definition missing 'source' table")
    lines = ['version: "1.1"', f'source: {source}']
    if defn.get("comment"):
        lines.append(f'comment: "{_yaml_esc(defn["comment"])}"')
    if defn.get("filter"):
        lines.append(f'filter: "{_yaml_esc(defn["filter"])}"')
    lines.append("dimensions:")
    for i, d in enumerate(defn.get("dimensions", [])):
        name = d.get("name", f"dim_{i}")
        expr = d.get("expr", name)
        lines.append(f'  - name: "{_yaml_esc(name)}"')
        lines.append(f'    expr: "{_yaml_esc(expr)}"')
        if d.get("comment"):
            lines.append(f'    comment: "{_yaml_esc(d["comment"])}"')
    lines.append("measures:")
    for i, m in enumerate(defn.get("measures", [])):
        name = m.get("name", f"measure_{i}")
        expr = m.get("expr", "COUNT(*)")
        lines.append(f'  - name: "{_yaml_esc(name)}"')
        lines.append(f'    expr: "{_yaml_esc(expr)}"')
        if m.get("comment"):
            lines.append(f'    comment: "{_yaml_esc(m["comment"])}"')
        if m.get("window"):
            w = m["window"]
            if isinstance(w, str):
                try:
                    w = json.loads(w)
                except (json.JSONDecodeError, TypeError):
                    w = None
            if isinstance(w, dict) and (w.get("order_by") or w.get("range") or w.get("rows")):
                lines.append("    window:")
                if w.get("order_by"):
                    lines.append(f'      order_by: "{_yaml_esc(w["order_by"])}"')
                if w.get("range"):
                    lines.append(f'      range: "{_yaml_esc(w["range"])}"')
                if w.get("rows"):
                    lines.append(f'      rows: "{_yaml_esc(w["rows"])}"')
    if defn.get("joins"):
        lines.append("joins:")
        for j in defn["joins"]:
            lines.append(f'  - name: "{_yaml_esc(j.get("name", ""))}"')
            lines.append(f'    source: {j.get("source", "")}')
            if j.get("on"):
                lines.append(f'    on: "{_yaml_esc(j["on"])}"')
    return "\n".join(lines) + "\n"


_agent_ref_cache: dict[str, dict] = {}

def _load_agent_reference(name: str, sections: list[str] | None = None) -> str:
    """Load a JSON agent reference file and return selected sections as prompt text."""
    if name not in _agent_ref_cache:
        candidates = [
            os.path.join(os.path.dirname(__file__), "..", "configurations", "agent_references", name),
            os.path.join(os.path.dirname(__file__), "configurations", "agent_references", name),
            os.path.join("configurations", "agent_references", name),
        ]
        for p in candidates:
            p = os.path.normpath(p)
            if os.path.isfile(p):
                with open(p) as f:
                    _agent_ref_cache[name] = json.load(f)
                break
        else:
            _agent_ref_cache[name] = {}

    ref = _agent_ref_cache[name]
    if not ref:
        return ""
    keys = sections or list(ref.keys())
    parts = []
    for k in keys:
        val = ref.get(k)
        if val is None:
            continue
        if isinstance(val, list):
            parts.append(f"\n### {k}")
            for item in val:
                parts.append(f"- {item}" if isinstance(item, str) else f"- {json.dumps(item)}")
        elif isinstance(val, str):
            parts.append(f"\n### {k}\n{val}")
        else:
            parts.append(f"\n### {k}\n{json.dumps(val, indent=2)}")
    return "\n".join(parts)


@app.post("/api/semantic-layer/definitions/{definition_id}/retry")
def retry_definition(definition_id: str):
    """Re-attempt a failed definition by asking AI to fix validation errors."""
    row = _fetch_definition(definition_id)
    version = int(row.get("version") or 1)
    if version >= 3:
        raise HTTPException(
            409,
            detail=f"Max retries reached (v{version}). Consider re-generating with different questions or tables.",
        )

    defn = (
        json.loads(row["json_definition"])
        if isinstance(row["json_definition"], str)
        else row["json_definition"]
    )
    source = defn.get("source", row.get("source_table", ""))
    validation_errors = row.get("validation_errors", "")

    # Fetch actual available columns so the AI knows what exists
    available_cols = ""
    parts = source.split(".")
    if len(parts) == 3:
        try:
            col_rows = execute_sql(
                f"SELECT column_name, data_type FROM `{parts[0]}`.information_schema.columns "
                f"WHERE table_catalog = '{parts[0]}' AND table_schema = '{parts[1]}' AND table_name = '{parts[2]}'"
            )
            available_cols = ", ".join(
                f"{r['column_name']} ({r.get('data_type', '')})" for r in col_rows
            )
        except Exception:
            pass

    src_cat, src_sch = _cat_sch_from_source(source)
    context = _build_sl_context([source], src_cat, src_sch)
    ref_rules = _load_agent_reference("metric_view_reference.json", ["yaml_syntax_rules", "anti_patterns"])
    prompt = f"""You are fixing a metric view definition that has SQL errors.

ORIGINAL DEFINITION:
{json.dumps(defn, indent=2)}

ERRORS:
{validation_errors}

AVAILABLE COLUMNS in {source}:
{available_cols}

TABLE METADATA:
{context}

REFERENCE: SYNTAX RULES & ANTI-PATTERNS (follow these strictly)
{ref_rules}

Fix the SQL expressions. Rules:
- Only reference columns listed in AVAILABLE COLUMNS above
- Use standard Spark SQL functions (DATEDIFF, ROUND, ABS, etc.) -- these are fine as functions but are NOT columns
- DATE_TRUNC requires a quoted interval: DATE_TRUNC('MONTH', col) not DATE_TRUNC(MONTH, col)
- If a needed column doesn't exist, remove that measure/dimension rather than inventing columns
- Every metric view must keep at least one measure and one dimension

OUTPUT: Return ONLY the corrected JSON definition (single object, not array)."""

    escaped = prompt.replace("'", "''")
    rows = execute_sql(
        f"SELECT AI_QUERY('{_DEFAULT_MODEL}', '{escaped}') as response", timeout=180
    )
    response = rows[0]["response"] if rows else ""
    logger.info(
        "Retry AI response (first 500 chars): %s",
        response[:500] if response else "<empty>",
    )

    new_defn = _parse_single_json(response)
    new_defn.setdefault("source", source)
    new_defn.setdefault("name", defn.get("name", ""))
    status, errs = _validate_definition(new_defn, new_defn.get("source", source))
    new_id = _update_definition_row(definition_id, new_defn, status, errs)

    return {
        "definition_id": new_id,
        "parent_id": definition_id,
        "status": status,
        "validation_errors": errs,
    }


class PutDefinitionRequest(BaseModel):
    json_definition: str


class CreateDefinitionRequest(BaseModel):
    target_catalog: str
    target_schema: str


class DropDefinitionRequest(BaseModel):
    target_catalog: str
    target_schema: str


class SuggestFixRequest(BaseModel):
    error_message: str


@app.put("/api/semantic-layer/definitions/{definition_id}")
def update_definition(definition_id: str, req: PutDefinitionRequest):
    """Save manual edits to a metric view definition's JSON."""
    _ensure_semantic_layer_tables()
    try:
        defn = json.loads(req.json_definition)
    except json.JSONDecodeError as e:
        raise HTTPException(400, detail=f"Invalid JSON: {e}")
    source = defn.get("source", "")
    status, errs = _validate_definition(defn, source) if source else ("pending", "")
    json_esc = json.dumps(defn).replace("'", "''")
    err_esc = errs.replace("'", "''")
    execute_sql(
        f"UPDATE {fq('metric_view_definitions')} "
        f"SET json_definition = '{json_esc}', status = '{status}', "
        f"validation_errors = '{err_esc}' "
        f"WHERE definition_id = '{definition_id}'"
    )
    return {"definition_id": definition_id, "status": status, "validation_errors": errs}


@app.post("/api/semantic-layer/definitions/{definition_id}/create")
def create_metric_view(definition_id: str, req: CreateDefinitionRequest):
    """Deploy a validated definition as a real UC metric view."""
    _ensure_semantic_layer_tables()
    row = _fetch_definition(definition_id)
    defn = json.loads(row["json_definition"]) if isinstance(row["json_definition"], str) else row["json_definition"]
    mv_name = defn.get("name") or row.get("metric_view_name", "")
    if not mv_name:
        raise HTTPException(400, detail="Definition has no metric view name")
    fq_mv = f"`{req.target_catalog}`.`{req.target_schema}`.`{mv_name}`"
    for item_type in ("dimensions", "measures"):
        for item in defn.get(item_type, []):
            if item.get("expr"):
                item["expr"] = _autofix_expr(item["expr"])
    if defn.get("filter"):
        defn["filter"] = _autofix_expr(defn["filter"])
    yaml_body = _definition_to_yaml(defn)
    sql = f"CREATE OR REPLACE VIEW {fq_mv}\nWITH METRICS LANGUAGE YAML AS $$\n{yaml_body}$$"
    try:
        execute_sql(sql, timeout=60)
    except Exception as e:
        raise HTTPException(400, detail=str(e))
    # Tag as draft for governance
    try:
        execute_sql(f"ALTER VIEW {fq_mv} SET TBLPROPERTIES ('certification_status' = 'draft', 'generated_by' = 'dbxmetagen')", timeout=15)
    except Exception:
        pass
    execute_sql(
        f"UPDATE {fq('metric_view_definitions')} "
        f"SET status = 'applied', applied_at = current_timestamp(), "
        f"deployed_catalog = '{req.target_catalog}', deployed_schema = '{req.target_schema}' "
        f"WHERE definition_id = '{definition_id}'"
    )
    return {"definition_id": definition_id, "status": "applied", "metric_view": fq_mv}


class CertifyRequest(BaseModel):
    target_catalog: str
    target_schema: str
    status: str = "certified"


@app.post("/api/semantic-layer/definitions/{definition_id}/certify")
def certify_metric_view(definition_id: str, req: CertifyRequest):
    """Promote a deployed metric view from draft to certified (or back)."""
    _ensure_semantic_layer_tables()
    row = _fetch_definition(definition_id)
    defn = json.loads(row["json_definition"]) if isinstance(row["json_definition"], str) else row["json_definition"]
    mv_name = defn.get("name") or row.get("metric_view_name", "")
    if not mv_name:
        raise HTTPException(400, detail="Definition has no metric view name")
    if row.get("status") != "applied":
        raise HTTPException(400, detail="Only applied metric views can be certified")
    fq_mv = f"`{req.target_catalog}`.`{req.target_schema}`.`{mv_name}`"
    cert_status = req.status.replace("'", "''")
    try:
        execute_sql(f"ALTER VIEW {fq_mv} SET TBLPROPERTIES ('certification_status' = '{cert_status}')", timeout=15)
    except Exception as e:
        raise HTTPException(400, detail=str(e))
    return {"definition_id": definition_id, "metric_view": fq_mv, "certification_status": cert_status}


@app.get("/api/semantic-layer/export-sql")
def export_metric_views_sql(catalog: Optional[str] = None, schema: Optional[str] = None):
    """Generate a .sql file with CREATE VIEW WITH METRICS statements for all applied definitions."""
    _ensure_semantic_layer_tables()
    rows = execute_sql(
        f"SELECT * FROM {fq('metric_view_definitions')} WHERE status = 'applied'"
    )
    if not rows:
        raise HTTPException(404, detail="No applied metric view definitions found")
    default_cat = catalog or CATALOG
    default_sch = schema or SCHEMA
    statements = []
    for row in rows:
        defn = json.loads(row["json_definition"]) if isinstance(row["json_definition"], str) else row["json_definition"]
        mv_name = defn.get("name") or row.get("metric_view_name", "")
        if not mv_name:
            continue
        mv_cat = row.get("deployed_catalog") or default_cat
        mv_sch = row.get("deployed_schema") or default_sch
        fq_mv = f"`{mv_cat}`.`{mv_sch}`.`{mv_name}`"
        yaml_body = _definition_to_yaml(defn)
        statements.append(f"CREATE OR REPLACE VIEW {fq_mv}\nWITH METRICS LANGUAGE YAML AS $$\n{yaml_body}$$")
    if not statements:
        raise HTTPException(404, detail="No valid definitions to export")
    body = ";\n\n".join(statements) + ";\n"
    return Response(
        content=body,
        media_type="text/sql",
        headers={"Content-Disposition": "attachment; filename=metric_views.sql"},
    )


@app.post("/api/semantic-layer/definitions/{definition_id}/improve")
def improve_definition(definition_id: str):
    """Ask AI to improve an existing validated/applied metric view definition."""
    _ensure_semantic_layer_tables()
    row = _fetch_definition(definition_id)
    defn = json.loads(row["json_definition"]) if isinstance(row["json_definition"], str) else row["json_definition"]
    source = defn.get("source", row.get("source_table", ""))
    src_cat, src_sch = _cat_sch_from_source(source) if source else (CATALOG, SCHEMA)
    context = _build_sl_context([source], src_cat, src_sch) if source else ""
    ref_rules = _load_agent_reference("metric_view_reference.json", ["measure_patterns", "yaml_syntax_rules"])

    prompt = f"""You are improving a metric view definition. Make it more comprehensive and useful.

CURRENT DEFINITION:
{json.dumps(defn, indent=2)}

TABLE METADATA:
{context}

REFERENCE: BEST PRACTICES
{ref_rules}

Improvements to make:
- Add missing measures that would be useful (ratios, rates, conditional aggregates)
- Improve dimension coverage (time-based truncations, categorizations)
- Ensure measure/dimension names are business-friendly
- Add FILTER-based conditional measures where relevant
- Keep existing measures/dimensions unless they are wrong
- Every metric view must have at least one measure and one dimension

OUTPUT: Return ONLY the improved JSON definition (single object, not array)."""

    escaped = prompt.replace("'", "''")
    rows = execute_sql(
        f"SELECT AI_QUERY('{_DEFAULT_MODEL}', '{escaped}') as response", timeout=180
    )
    response = rows[0]["response"] if rows else ""
    new_defn = _parse_single_json(response)
    new_defn.setdefault("source", source)
    new_defn.setdefault("name", defn.get("name", ""))
    status, errs = _validate_definition(new_defn, new_defn.get("source", source))
    new_id = _update_definition_row(definition_id, new_defn, status, errs)
    return {"definition_id": new_id, "parent_id": definition_id, "status": status, "validation_errors": errs}


@app.post("/api/semantic-layer/definitions/{definition_id}/drop")
def drop_metric_view(definition_id: str, req: DropDefinitionRequest):
    """Drop a deployed metric view from Unity Catalog."""
    _ensure_semantic_layer_tables()
    row = _fetch_definition(definition_id)
    defn = json.loads(row["json_definition"]) if isinstance(row["json_definition"], str) else row["json_definition"]
    mv_name = defn.get("name") or row.get("metric_view_name", "")
    if not mv_name:
        raise HTTPException(400, detail="Definition has no metric view name")
    fq_mv = f"`{req.target_catalog}`.`{req.target_schema}`.`{mv_name}`"
    try:
        execute_sql(f"DROP VIEW IF EXISTS {fq_mv}", timeout=30)
    except Exception as e:
        raise HTTPException(400, detail=f"Failed to drop view: {e}")
    execute_sql(
        f"UPDATE {fq('metric_view_definitions')} "
        f"SET status = 'validated', applied_at = NULL "
        f"WHERE definition_id = '{definition_id}'"
    )
    return {"definition_id": definition_id, "status": "validated", "dropped": fq_mv}


@app.post("/api/semantic-layer/definitions/{definition_id}/suggest-fix")
def suggest_fix(definition_id: str, req: SuggestFixRequest):
    """Ask AI to suggest a fix for a definition that failed to create."""
    _ensure_semantic_layer_tables()
    row = _fetch_definition(definition_id)
    defn = json.loads(row["json_definition"]) if isinstance(row["json_definition"], str) else row["json_definition"]
    source = defn.get("source", row.get("source_table", ""))

    prompt = f"""A metric view definition failed to deploy with this error:

ERROR:
{req.error_message}

DEFINITION:
{json.dumps(defn, indent=2)}

Fix the definition so it deploys successfully. Rules:
- DATE_TRUNC requires a quoted interval: DATE_TRUNC('MONTH', col)
- Only use columns that exist in the source table
- All string literals must be single-quoted
- Output ONLY the corrected JSON definition (single object, not array)."""

    escaped = prompt.replace("'", "''")
    rows = execute_sql(
        f"SELECT AI_QUERY('{_DEFAULT_MODEL}', '{escaped}') as response", timeout=180
    )
    response = rows[0]["response"] if rows else ""
    try:
        suggested = _parse_single_json(response)
        return {"suggested_json": json.dumps(suggested, indent=2)}
    except Exception:
        return {"suggested_json": response}


# ---------------------------------------------------------------------------
# Genie Builder endpoints
# ---------------------------------------------------------------------------


@app.post("/api/genie/generate-questions")
def genie_generate_questions(req: SuggestQuestionsRequest):
    """Generate business-user-friendly questions for the selected tables using an LLM."""
    wh = os.environ.get("WAREHOUSE_ID", "")
    if not wh:
        raise HTTPException(500, detail="WAREHOUSE_ID not configured")
    from dbxmetagen.genie.context import GenieContextAssembler
    from langchain_community.chat_models import ChatDatabricks

    ws = _get_effective_client()
    assembler = GenieContextAssembler(ws, wh, CATALOG, SCHEMA)
    ctx = assembler.assemble(
        req.table_identifiers,
        questions=None,
        metric_view_names=req.metric_view_names,
    )

    biz_ctx_block = ""
    if req.business_context and req.business_context.strip():
        biz_ctx_block = f"\nBUSINESS CONTEXT (provided by the user -- this defines the semantic frame for all analysis):\n{req.business_context.strip()}\n"

    if req.purpose == "metric_views":
        prompt = f"""You are a business intelligence strategist. Your task is to generate questions that would drive the creation of reusable KPI metric views.
{biz_ctx_block}
Below is metadata about available tables and their business context.

{ctx.get('context_text', '')}

Generate exactly {req.count} questions that a BUSINESS LEADER would ask to track organizational performance. Rules:
- Ground every question in the ENTITY TYPES and RELATIONSHIPS described in the metadata (e.g., if the data contains Encounters, Patients, Providers -- ask about patient visit patterns, provider utilization, encounter outcomes)
- Every question MUST be answerable using ONLY the tables and columns described above -- do not invent data that isn't present
- Use the domain and subdomain classifications to frame questions in the right business context
- Think about what a CEO, CFO, VP, or department head would ask in a weekly review meeting
- Focus on measurable outcomes: revenue growth, cost efficiency, customer satisfaction, operational throughput, quality metrics
- Frame questions around time-based trends ("How has X changed over the past quarter?"), comparisons ("Which segment leads in Y?"), and thresholds ("Are we meeting our Z target?")
- Prefer questions that naturally decompose into a measure (SUM, AVG, COUNT) and dimensions (time, category, region)
- Do NOT mention column names, table names, or SQL concepts -- use business language only
- A business user who USES the data but doesn't know the data model should understand every question

Return ONLY a JSON array of strings, no other text."""
    else:
        prompt = f"""You are a data analyst helping business users explore their data through a natural language SQL interface (Databricks Genie).
{biz_ctx_block}
Below is metadata about the available tables, columns, relationships, and metric views.

{ctx.get('context_text', '')}

Generate exactly {req.count} questions that a BUSINESS USER would naturally ask. Rules:
- Ground every question in the ENTITY TYPES and RELATIONSHIPS described in the metadata -- if the data models Patients, Orders, Claims, etc., ask about those specific business concepts
- Every question MUST be answerable using ONLY the tables and columns described above -- do not invent data that isn't present
- Use the domain and subdomain classifications to frame questions in the right business context
- Questions should be outcome-oriented and insight-driven (e.g. "What are the top performing regions by revenue this quarter?")
- Do NOT reference column names, table names, or technical schema details
- Focus on trends, comparisons, rankings, anomalies, and KPIs
- Vary the question types: aggregations, time-series trends, top-N, filters, comparisons
- A business user who USES the data but doesn't know the data model should understand every question

Return ONLY a JSON array of strings, no other text."""

    llm = ChatDatabricks(endpoint=req.model_endpoint, temperature=0.7, max_tokens=2048)
    response = llm.invoke(prompt)
    content = response.content.strip()
    if content.startswith("```"):
        content = content.split("\n", 1)[1] if "\n" in content else content[3:]
        content = content.rsplit("```", 1)[0]
    questions = json.loads(content)
    return {"questions": questions[:req.count]}


@app.post("/api/genie/generate")
def genie_generate(req: GenieGenerateRequest):
    """Start the Genie builder agent as a background task."""
    wh = os.environ.get("WAREHOUSE_ID", "")
    if not wh:
        raise HTTPException(500, detail="WAREHOUSE_ID not configured")

    task_id = str(_uuid.uuid4())[:12]
    started_at = time.time()
    _genie_tasks[task_id] = {
        "status": "running",
        "stage": "starting",
        "created": started_at,
        "started_at": started_at,
        "round": 0,
    }

    # Total wall-clock backstop: context assembly ~60s + LLM call ~300s + SQL validation ~60s + recovery ~300s
    _MONITOR_WALL_TIMEOUT = 600

    def _run():
        try:
            from dbxmetagen.genie.context import GenieContextAssembler
            from dbxmetagen.genie.agent import run_genie_agent

            ws = get_workspace_client()
            progress_q: queue.Queue = queue.Queue()

            _genie_tasks[task_id]["stage"] = "gathering_context"
            ctx_t0 = time.time()
            assembler = GenieContextAssembler(ws, wh, CATALOG, SCHEMA)
            ctx = assembler.assemble(
                req.table_identifiers,
                req.questions or None,
                metric_view_names=req.metric_view_names,
            )
            ctx_elapsed = round(time.time() - ctx_t0, 1)
            ctx_len = len(ctx.get("context_text", ""))
            logger.info(
                "Genie context assembly: %.1fs, %d tables, %d join_specs, %d context_text chars",
                ctx_elapsed, len(req.table_identifiers), len(ctx.get("join_specs", [])), ctx_len,
            )
            if ctx_len < 100:
                raise ValueError(
                    "No metadata found for the selected tables. "
                    "Ensure the knowledge base pipeline has been run (build_knowledge_base) "
                    "and the correct catalog/schema is configured."
                )

            if req.business_context and req.business_context.strip():
                biz_block = (
                    "\n\nBUSINESS CONTEXT (provided by the user -- this defines the semantic frame for all analysis):\n"
                    + req.business_context.strip()
                    + "\n"
                )
                ctx["context_text"] = biz_block + ctx.get("context_text", "")

            if req.kpi_names:
                _ensure_kpi_table()
                kpi_rows = execute_sql(f"SELECT name, description, formula, domain FROM {fq('kpi_definitions')}")
                sel = {n.lower() for n in req.kpi_names}
                matched = [r for r in kpi_rows if r.get("name", "").lower() in sel]
                if matched:
                    kpi_block = "\n\nBUSINESS KPIs (use these to inform measures, expressions, and sample questions):\n"
                    for k in matched:
                        kpi_block += f"- {k['name']}: {k.get('description', '')} | Formula: {k.get('formula', 'N/A')}\n"
                    ctx["context_text"] = ctx.get("context_text", "") + kpi_block

            _genie_tasks[task_id]["stage"] = "agent_running"

            def _monitor_progress():
                while True:
                    # Total wall-clock backstop
                    remaining = _MONITOR_WALL_TIMEOUT - (time.time() - started_at)
                    if remaining <= 0:
                        elapsed = round(time.time() - started_at)
                        _genie_tasks[task_id].update({
                            "status": "error",
                            "error": (
                                f"Generation timed out after {elapsed}s. "
                                "Try selecting fewer tables or simplifying the request."
                            ),
                            "elapsed_seconds": elapsed,
                            "rounds_completed": 0,
                        })
                        return
                    try:
                        event = progress_q.get(timeout=min(remaining, 30))
                    except queue.Empty:
                        continue

                    stage = event.get("stage", "running")

                    if stage == "done":
                        _genie_tasks[task_id].update({
                            "status": "done",
                            "stage": "done",
                            "result": event.get("result"),
                            "warnings": event.get("warnings"),
                            "elapsed_seconds": event.get("elapsed_seconds"),
                            "rounds_completed": event.get("rounds_completed"),
                        })
                        return

                    if stage == "error":
                        _genie_tasks[task_id].update({
                            "status": "error",
                            "error": event.get("message", "Unknown error"),
                            "elapsed_seconds": event.get("elapsed_seconds"),
                            "rounds_completed": event.get("rounds_completed"),
                        })
                        return

                    _genie_tasks[task_id]["stage"] = stage
                    if "round" in event:
                        _genie_tasks[task_id]["round"] = event["round"]

            monitor = threading.Thread(target=_monitor_progress, daemon=True)
            monitor.start()

            run_genie_agent(
                ws, wh, ctx, progress_q,
                model_endpoint=req.model_endpoint,
                refinement_feedback=req.refinement_feedback,
                prior_result=req.prior_result,
            )
        except Exception as e:
            logger.error("Genie builder error: %s", e, exc_info=True)
            elapsed = round(time.time() - started_at)
            task = _genie_tasks.get(task_id)
            if not task or task.get("status") == "error":
                return  # task already cleaned up or monitor already recorded the error
            rnd = max(task.get("round", 0), task.get("rounds_completed", 0))
            task.update({
                "status": "error",
                "error": str(e),
                "elapsed_seconds": elapsed,
                "rounds_completed": rnd,
            })

    threading.Thread(target=_run, daemon=True).start()

    # Clean up old tasks (> 30 min)
    cutoff = time.time() - 1800
    for tid in list(_genie_tasks):
        if _genie_tasks.get(tid, {}).get("created", 0) < cutoff:
            _genie_tasks.pop(tid, None)

    return {"task_id": task_id}


@app.get("/api/genie/tasks/{task_id}")
def genie_task_status(task_id: str):
    """Poll status of a Genie builder task."""
    cutoff = time.time() - 1800
    for tid in list(_genie_tasks):
        if tid != task_id and _genie_tasks.get(tid, {}).get("created", 0) < cutoff:
            _genie_tasks.pop(tid, None)
    task = _genie_tasks.get(task_id)
    if not task:
        raise HTTPException(404, detail="Task not found")
    resp = dict(task)
    if resp.get("status") == "running" and "started_at" in resp:
        resp["elapsed_seconds"] = round(time.time() - resp["started_at"])
    resp.pop("started_at", None)
    return resp


from dbxmetagen.genie.schema import build_serialized_space


def _transform_to_genie_schema(raw: dict) -> dict:
    """Convert agent output into the Databricks Genie API format via Pydantic whitelist."""
    return build_serialized_space(raw)


def _strip_field(obj, field_name):
    """Recursively remove a field from all dicts in the structure."""
    if isinstance(obj, dict):
        obj.pop(field_name, None)
        for v in obj.values():
            _strip_field(v, field_name)
    elif isinstance(obj, list):
        for item in obj:
            _strip_field(item, field_name)


def _validate_serialized_space(ss: dict) -> list[str]:
    """Validate the transformed serialized_space before sending to Genie API."""
    errors = []
    if "data_sources" not in ss:
        errors.append("Missing required 'data_sources' section")
    else:
        ds = ss["data_sources"]
        if not isinstance(ds, dict):
            errors.append("'data_sources' must be a dict")
        elif not ds.get("tables") and not ds.get("metric_views"):
            errors.append("'data_sources' must have at least one table or metric_view")
        valid_ids = set()
        for i, tbl in enumerate(ds.get("tables", [])):
            if not tbl.get("identifier"):
                errors.append(f"data_sources.tables[{i}] missing 'identifier'")
            else:
                valid_ids.add(tbl["identifier"])
        for i, mv in enumerate(ds.get("metric_views", [])):
            if not mv.get("identifier"):
                errors.append(f"data_sources.metric_views[{i}] missing 'identifier'")
            else:
                valid_ids.add(mv["identifier"])
        # Join spec references must be in data_sources
        for i, j in enumerate(ss.get("instructions", {}).get("join_specs", [])):
            left_id = j.get("left", {}).get("identifier") if isinstance(j.get("left"), dict) else None
            right_id = j.get("right", {}).get("identifier") if isinstance(j.get("right"), dict) else None
            if left_id and left_id not in valid_ids:
                errors.append(f"join_specs[{i}].left.identifier '{left_id}' not in data_sources")
            if right_id and right_id not in valid_ids:
                errors.append(f"join_specs[{i}].right.identifier '{right_id}' not in data_sources")
    inst = ss.get("instructions", {})
    if not isinstance(inst, dict):
        errors.append("'instructions' must be a dict")
    return errors


def _collect_valid_identifiers(ss: dict) -> set[str]:
    """Collect all identifiers (full and short names) from data_sources."""
    ds = ss.get("data_sources", {})
    ids: set[str] = set()
    for t in ds.get("tables", []):
        ident = t.get("identifier", "")
        if ident:
            ids.add(ident)
            ids.add(ident.split(".")[-1])
    for m in ds.get("metric_views", []):
        ident = m.get("identifier", "")
        if ident:
            ids.add(ident)
            ids.add(ident.split(".")[-1])
    return ids


def _build_genie_from_clause(ss: dict) -> str | None:
    """Build a FROM clause with JOINs from data_sources + join_specs."""
    ds = ss.get("data_sources", {})
    table_ids = [t["identifier"] for t in ds.get("tables", []) if t.get("identifier")]
    table_ids += [m["identifier"] for m in ds.get("metric_views", []) if m.get("identifier")]
    if not table_ids:
        return None

    def _quote(ident: str) -> str:
        return ".".join(f"`{p}`" for p in ident.split("."))

    base = table_ids[0]
    base_alias = base.split(".")[-1]
    parts = [f"{_quote(base)} AS `{base_alias}`"]

    join_specs = ss.get("instructions", {}).get("join_specs", [])
    joined: set[str] = {base}
    for j in join_specs:
        left_id = j.get("left", {}).get("identifier", "")
        right_id = j.get("right", {}).get("identifier", "")
        join_sql = j.get("sql", [])
        if not left_id or not right_id or not join_sql:
            continue
        if left_id in joined and right_id not in joined:
            alias = right_id.split(".")[-1]
            cond = " AND ".join(join_sql)
            parts.append(f"LEFT JOIN {_quote(right_id)} AS `{alias}` ON {cond}")
            joined.add(right_id)
        elif right_id in joined and left_id not in joined:
            alias = left_id.split(".")[-1]
            cond = " AND ".join(join_sql)
            parts.append(f"LEFT JOIN {_quote(left_id)} AS `{alias}` ON {cond}")
            joined.add(left_id)

    for ident in table_ids:
        if ident not in joined:
            alias = ident.split(".")[-1]
            parts.append(f"LEFT JOIN {_quote(ident)} AS `{alias}` ON 1=1")

    return " ".join(parts)


def _validate_sql_expressions(ss: dict, warehouse_id: str) -> dict:
    """Dry-run example_question_sqls and strip broken ones.

    Snippets (measures/expressions/filters) are only warned about since they
    are SQL fragments that Genie embeds into its own query context.
    """
    inst = ss.get("instructions", {})

    example_sqls = inst.get("example_question_sqls", [])
    valid_examples = []
    for ex in example_sqls:
        sqls = ex.get("sql", [])
        sql = (sqls[0] if sqls else "").strip().rstrip(";")
        if not sql:
            continue
        try:
            execute_sql(f"{sql} LIMIT 0", warehouse_id=warehouse_id, timeout=15)
            valid_examples.append(ex)
        except Exception:
            logger.warning("Stripped invalid example_sql: %s", sql[:120])
    inst["example_question_sqls"] = valid_examples

    return ss


def _strip_out_of_scope_sql(ss: dict) -> dict:
    """Remove SQL entries that reference tables not in data_sources."""
    valid_ids = _collect_valid_identifiers(ss)
    if not valid_ids:
        return ss

    from dbxmetagen.genie.schema import _extract_table_refs_from_sql

    def _refs_ok(sql_list: list[str]) -> bool:
        for sql in sql_list:
            refs = _extract_table_refs_from_sql(sql)
            for ref in refs:
                short = ref.split(".")[-1]
                if ref not in valid_ids and short not in valid_ids:
                    return False
        return True

    inst = ss.get("instructions", {})
    examples = inst.get("example_question_sqls", [])
    inst["example_question_sqls"] = [
        ex for ex in examples if _refs_ok(ex.get("sql", []))
    ]

    snippets = inst.get("sql_snippets") or {}
    for category in ("measures", "expressions", "filters"):
        items = snippets.get(category, [])
        snippets[category] = [it for it in items if _refs_ok(it.get("sql", []))]

    # Filter join_specs: check table.column refs in join SQL
    join_specs = inst.get("join_specs", [])
    if join_specs:
        valid_short_lower = {v.split(".")[-1].lower() for v in valid_ids}
        valid_lower = {v.lower() for v in valid_ids} | valid_short_lower
        logger.warning("[join-diag] valid_short_lower: %s", sorted(valid_short_lower)[:20])
        valid_joins = []
        for js in join_specs:
            sql_list = js.get("sql", [])
            join_refs = set()
            for sql in sql_list:
                # Handle both `table`.`col` (backtick-quoted) and table.col (bare)
                join_refs.update(m.lower() for m in re.findall(r'`(\w+)`\.`?\w+`?', sql))
                join_refs.update(m.lower() for m in re.findall(r'\b(\w+)\.\w+', sql))
            if join_refs and all(ref in valid_lower for ref in join_refs):
                valid_joins.append(js)
            elif not join_refs:
                valid_joins.append(js)
            else:
                bad = join_refs - valid_lower
                left_id = js.get("left", {}).get("identifier", "?")
                right_id = js.get("right", {}).get("identifier", "?")
                logger.warning(
                    "[join-diag] STRIPPED join (%s <-> %s): extracted refs %s, bad refs %s, SQL %s",
                    left_id, right_id, join_refs, bad, sql_list,
                )
        inst["join_specs"] = valid_joins

    return ss


def _validate_data_sources_exist(ss: dict, warehouse_id: str) -> list[str]:
    """Run SELECT 1 FROM identifier LIMIT 1 for each data source; return list of errors."""
    errors = []
    ds = ss.get("data_sources", {})
    identifiers = [t.get("identifier") for t in ds.get("tables", []) if t.get("identifier")]
    identifiers += [m.get("identifier") for m in ds.get("metric_views", []) if m.get("identifier")]
    for ident in identifiers:
        quoted = ".".join(f"`{p}`" for p in ident.split("."))
        try:
            execute_sql(f"SELECT 1 FROM {quoted} LIMIT 1", warehouse_id=warehouse_id, timeout=15)
        except HTTPException as e:
            errors.append(f"data_sources identifier '{ident}': {e.detail}")
        except Exception as e:
            errors.append(f"data_sources identifier '{ident}': {e}")
    return errors


@app.post("/api/genie/create")
def genie_create(req: GenieCreateRequest):
    """Create or update a Genie space via the Databricks REST API."""
    # Capture prebuilt join identifier-pairs before Pydantic strips _prebuilt
    _raw_joins = (req.serialized_space.get("instructions", {}).get("join_specs")
                  or req.serialized_space.get("join_specs") or [])
    prebuilt_pairs: set[tuple[str, str]] = set()
    for j in _raw_joins:
        if j.get("_prebuilt"):
            l_id = j.get("left", {}).get("identifier", "")
            r_id = j.get("right", {}).get("identifier", "")
            if l_id and r_id:
                prebuilt_pairs.add(tuple(sorted([l_id.lower(), r_id.lower()])))
    transformed = _transform_to_genie_schema(req.serialized_space)

    _pre_strip_joins = transformed.get("instructions", {}).get("join_specs", [])
    logger.warning(
        "[join-diag] after build_serialized_space: %d joins. SQL samples: %s",
        len(_pre_strip_joins),
        [js.get("sql", [])[:1] for js in _pre_strip_joins[:3]],
    )

    validation_errors = _validate_serialized_space(transformed)
    if validation_errors:
        raise HTTPException(
            400, detail=f"Invalid serialized_space: {'; '.join(validation_errors)}"
        )

    ws = get_workspace_client()
    wh = req.warehouse_id or os.environ.get("WAREHOUSE_ID", "")
    if not wh:
        raise HTTPException(500, detail="WAREHOUSE_ID not configured")

    exist_errors = _validate_data_sources_exist(transformed, wh)
    if exist_errors:
        raise HTTPException(400, detail="Data source validation failed: " + "; ".join(exist_errors))

    transformed = _strip_out_of_scope_sql(transformed)

    _post_strip_joins = transformed.get("instructions", {}).get("join_specs", [])
    logger.warning(
        "[join-diag] after _strip_out_of_scope_sql: %d joins (was %d)",
        len(_post_strip_joins), len(_pre_strip_joins),
    )

    transformed = _validate_sql_expressions(transformed, wh)

    _inst = transformed.get("instructions", {})
    _snip = _inst.get("sql_snippets", {}) or {}
    logger.info(
        "Genie deploy payload: %d tables, %d MVs, %d joins, %d measures, %d filters, %d expressions, %d examples",
        len(transformed.get("data_sources", {}).get("tables", [])),
        len(transformed.get("data_sources", {}).get("metric_views", [])),
        len(_inst.get("join_specs", [])),
        len(_snip.get("measures", [])),
        len(_snip.get("filters", [])),
        len(_snip.get("expressions", [])),
        len(_inst.get("example_question_sqls", [])),
    )
    logger.debug("Genie serialized_space (first 2000 chars): %s", json.dumps(transformed)[:2000])

    def _do_genie_request(space_json):
        body = {
            "title": req.title,
            "warehouse_id": wh,
            "serialized_space": json.dumps(space_json),
        }
        if req.description:
            body["description"] = req.description
        if req.space_id:
            ws.api_client.do(
                "PATCH", f"/api/2.0/genie/spaces/{req.space_id}", body=body
            )
            return {"space_id": req.space_id, "title": req.title, "updated": True}
        else:
            resp = ws.api_client.do("POST", "/api/2.0/genie/spaces", body=body)
            return {
                "space_id": resp.get("space_id", resp.get("id")),
                "title": req.title,
                "updated": False,
            }

    deploy_warnings: list[str] = []

    _MAX_GENIE_RETRIES = 10
    last_err: Exception | None = None
    for attempt in range(_MAX_GENIE_RETRIES + 1):
        try:
            result = _do_genie_request(transformed)
            # Persist to tracking table
            try:
                _ensure_genie_tracking_table()
                space_id = result["space_id"]
                ds = transformed.get("data_sources", {})
                table_ids = [t.get("identifier") for t in ds.get("tables", []) if t.get("identifier")]
                table_ids += [m.get("identifier") for m in ds.get("metric_views", []) if m.get("identifier")]
                config_str = json.dumps(transformed).replace("'", "''")
                title_esc = req.title.replace("'", "''")
                arr_literal = ",".join(f"'{t}'" for t in table_ids)
                if result.get("updated"):
                    old_rows = execute_sql(
                        f"SELECT COALESCE(version, 1) as version FROM {fq('genie_spaces')} "
                        f"WHERE space_id = '{space_id}' AND COALESCE(status, 'active') = 'active' "
                        f"AND deleted_at IS NULL ORDER BY version DESC LIMIT 1",
                        timeout=15,
                    )
                    old_ver = int(old_rows[0]["version"]) if old_rows else 1
                    execute_sql(
                        f"UPDATE {fq('genie_spaces')} SET status = 'superseded', updated_at = current_timestamp() "
                        f"WHERE space_id = '{space_id}' AND COALESCE(status, 'active') = 'active'",
                        timeout=30,
                    )
                    execute_sql(
                        f"INSERT INTO {fq('genie_spaces')} "
                        f"(space_id, title, tables, config_json, version, status, parent_space_id, created_at, updated_at, deleted_at) VALUES "
                        f"('{space_id}', '{title_esc}', ARRAY({arr_literal}), "
                        f"'{config_str}', {old_ver + 1}, 'active', '{space_id}', current_timestamp(), current_timestamp(), NULL)",
                        timeout=30,
                    )
                else:
                    execute_sql(
                        f"INSERT INTO {fq('genie_spaces')} "
                        f"(space_id, title, tables, config_json, version, status, parent_space_id, created_at, updated_at, deleted_at) VALUES "
                        f"('{space_id}', '{title_esc}', ARRAY({arr_literal}), "
                        f"'{config_str}', 1, 'active', NULL, current_timestamp(), current_timestamp(), NULL)",
                        timeout=30,
                    )
                logger.info("Tracked genie space %s in genie_spaces table", space_id)
            except Exception as track_err:
                logger.warning("Failed to track genie space: %s", track_err)
            final_joins = transformed.get("instructions", {}).get("join_specs", [])
            final_tables = len(transformed.get("data_sources", {}).get("tables", []))
            final_mvs = len(transformed.get("data_sources", {}).get("metric_views", []))
            logger.info(
                "Genie deploy SUCCESS: space_id=%s, %d tables, %d MVs, %d joins survived, %d warnings",
                result.get("space_id"), final_tables, final_mvs, len(final_joins), len(deploy_warnings),
            )
            result["join_count"] = len(final_joins)
            result["table_count"] = final_tables
            result["mv_count"] = final_mvs
            # Read-back verification: confirm joins actually persisted
            persisted_join_count = None
            try:
                space_id = result.get("space_id")
                if space_id and final_joins:
                    rb = ws.api_client.do(
                        "GET",
                        f"/api/2.0/genie/spaces/{space_id}",
                        query={"include_serialized_space": "true"},
                    )
                    rb_ss = rb.get("serialized_space", "")
                    if isinstance(rb_ss, str) and rb_ss:
                        rb_parsed = json.loads(rb_ss)
                    else:
                        rb_parsed = rb_ss if isinstance(rb_ss, dict) else {}
                    rb_joins = (rb_parsed.get("instructions", {}).get("join_specs", [])
                                or rb_parsed.get("data_sources", {}).get("join_specs", []))
                    persisted_join_count = len(rb_joins)
                    logger.info(
                        "Genie read-back: %d joins persisted (sent %d)",
                        persisted_join_count, len(final_joins),
                    )
                    if persisted_join_count == 0 and len(final_joins) > 0:
                        deploy_warnings.append(
                            f"Joins did NOT persist: sent {len(final_joins)} but API returned 0. "
                            "The Genie API may have silently dropped them."
                        )
            except Exception as rb_err:
                logger.warning("Genie read-back failed: %s", rb_err)
            if persisted_join_count is not None:
                result["persisted_join_count"] = persisted_join_count
            if deploy_warnings:
                result["warnings"] = deploy_warnings
            return result
        except Exception as e:
            last_err = e
            err_str = str(e)
            m = re.search(r"Cannot find field: (\w+)", err_str)
            if m:
                bad_field = m.group(1)
                logger.warning(
                    "Attempt %d: stripping unknown field '%s'", attempt + 1, bad_field
                )
                target = transformed.get("instructions", {}).get("sql_snippets", {})
                if target:
                    _strip_field(target, bad_field)
                else:
                    _strip_field(transformed, bad_field)
                deploy_warnings.append(f"Stripped unknown API field: {bad_field}")
                continue
            if "parse export proto" in err_str.lower() or "failed to parse" in err_str.lower():
                logger.warning("Attempt %d proto error (full): %s", attempt + 1, err_str[:500])
                inst = transformed.get("instructions", {})
                snippets = inst.get("sql_snippets", {})
                examples = inst.get("example_question_sqls", [])
                join_specs = inst.get("join_specs", [])
                # Phase 1: strip ALL non-empty snippet categories in one pass
                stripped = False
                for cat in ("expressions", "filters", "measures"):
                    if snippets.get(cat):
                        removed_items = snippets.pop(cat)
                        logger.warning("Attempt %d: stripped %d %s due to proto parse error", attempt + 1, len(removed_items), cat)
                        deploy_warnings.append(f"Stripped {len(removed_items)} {cat} (proto error)")
                        stripped = True
                if stripped:
                    if not snippets:
                        inst.pop("sql_snippets", None)
                    continue
                # Phase 2: strip example_sql from the end
                if examples:
                    removed_ex = examples.pop()
                    logger.warning("Attempt %d: stripped example_sql entry due to proto parse error", attempt + 1)
                    deploy_warnings.append("Stripped example_sql entry (proto error)")
                    if not examples:
                        inst.pop("example_question_sqls", None)
                    continue
                # Phase 3: strip non-prebuilt joins
                def _is_prebuilt(j):
                    l = j.get("left", {}).get("identifier", "").lower()
                    r = j.get("right", {}).get("identifier", "").lower()
                    return tuple(sorted([l, r])) in prebuilt_pairs
                non_prebuilt = [j for j in join_specs if not _is_prebuilt(j)]
                if non_prebuilt:
                    removed = non_prebuilt[-1]
                    join_specs.remove(removed)
                    left_id = removed.get("left", {}).get("identifier", "?")
                    right_id = removed.get("right", {}).get("identifier", "?")
                    logger.warning(
                        "Attempt %d: removing agent join_spec (%s <-> %s) as last resort",
                        attempt + 1, left_id, right_id,
                    )
                    deploy_warnings.append(f"Removed agent join: {left_id} <-> {right_id}")
                    continue
                # Phase 4: strip ALL remaining joins (including prebuilt) as nuclear fallback
                if join_specs:
                    logger.warning("Attempt %d: stripping ALL %d remaining joins as nuclear fallback", attempt + 1, len(join_specs))
                    deploy_warnings.append(f"Stripped all {len(join_specs)} join_specs (proto error)")
                    join_specs.clear()
                    continue
                break
            break
    logger.error("Genie create/update failed: %s", last_err)
    detail = f"Failed to create/update Genie space: {last_err}"
    if deploy_warnings:
        detail += f" (warnings: {'; '.join(deploy_warnings)})"
    raise HTTPException(500, detail=detail)


# ---------------------------------------------------------------------------
# Genie Space updater AI assist
# ---------------------------------------------------------------------------

@app.post("/api/genie/update-assist")
def genie_update_assist(req: GenieUpdateAssistRequest):
    """AI-assisted generation of a single section of a Genie space definition."""
    valid_sections = {"joins", "instructions", "questions", "measures", "filters", "expressions", "example_sql", "synonyms"}
    if req.section not in valid_sections:
        raise HTTPException(400, detail=f"Invalid section '{req.section}'. Must be one of: {', '.join(sorted(valid_sections))}")
    if not req.table_identifiers:
        raise HTTPException(400, detail="table_identifiers required")
    ws = _get_effective_client()
    wh = os.environ.get("WAREHOUSE_ID", "")
    if not wh:
        raise HTTPException(500, detail="WAREHOUSE_ID not configured")
    cat = os.environ.get("CATALOG_NAME", "")
    sch = os.environ.get("SCHEMA_NAME", "")
    from dbxmetagen.genie.context import generate_section_assist
    result = generate_section_assist(
        ws=ws, warehouse_id=wh, catalog=cat, schema=sch,
        section=req.section,
        table_identifiers=req.table_identifiers,
        existing_items=req.existing_items,
        user_prompt=req.user_prompt,
        model_endpoint=req.model_endpoint,
    )
    if "error" in result and len(result) == 1:
        raise HTTPException(500, detail=result["error"])
    return result


@app.post("/api/genie/enrich-description")
def genie_enrich_description(req: GenieEnrichDescriptionRequest):
    """LLM-powered enrichment of a Genie table description using KB metadata."""
    if not req.table_identifier:
        raise HTTPException(400, detail="table_identifier required")
    wh = os.environ.get("WAREHOUSE_ID", "")
    if not wh:
        raise HTTPException(500, detail="WAREHOUSE_ID not configured")

    tbl_safe = _safe_sql_str(req.table_identifier)
    table_rows = execute_sql(
        f"SELECT comment, domain, subdomain FROM {fq('table_knowledge_base')} WHERE table_name = {tbl_safe} LIMIT 1"
    )
    col_rows = execute_sql(
        f"SELECT column_name, data_type, comment FROM {fq('column_knowledge_base')} WHERE table_name = {tbl_safe} ORDER BY column_name"
    )

    if not table_rows and not col_rows:
        raise HTTPException(404, detail=f"No KB data found for {req.table_identifier}")

    kb_context_parts = []
    if table_rows:
        t = table_rows[0]
        kb_context_parts.append(f"Table comment: {t.get('comment', 'N/A')}")
        if t.get("domain"):
            kb_context_parts.append(f"Domain: {t['domain']}")
        if t.get("subdomain"):
            kb_context_parts.append(f"Subdomain: {t['subdomain']}")
    if col_rows:
        col_lines = [f"  - {c['column_name']} ({c.get('data_type', '?')}): {c.get('comment', '')}" for c in col_rows]
        kb_context_parts.append("Columns:\n" + "\n".join(col_lines))

    kb_context = "\n".join(kb_context_parts)
    existing = req.existing_description or "(none)"

    from langchain_community.chat_models import ChatDatabricks
    model = os.environ.get("LLM_MODEL", "databricks-claude-sonnet-4-6")
    llm = ChatDatabricks(endpoint=model, temperature=0.1, max_tokens=2048, max_retries=1, request_timeout=60)
    messages = [
        {"role": "system", "content": (
            "You write concise, Genie-optimized table descriptions for Databricks Genie Spaces. "
            "A good description helps Genie understand what the table contains, its business purpose, "
            "and key columns so it can generate accurate SQL. Keep it under 3 sentences."
        )},
        {"role": "user", "content": (
            f"Enrich this table description using the knowledge base metadata below.\n\n"
            f"Table: {req.table_identifier}\n"
            f"Current description: {existing}\n\n"
            f"=== Knowledge Base Metadata ===\n{kb_context}\n\n"
            f"Write an improved description that incorporates the KB context. "
            f"Output ONLY the description text, no quotes or explanation."
        )},
    ]
    result = llm.invoke(messages)
    content = (getattr(result, "content", "") or "").strip().strip('"').strip("'")
    return {"description": content}


@app.get("/api/genie/uc-comment")
def genie_uc_comment(table_identifier: str):
    """Fetch the live Unity Catalog comment for a table."""
    parts = table_identifier.split(".")
    if len(parts) != 3:
        raise HTTPException(400, detail="table_identifier must be catalog.schema.table")
    cat, sch, tbl = parts
    for p in (cat, sch, tbl):
        _validate_filter(p, "identifier_part")
    rows = execute_sql(
        f"SELECT comment FROM {cat}.information_schema.tables "
        f"WHERE table_schema = {_safe_sql_str(sch)} AND table_name = {_safe_sql_str(tbl)} LIMIT 1"
    )
    comment = rows[0].get("comment") if rows else None
    return {"comment": comment}


@app.get("/api/genie/table-columns")
def genie_table_columns(table_identifier: str):
    """Fetch column names, types, and KB comments for a table."""
    tbl_safe = _safe_sql_str(table_identifier)
    rows = execute_sql(
        f"SELECT column_name, data_type, comment "
        f"FROM {fq('column_knowledge_base')} "
        f"WHERE table_name = {tbl_safe} ORDER BY column_name"
    )
    return {"columns": rows or []}


# ---------------------------------------------------------------------------
# Genie Space tracking endpoints
# ---------------------------------------------------------------------------

def _ensure_genie_tracking_table():
    try:
        execute_sql(f"""
            CREATE TABLE IF NOT EXISTS {fq('genie_spaces')} (
                space_id STRING, title STRING, tables ARRAY<STRING>,
                config_json STRING, version INT,
                status STRING, parent_space_id STRING,
                created_at TIMESTAMP, updated_at TIMESTAMP, deleted_at TIMESTAMP
            )
        """, timeout=30)
        for col_name, typ in [("status", "STRING"), ("parent_space_id", "STRING")]:
            try:
                execute_sql(f"ALTER TABLE {fq('genie_spaces')} ADD COLUMN {col_name} {typ}", timeout=15)
            except Exception:
                pass
    except Exception as e:
        logger.warning("Could not create genie_spaces tracking table: %s", e)


@app.get("/api/genie/spaces")
def list_genie_spaces():
    _ensure_genie_tracking_table()
    return execute_sql(f"""
        SELECT space_id, title, tables, config_json, COALESCE(version, 1) as version, created_at, updated_at
        FROM {fq('genie_spaces')}
        WHERE deleted_at IS NULL AND COALESCE(status, 'active') = 'active'
        ORDER BY updated_at DESC
    """)


@app.post("/api/genie/spaces/track")
def track_genie_space(space_id: str, title: str, tables: list[str], config_json: str = ""):
    _ensure_genie_tracking_table()
    cfg_esc = config_json.replace("'", "''")
    arr_literal = ",".join("'" + t + "'" for t in tables)
    execute_sql(
        f"INSERT INTO {fq('genie_spaces')} "
        f"(space_id, title, tables, config_json, version, status, parent_space_id, created_at, updated_at, deleted_at) VALUES "
        f"('{space_id}', '{title}', ARRAY({arr_literal}), "
        f"'{cfg_esc}', 1, 'active', NULL, current_timestamp(), current_timestamp(), NULL)",
        timeout=30,
    )
    return {"ok": True}


def _parse_serialized_space(raw) -> dict:
    """Robustly parse a serialized_space value that may be a dict, JSON string, double-encoded, or wrapped API response."""
    def _unwrap(d: dict) -> dict:
        if "serialized_space" in d and ("space_id" in d or "title" in d):
            return _parse_serialized_space(d["serialized_space"])
        return d

    if isinstance(raw, dict):
        return _unwrap(raw)
    if not isinstance(raw, str):
        return {}
    try:
        parsed = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return {}
    if isinstance(parsed, str):
        try:
            parsed = json.loads(parsed)
        except (json.JSONDecodeError, TypeError):
            return {}
    if isinstance(parsed, dict):
        return _unwrap(parsed)
    return {}


@app.get("/api/genie/spaces/{space_id}/definition")
def get_genie_space_definition(space_id: str):
    """Return the full serialized_space for a Genie space.

    Checks the tracked genie_spaces table first (has config_json).
    Falls back to fetching live from the Databricks Genie API.
    If tracked config_json is empty/corrupt, fetches live and backfills.
    """
    _ensure_genie_tracking_table()
    tracked_rows = execute_sql(
        f"SELECT title, config_json, COALESCE(version, 1) as version "
        f"FROM {fq('genie_spaces')} "
        f"WHERE space_id = '{space_id}' AND deleted_at IS NULL "
        f"AND COALESCE(status, 'active') = 'active' "
        f"ORDER BY version DESC LIMIT 1",
        timeout=15,
    )
    tracked_row = tracked_rows[0] if tracked_rows else None
    tracked_ss = {}
    if tracked_row and tracked_row.get("config_json"):
        tracked_ss = _parse_serialized_space(tracked_row["config_json"])
        logger.info("get_genie_space_definition: tracked space %s, config_json type=%s, parsed keys=%s",
                     space_id, type(tracked_row["config_json"]).__name__, list(tracked_ss.keys())[:10])

    cj_type = type(tracked_row["config_json"]).__name__ if tracked_row and tracked_row.get("config_json") else None
    cj_len = len(str(tracked_row.get("config_json", ""))) if tracked_row else 0

    if tracked_ss:
        return {
            "space_id": space_id,
            "title": tracked_row.get("title", ""),
            "description": tracked_ss.get("description", ""),
            "serialized_space": tracked_ss,
            "tracked": tracked_row is not None,
            "version": int(tracked_row.get("version", 1)) if tracked_row else 1,
            "_debug": {"source": "tracked", "config_json_type": cj_type, "config_json_len": cj_len, "parsed_keys": list(tracked_ss.keys())[:10]},
        }

    # Tracked config_json was missing/empty/corrupt -- fetch live from Genie API
    try:
        ws = get_workspace_client()
        resp = ws.api_client.do("GET", f"/api/2.0/genie/spaces/{space_id}?include_serialized_space=true")
        ss_raw = resp.get("serialized_space", "{}")
        ss = _parse_serialized_space(ss_raw)
        logger.info("get_genie_space_definition: API space %s, raw type=%s, parsed keys=%s, resp_keys=%s",
                     space_id, type(ss_raw).__name__, list(ss.keys())[:10], list(resp.keys())[:15])

        # Backfill tracked row's config_json if it was empty
        if tracked_row and ss:
            try:
                backfill = json.dumps(ss).replace("'", "''")
                execute_sql(
                    f"UPDATE {fq('genie_spaces')} SET config_json = '{backfill}', updated_at = current_timestamp() "
                    f"WHERE space_id = '{space_id}' AND COALESCE(status, 'active') = 'active' AND deleted_at IS NULL",
                    timeout=30,
                )
                logger.info("Backfilled config_json for tracked space %s", space_id)
            except Exception as bf_err:
                logger.warning("Failed to backfill config_json for %s: %s", space_id, bf_err)

        return {
            "space_id": space_id,
            "title": resp.get("title", resp.get("display_name", tracked_row.get("title", "") if tracked_row else "")),
            "description": resp.get("description", ""),
            "serialized_space": ss,
            "tracked": tracked_row is not None,
            "version": int(tracked_row.get("version", 1)) if tracked_row else 1,
            "_debug": {"source": "live_api", "config_json_type": cj_type, "config_json_len": cj_len, "parsed_keys": list(ss.keys())[:10], "raw_resp_keys": list(resp.keys())[:15], "ss_raw_type": type(ss_raw).__name__, "backfilled": tracked_row is not None and bool(ss)},
        }
    except Exception as e:
        raise HTTPException(404, detail=f"Could not load Genie space {space_id}: {e}")


@app.delete("/api/genie/spaces/{space_id}")
def delete_genie_space(space_id: str):
    _ensure_genie_tracking_table()
    try:
        ws = get_workspace_client()
        ws.api_client.do("DELETE", f"/api/2.0/genie/spaces/{space_id}")
    except Exception as e:
        logger.warning("Could not delete Genie space %s from Databricks: %s", space_id, e)
    execute_sql(
        f"UPDATE {fq('genie_spaces')} SET deleted_at = current_timestamp() "
        f"WHERE space_id = '{space_id}' AND deleted_at IS NULL",
        timeout=30,
    )
    return {"ok": True, "space_id": space_id}


# ---------------------------------------------------------------------------
# KPI Library endpoints
# ---------------------------------------------------------------------------

def _ensure_kpi_table():
    try:
        execute_sql(f"""
            CREATE TABLE IF NOT EXISTS {fq('kpi_definitions')} (
                kpi_id STRING, name STRING, description STRING,
                formula STRING, target_tables ARRAY<STRING>,
                domain STRING, source STRING,
                created_at TIMESTAMP, updated_at TIMESTAMP,
                validation_status STRING, validation_error STRING
            )
        """, timeout=30)
    except Exception as e:
        logger.warning("Could not create kpi_definitions table: %s", e)
    try:
        execute_sql(f"ALTER TABLE {fq('kpi_definitions')} ADD COLUMNS (validation_status STRING, validation_error STRING)", timeout=15)
    except Exception:
        pass
    try:
        execute_sql(f"ALTER TABLE {fq('kpi_definitions')} ADD COLUMNS (profile_id STRING)", timeout=15)
    except Exception:
        pass


def _validate_kpi_formula(formula: str, target_tables: list[str]) -> tuple[str, str]:
    """Dry-run a KPI formula against each target table. Returns (status, error)."""
    if not formula or not target_tables:
        return "skipped", ""
    for table in target_tables:
        try:
            execute_sql(f"SELECT {formula} FROM {table} LIMIT 0", timeout=30)
        except Exception as e:
            return "invalid", f"Against {table}: {e}"
    return "valid", ""


class KpiRequest(BaseModel):
    name: str
    description: str = ""
    formula: str = ""
    target_tables: list[str] = []
    domain: str = ""
    profile_id: Optional[str] = None


class KpiSuggestRequest(BaseModel):
    table_identifiers: list[str]
    count: int = 8
    model_endpoint: str = _LLM_MODEL
    business_context: Optional[str] = None
    questions: list[str] = []
    profile_id: Optional[str] = None


@app.get("/api/kpis")
def list_kpis(profile_id: str = None):
    _ensure_kpi_table()
    where = f" WHERE profile_id = '{profile_id.replace(chr(39), chr(39)*2)}'" if profile_id else ""
    return execute_sql(f"SELECT * FROM {fq('kpi_definitions')}{where} ORDER BY updated_at DESC")


@app.post("/api/kpis")
def create_kpi(req: KpiRequest):
    _ensure_kpi_table()
    kpi_id = str(_uuid.uuid4())[:12]
    name_esc = req.name.replace("'", "''")
    desc_esc = req.description.replace("'", "''")
    formula_esc = req.formula.replace("'", "''")
    arr = ",".join("'" + t + "'" for t in req.target_tables)
    v_status, v_error = _validate_kpi_formula(req.formula, req.target_tables)
    v_error_esc = v_error.replace("'", "''")
    pid = req.profile_id or ""
    execute_sql(
        f"INSERT INTO {fq('kpi_definitions')} "
        f"(kpi_id, name, description, formula, target_tables, domain, source, "
        f"created_at, updated_at, validation_status, validation_error, profile_id) VALUES "
        f"('{kpi_id}', '{name_esc}', '{desc_esc}', '{formula_esc}', "
        f"ARRAY({arr}), '{req.domain}', 'manual', current_timestamp(), current_timestamp(), "
        f"'{v_status}', '{v_error_esc}', '{pid}')",
        timeout=30,
    )
    return {"kpi_id": kpi_id, "name": req.name, "validation_status": v_status, "validation_error": v_error}


@app.put("/api/kpis/{kpi_id}")
def update_kpi(kpi_id: str, req: KpiRequest):
    _ensure_kpi_table()
    name_esc = req.name.replace("'", "''")
    desc_esc = req.description.replace("'", "''")
    formula_esc = req.formula.replace("'", "''")
    arr = ",".join("'" + t + "'" for t in req.target_tables)
    v_status, v_error = _validate_kpi_formula(req.formula, req.target_tables)
    v_error_esc = v_error.replace("'", "''")
    pid = req.profile_id or ""
    execute_sql(
        f"UPDATE {fq('kpi_definitions')} SET name = '{name_esc}', description = '{desc_esc}', "
        f"formula = '{formula_esc}', target_tables = ARRAY({arr}), domain = '{req.domain}', "
        f"validation_status = '{v_status}', validation_error = '{v_error_esc}', "
        f"profile_id = '{pid}', "
        f"updated_at = current_timestamp() WHERE kpi_id = '{kpi_id}'",
        timeout=30,
    )
    return {"ok": True, "validation_status": v_status, "validation_error": v_error}


@app.delete("/api/kpis")
def delete_all_kpis():
    _ensure_kpi_table()
    execute_sql(f"DELETE FROM {fq('kpi_definitions')}", timeout=30)
    return {"ok": True}


@app.delete("/api/kpis/{kpi_id}")
def delete_kpi(kpi_id: str):
    _ensure_kpi_table()
    execute_sql(f"DELETE FROM {fq('kpi_definitions')} WHERE kpi_id = '{kpi_id}'", timeout=30)
    return {"ok": True}


def _build_kpi_context(assembler, table_identifiers: list[str]) -> str:
    """Build condensed entity-first context for KPI generation.

    Leads with entity types and relationships, then per-entity summarized
    columns grouped by role (measures, dimensions, identifiers).  Omits full
    column listings to reduce noise and let the LLM focus on business concepts.
    """
    table_meta = assembler._get_table_metadata(table_identifiers)
    column_meta = assembler._get_column_metadata(table_identifiers)
    fk_rows = assembler._get_fk_predictions(table_identifiers)
    entity_rows = assembler._get_ontology_entities(table_identifiers)
    entity_rels = assembler._get_entity_relationships(table_identifiers)

    col_by_table: dict[str, list] = {}
    for c in column_meta:
        col_by_table.setdefault(c["table_name"], []).append(c)

    entity_map: dict[str, dict] = {}
    for e in entity_rows:
        src = e.get("source_tables") or []
        if isinstance(src, str):
            src = [src]
        for t in src:
            entity_map[t] = e
            entity_map[t.split(".")[-1]] = e

    parts: list[str] = []

    # Entity overview first
    if entity_rows:
        parts.append("ENTITIES (the core business objects in this data):")
        for e in entity_rows:
            desc = e.get("description", "")
            parts.append(f"  {e['entity_type']}: {desc}" if desc else f"  {e['entity_type']}")

    if entity_rels:
        parts.append("\nENTITY RELATIONSHIPS:")
        for r in entity_rels:
            card = r.get("cardinality", "")
            parts.append(f"  {r.get('src_type', '')} --{r.get('relationship', '')}--> {r.get('dst_type', '')}" + (f" ({card})" if card else ""))

    # Per-table: domain + summarized columns by role
    measure_keywords = {"amount", "price", "cost", "revenue", "total", "charge", "fee", "balance", "salary", "quantity", "count", "sum", "rate", "percent", "score", "value"}
    temporal_types = {"DATE", "TIMESTAMP", "DATETIME"}
    id_keywords = {"_id", "id", "key", "code", "number", "num", "no"}

    for t in table_meta:
        tname = t["table_name"]
        ent = entity_map.get(tname) or entity_map.get(tname.split(".")[-1])
        header = f"\n{tname}"
        if t.get("domain"):
            header += f" (Domain: {t['domain']}/{t.get('subdomain', '')})"
        if ent:
            header += f" Entity: {ent['entity_type']}"
        if t.get("comment"):
            header += f" -- {t['comment']}"
        parts.append(header)

        cols = col_by_table.get(tname, [])
        measures, dimensions, identifiers = [], [], []
        for c in cols:
            cn = c["column_name"].lower()
            dt = (c.get("data_type") or "").upper()
            comment = c.get("comment") or ""
            label = c["column_name"]
            if dt:
                label += f" {dt}"
            if comment:
                label += f" -- {comment}"
            if any(kw in cn for kw in id_keywords):
                identifiers.append(label)
            elif dt in temporal_types or "date" in cn or "time" in cn:
                dimensions.append(label)
            elif any(kw in cn for kw in measure_keywords) or dt in ("DECIMAL", "DOUBLE", "FLOAT", "INT", "BIGINT", "SMALLINT"):
                measures.append(label)
            else:
                dimensions.append(label)
        if identifiers:
            parts.append(f"  Identifiers: {'; '.join(identifiers[:8])}")
        if measures:
            parts.append(f"  Measure columns: {'; '.join(measures[:12])}")
        if dimensions:
            parts.append(f"  Dimension columns: {'; '.join(dimensions[:12])}")

    if fk_rows:
        parts.append("\nFOREIGN KEY RELATIONSHIPS:")
        for fk in fk_rows:
            parts.append(f"  {fk['src_table']}.{fk['src_column']} -> {fk['dst_table']}.{fk['dst_column']}")

    return "\n".join(parts)


@app.post("/api/kpis/suggest")
def suggest_kpis(req: KpiSuggestRequest):
    wh = os.environ.get("WAREHOUSE_ID", "")
    if not wh:
        raise HTTPException(500, detail="WAREHOUSE_ID not configured")
    from dbxmetagen.genie.context import GenieContextAssembler
    from langchain_community.chat_models import ChatDatabricks

    ws = _get_effective_client()
    assembler = GenieContextAssembler(ws, wh, CATALOG, SCHEMA)
    kpi_context = _build_kpi_context(assembler, req.table_identifiers)

    biz_ctx_block = ""
    if req.business_context and req.business_context.strip():
        biz_ctx_block = f"\nBUSINESS CONTEXT (provided by the user -- this defines the semantic frame for all analysis):\n{req.business_context.strip()}\n"

    questions_block = ""
    if req.questions:
        q_list = "\n".join(f"  - {q}" for q in req.questions)
        questions_block = f"\nBUSINESS QUESTIONS (the KPIs you generate should help answer these):\n{q_list}\n"

    prompt = f"""You are a business intelligence architect. Given the data model below, suggest {req.count} concrete KPIs that would bridge the semantic gap between raw data and business meaning.
{biz_ctx_block}
{kpi_context}
{questions_block}
Rules:
- Ground each KPI in the ENTITY TYPES described in the metadata -- if there are Patients, Encounters, Claims, Orders, etc., the KPIs should measure aspects of those specific entities
- Use the RELATIONSHIPS between entities to suggest cross-entity KPIs (e.g., encounters per patient, revenue per provider, claims per policy)
- Each KPI's formula MUST reference only columns that exist in the provided table metadata -- do not invent columns
- Align KPIs with the domain/subdomain classifications shown in the metadata
- If BUSINESS QUESTIONS are provided, prioritize KPIs that directly support answering those questions
- Think like a business user who works WITH the data but doesn't know the schema -- KPIs should be framed in business language

For each KPI provide:
- name: concise business name (e.g. "Monthly Revenue Growth Rate")
- description: 1-2 sentences explaining what it measures and why it matters
- formula: SQL expression using bare column names only -- NO catalog, schema, or table prefixes (e.g. SUM(total_amount), not SUM(schema.table.total_amount)). The source table is specified separately.
- domain: business domain it belongs to (e.g. sales, finance, operations)

Return ONLY a JSON array of objects with keys: name, description, formula, domain. No other text."""

    llm = ChatDatabricks(endpoint=req.model_endpoint, temperature=0.5, max_tokens=4096)
    response = llm.invoke(prompt)
    content = response.content.strip()
    if content.startswith("```"):
        content = content.split("\n", 1)[1] if "\n" in content else content[3:]
        content = content.rsplit("```", 1)[0]
    kpis = json.loads(content)
    for kpi in kpis:
        tables = req.table_identifiers[:1] if req.table_identifiers else []
        v_status, v_error = _validate_kpi_formula(kpi.get("formula", ""), tables)
        kpi["validation_status"] = v_status
        kpi["validation_error"] = v_error
    return {"kpis": kpis[:req.count]}


# ---------------------------------------------------------------------------
# Metadata Intelligence Agent endpoints
# ---------------------------------------------------------------------------

VS_ENDPOINT = os.environ.get("VECTOR_SEARCH_ENDPOINT", "dbxmetagen-vs")
VS_INDEX_SUFFIX = os.environ.get("VECTOR_SEARCH_INDEX", "metadata_vs_index")

_api_vsc = None
_api_vs_indexes: dict = {}


def _get_api_vs_index(index_name: str):
    """Return a cached VectorSearchIndex for the API layer."""
    global _api_vsc
    if index_name in _api_vs_indexes:
        return _api_vs_indexes[index_name]
    if _api_vsc is None:
        from databricks.vector_search.client import VectorSearchClient
        ws = get_workspace_client()
        client_id = os.environ.get("DATABRICKS_CLIENT_ID")
        client_secret = os.environ.get("DATABRICKS_CLIENT_SECRET")
        if client_id and client_secret:
            _api_vsc = VectorSearchClient(
                workspace_url=ws.config.host,
                service_principal_client_id=client_id,
                service_principal_client_secret=client_secret,
            )
        else:
            _token = os.environ.get("DATABRICKS_TOKEN")
            if not _token:
                headers = ws.config.authenticate()
                _token = headers.get("Authorization", "").removeprefix("Bearer ")
            _api_vsc = VectorSearchClient(workspace_url=ws.config.host, personal_access_token=_token)
    idx = _api_vsc.get_index(endpoint_name=VS_ENDPOINT, index_name=index_name)
    _api_vs_indexes[index_name] = idx
    return idx


class AgentChatRequest(BaseModel):
    message: str
    history: list = []
    mode: str = "quick"
    session_id: str = ""


VALID_AGENT_MODES = {"quick", "deep", "graphrag", "baseline"}


@app.post("/api/agent/chat")
async def agent_chat(req: AgentChatRequest):
    t0 = time.time()
    from agent.guardrails import validate_input
    ok, err = validate_input(req.message)
    if not ok:
        raise HTTPException(400, detail=err)
    try:
        from agent.metadata_agent import run_metadata_agent
    except ImportError as e:
        raise HTTPException(503, detail=f"Agent not available: {e}")
    mode = req.mode if req.mode in VALID_AGENT_MODES else "quick"
    try:
        result = await run_metadata_agent(req.message, history=req.history, mode=mode, session_id=req.session_id or None)
        if isinstance(result, dict):
            result["elapsed_ms"] = int((time.time() - t0) * 1000)
        return result
    except Exception as exc:
        msg = str(exc)
        if "REQUEST_LIMIT_EXCEEDED" in msg or "429" in msg or "RateLimitError" in msg:
            raise HTTPException(429, detail="Model rate limit exceeded. Try again shortly.") from exc
        logger.error("Metadata agent error: %s", exc)
        raise HTTPException(500, detail=f"Agent error: {msg}") from exc


# ---------------------------------------------------------------------------
# Plot generation from agent responses
# ---------------------------------------------------------------------------

class PlotRequest(BaseModel):
    content: str
    history: list = []


@app.post("/api/agent/plot")
def agent_plot(req: PlotRequest):
    """Generate a chart specification from an agent response."""
    if not req.content:
        return {"no_data": True, "reason": "No content provided"}
    try:
        from agent.metadata_agent import create_plot_spec
        spec = create_plot_spec(req.content, req.history)
        return spec
    except Exception as e:
        logger.error("Plot agent error: %s", e, exc_info=True)
        return {"no_data": True, "reason": str(e)}


# ---------------------------------------------------------------------------
# Task-based deep analysis (background task + polling, avoids HTTP timeout)
# ---------------------------------------------------------------------------

_deep_tasks: dict[str, dict] = {}


@app.post("/api/agent/deep/submit")
def agent_deep_submit(req: AgentChatRequest):
    """Submit a deep analysis (graphrag/baseline) as a background task.

    Returns {"task_id": "..."} immediately. Poll GET /api/agent/deep/task/{task_id}
    for progress and results.
    """
    from agent.guardrails import validate_input
    ok, err = validate_input(req.message)
    if not ok:
        raise HTTPException(400, detail=err)
    mode = req.mode if req.mode in ("graphrag", "baseline") else "graphrag"
    try:
        from agent.deep_analysis import run_deep_analysis_streaming
    except ImportError as e:
        raise HTTPException(503, detail=f"Deep analysis agent not available: {e}")

    task_id = str(_uuid.uuid4())[:12]
    _deep_tasks[task_id] = {"status": "running", "stage": "starting", "message": "", "steps": [], "elapsed_ms": 0, "created": time.time()}

    progress_q, cancel_event = run_deep_analysis_streaming(req.message, mode=mode, history=req.history, session_id=req.session_id or None)

    _DEEP_WALL_TIMEOUT = 300  # 5-minute absolute max (new pipeline typically finishes in ~90s)

    def _monitor():
        wall_deadline = time.time() + _DEEP_WALL_TIMEOUT
        try:
            while True:
                if time.time() > wall_deadline:
                    cancel_event.set()
                    elapsed_s = int(time.time() - _deep_tasks[task_id]["created"])
                    _deep_tasks[task_id].update({
                        "status": "error",
                        "error": f"Analysis timed out after {elapsed_s}s. Try a simpler question.",
                        "elapsed_ms": elapsed_s * 1000,
                    })
                    return
                remaining = max(wall_deadline - time.time(), 1)
                try:
                    event = progress_q.get(timeout=min(remaining, 30))
                except queue.Empty:
                    continue
                if event.get("stage") == "done":
                    created = _deep_tasks[task_id]["created"]
                    prev_steps = _deep_tasks[task_id].get("steps", [])
                    _deep_tasks[task_id] = {
                        "status": "done",
                        "stage": "done",
                        "answer": event.get("answer", event.get("response", "")),
                        "tool_calls": event.get("tool_calls", []),
                        "mode": event.get("mode", mode),
                        "routing_trace": event.get("routing_trace"),
                        "graph_data": event.get("graph_data"),
                        "timing": event.get("timing"),
                        "intent": event.get("intent"),
                        "steps": prev_steps,
                        "created": created,
                        "elapsed_ms": int((time.time() - created) * 1000),
                    }
                    return
                if event.get("stage") == "error":
                    _deep_tasks[task_id] = {
                        **_deep_tasks[task_id],
                        "status": "error",
                        "error": event.get("message", "Unknown error"),
                    }
                    return
                stage = event.get("stage", "running")
                msg = event.get("message", "")
                _deep_tasks[task_id]["stage"] = stage
                _deep_tasks[task_id]["message"] = msg
                _deep_tasks[task_id].setdefault("steps", []).append({
                    "stage": stage, "message": msg, "ts": time.time(),
                })
                _deep_tasks[task_id]["elapsed_ms"] = int((time.time() - _deep_tasks[task_id]["created"]) * 1000)
        except Exception as e:
            logger.error("Deep task monitor error: %s", e, exc_info=True)
            _deep_tasks[task_id] = {
                **_deep_tasks[task_id],
                "status": "error",
                "error": str(e),
            }

    threading.Thread(target=_monitor, daemon=True).start()

    cutoff = time.time() - 600
    for tid in list(_deep_tasks):
        if _deep_tasks.get(tid, {}).get("created", 0) < cutoff:
            _deep_tasks.pop(tid, None)

    return {"task_id": task_id}


@app.get("/api/agent/deep/task/{task_id}")
def agent_deep_poll(task_id: str):
    """Poll a deep analysis task for status/progress/result."""
    cutoff = time.time() - 600
    for tid in list(_deep_tasks):
        if tid != task_id and _deep_tasks.get(tid, {}).get("created", 0) < cutoff:
            _deep_tasks.pop(tid, None)
    task = _deep_tasks.get(task_id)
    if not task:
        raise HTTPException(404, detail="Task not found")
    return task


# ---------------------------------------------------------------------------
# SSE streaming deep analysis (LangGraph-based, replaces submit/poll for new UI)
# ---------------------------------------------------------------------------

_DEEP_STREAM_TIMEOUT = 300  # 5-minute wall-clock max for SSE stream


@app.post("/api/agent/deep/stream")
async def agent_deep_stream(req: AgentChatRequest, request: Request):
    """SSE endpoint: streams stage progress, token-level output, and final result.

    Event types:
      event: stage     -- {"stage": "...", "message": "..."}
      event: progress  -- {"stage": "gathering", "message": "Step 2/7: ..."}
      event: token     -- {"content": "partial text"}
      event: done      -- {"answer": "...", "tool_calls": [...], "graph_data": {...}, ...}
      event: error     -- {"message": "..."}
    """
    from agent.guardrails import validate_input
    ok, err = validate_input(req.message)
    if not ok:
        raise HTTPException(400, detail=err)

    mode = req.mode if req.mode in ("graphrag", "baseline") else "graphrag"

    try:
        from agent.deep_analysis_graph import get_graph, NODE_STAGE_MAP
    except ImportError as e:
        raise HTTPException(503, detail=f"Deep analysis graph not available: {e}")

    def _sse(event_type: str, data: dict) -> str:
        return f"event: {event_type}\ndata: {json.dumps(data, default=str)}\n\n"

    async def event_generator():
        # Fix #1: ensure MLflow context is set in the async worker thread
        try:
            from agent.tracing import ensure_mlflow_context
            ensure_mlflow_context()
        except Exception:
            pass

        graph = get_graph()
        initial_state = {
            "query": req.message,
            "history": req.history or [],
            "session_id": req.session_id or "",
            "mode": mode,
        }

        t_start = time.time()
        deadline = t_start + _DEEP_STREAM_TIMEOUT
        answer_tokens: list[str] = []
        root_run_id: str | None = None

        try:
            async for event in graph.astream_events(initial_state, version="v2"):
                # Fix #3: wall-clock timeout check
                if time.time() > deadline:
                    elapsed_s = int(time.time() - t_start)
                    logger.error("SSE deep stream timed out after %ds", elapsed_s)
                    yield _sse("error", {"message": f"Analysis timed out after {elapsed_s}s. Try a simpler question."})
                    return

                # Fix #5: abort if client disconnected
                if await request.is_disconnected():
                    logger.info("SSE client disconnected, aborting stream")
                    return

                kind = event["event"]
                name = event.get("name", "")

                # Fix #2: capture root run_id from the first LangGraph chain start
                if kind == "on_chain_start" and name == "LangGraph" and root_run_id is None:
                    root_run_id = event.get("run_id")

                if kind == "on_chain_start" and name in NODE_STAGE_MAP:
                    yield _sse("stage", NODE_STAGE_MAP[name])

                elif kind == "on_custom_event" and name == "progress":
                    yield _sse("progress", event["data"])

                elif kind == "on_chat_model_stream":
                    chunk = event.get("data", {}).get("chunk")
                    if chunk:
                        content = getattr(chunk, "content", "") or ""
                        if content:
                            answer_tokens.append(content)
                            yield _sse("token", {"content": content})

                elif kind == "on_chain_end" and name == "LangGraph":
                    output = event.get("data", {}).get("output", {})
                    final_answer = "".join(answer_tokens) if answer_tokens else output.get("answer", "")
                    elapsed_ms = int((time.time() - t_start) * 1000)

                    timing = output.get("timing") or {}
                    yield _sse("done", {
                        "answer": final_answer,
                        "tool_calls": output.get("tool_calls", []),
                        "graph_data": output.get("graph_data"),
                        "timing": timing,
                        "token_usage": timing.get("token_usage"),
                        "mode": output.get("mode", mode),
                        "intent": output.get("intent_type"),
                        "trace_id": root_run_id,
                        "elapsed_ms": elapsed_ms,
                    })

        except Exception as e:
            logger.error("SSE deep stream error: %s", e, exc_info=True)
            yield _sse("error", {"message": str(e)})

    return StreamingResponse(event_generator(), media_type="text/event-stream")


@app.get("/api/agent/stats")
def agent_stats():
    """Summary statistics for the agent landing page."""
    stats = {"lakebase_connected": pg_configured()}
    try:
        rows = execute_sql(f"SELECT COUNT(*) AS cnt FROM {fq('table_knowledge_base')}")
        stats["tables_profiled"] = rows[0]["cnt"] if rows else 0
    except Exception:
        stats["tables_profiled"] = 0
    try:
        rows = execute_sql(f"SELECT COUNT(DISTINCT entity_type) AS cnt FROM {fq('ontology_entities')}")
        stats["entity_types"] = rows[0]["cnt"] if rows else 0
    except Exception:
        stats["entity_types"] = 0
    try:
        rows = execute_sql(f"SELECT COUNT(*) AS cnt FROM {fq('fk_predictions')} WHERE final_confidence >= 0.5")
        stats["fk_predictions"] = rows[0]["cnt"] if rows else 0
    except Exception:
        stats["fk_predictions"] = 0
    try:
        rows = execute_sql(f"SELECT COUNT(*) AS cnt FROM {fq('metric_view_definitions')} WHERE status IN ('validated','applied')")
        stats["metric_views"] = rows[0]["cnt"] if rows else 0
    except Exception:
        stats["metric_views"] = 0
    try:
        rows = execute_sql(f"SELECT doc_type, COUNT(*) AS cnt FROM {fq('metadata_documents')} GROUP BY doc_type")
        stats["vs_documents"] = sum(int(r["cnt"]) for r in rows) if rows else 0
        stats["vs_by_type"] = {r["doc_type"]: int(r["cnt"]) for r in rows} if rows else {}
    except Exception as e:
        logger.warning("metadata_documents query failed in agent_stats: %s", e)
        stats["vs_documents"] = 0
        stats["vs_by_type"] = {}
    if stats["vs_documents"] == 0:
        try:
            ws = get_workspace_client()
            vs_index_name = f"{CATALOG}.{SCHEMA}.{VS_INDEX_SUFFIX}"
            idx = ws.vector_search_indexes.get_index(vs_index_name)
            if idx.status and hasattr(idx.status, "indexed_row_count"):
                count = idx.status.indexed_row_count or 0
                if count > 0:
                    stats["vs_documents"] = count
                    stats["vs_source"] = "index"
        except Exception:
            pass
    return stats


@app.get("/api/agent/suggestions")
def agent_suggestions():
    """Context-aware suggestion chips."""
    suggestions = [
        {"label": "What tables exist in my catalog?", "query": "What tables exist in my catalog?"},
        {"label": "Show me the data quality summary", "query": "Show me the data quality summary for all tables"},
    ]
    try:
        ents = execute_sql(f"SELECT COUNT(*) AS cnt FROM {fq('ontology_entities')}")
        if ents and ents[0]["cnt"] and int(ents[0]["cnt"]) > 0:
            suggestions.append({"label": "What entity types were discovered?", "query": "What entity types were discovered and which tables do they map to?"})
    except Exception:
        pass
    try:
        fks = execute_sql(f"SELECT COUNT(*) AS cnt FROM {fq('fk_predictions')} WHERE final_confidence >= 0.5")
        if fks and fks[0]["cnt"] and int(fks[0]["cnt"]) > 0:
            suggestions.append({"label": "How are my tables related?", "query": "Show me the foreign key relationships between tables"})
    except Exception:
        pass
    try:
        mvs = execute_sql(f"SELECT COUNT(*) AS cnt FROM {fq('metric_view_definitions')} WHERE status IN ('validated','applied')")
        if mvs and mvs[0]["cnt"] and int(mvs[0]["cnt"]) > 0:
            suggestions.append({"label": "What metric views are available?", "query": "List all available metric views with their source tables and measures"})
    except Exception:
        pass
    suggestions.append({"label": "Which columns contain PII or PHI?", "query": "Which columns contain PII or PHI data?"})
    return suggestions


@app.get("/api/agent/domain-stats")
def agent_domain_stats():
    """Domain-level breakdowns for the agent stats panel."""
    result = {}
    try:
        rows = execute_sql(f"SELECT domain, COUNT(*) AS cnt FROM {fq('table_knowledge_base')} GROUP BY domain ORDER BY cnt DESC LIMIT 20")
        result["tables_by_domain"] = [{"domain": r.get("domain", "unknown"), "count": int(r["cnt"])} for r in rows] if rows else []
    except Exception:
        result["tables_by_domain"] = []
    try:
        rows = execute_sql(f"SELECT entity_type, COUNT(*) AS cnt FROM {fq('ontology_entities')} GROUP BY entity_type ORDER BY cnt DESC LIMIT 20")
        result["entities_by_type"] = [{"type": r.get("entity_type", "unknown"), "count": int(r["cnt"])} for r in rows] if rows else []
    except Exception:
        result["entities_by_type"] = []
    try:
        rows = execute_sql(f"""
            SELECT t.domain, COUNT(*) AS cnt
            FROM {fq('fk_predictions')} f
            JOIN {fq('table_knowledge_base')} t ON f.src_table = t.table_name
            WHERE f.final_confidence >= 0.5
            GROUP BY t.domain ORDER BY cnt DESC LIMIT 15
        """)
        result["fk_by_domain"] = [{"domain": r.get("domain", "unknown"), "count": int(r["cnt"])} for r in rows] if rows else []
    except Exception:
        result["fk_by_domain"] = []
    return result


# ---------------------------------------------------------------------------
# Vector Search endpoints
# ---------------------------------------------------------------------------


class VectorSearchRequest(BaseModel):
    query: str
    doc_type: Optional[str] = None
    num_results: int = 5
    query_type: str = "ANN"


@app.get("/api/vector/status")
def vector_status():
    """Get VS endpoint and index status + document counts."""
    vs_index_name = f"{CATALOG}.{SCHEMA}.{VS_INDEX_SUFFIX}"
    result: dict = {"endpoint_name": VS_ENDPOINT, "index_name": vs_index_name}
    try:
        ws = get_workspace_client()
        ep = ws.vector_search_endpoints.get_endpoint(VS_ENDPOINT)
        result["endpoint_state"] = ep.endpoint_status.state.value if ep.endpoint_status else "UNKNOWN"
    except Exception as e:
        err = str(e)
        logger.warning("VS endpoint check failed for '%s': %s", VS_ENDPOINT, err)
        if "RESOURCE_DOES_NOT_EXIST" in err or "does not exist" in err.lower() or "not found" in err.lower():
            result["endpoint_state"] = "NOT_FOUND"
        else:
            result["endpoint_state"] = "ERROR"
        result["endpoint_error"] = err
    try:
        ws = get_workspace_client()
        idx = ws.vector_search_indexes.get_index(vs_index_name)
        result["index_status"] = str(idx.status) if idx.status else "UNKNOWN"
    except Exception as e:
        err = str(e)
        if "RESOURCE_DOES_NOT_EXIST" in err or "does not exist" in err.lower() or "not found" in err.lower():
            result["index_status"] = "NOT_FOUND"
        else:
            result["index_status"] = "ERROR"
        result["index_error"] = err
    try:
        rows = execute_sql(f"SELECT doc_type, COUNT(*) AS cnt FROM {fq('metadata_documents')} GROUP BY doc_type ORDER BY cnt DESC")
        result["doc_counts"] = {r["doc_type"]: int(r["cnt"]) for r in rows} if rows else {}
        result["total_documents"] = sum(result["doc_counts"].values())
    except Exception:
        result["doc_counts"] = {}
        result["total_documents"] = 0
    return result


@app.post("/api/vector/search")
def vector_search(req: VectorSearchRequest):
    """Execute a similarity search against the metadata VS index."""
    vs_index_name = f"{CATALOG}.{SCHEMA}.{VS_INDEX_SUFFIX}"
    try:
        index = _get_api_vs_index(vs_index_name)
        kwargs = dict(
            query_text=req.query,
            columns=["doc_id", "doc_type", "content", "table_name", "domain", "entity_type", "confidence_score"],
            num_results=min(max(req.num_results, 1), 20),
        )
        if req.doc_type:
            kwargs["filters"] = {"doc_type": req.doc_type}
        if req.query_type == "HYBRID":
            kwargs["query_type"] = "HYBRID"
        results = index.similarity_search(**kwargs)
        matches = []
        cols = results.get("manifest", {}).get("columns", [])
        col_names = [c.get("name", f"col{i}") for i, c in enumerate(cols)] if cols else []
        for row in results.get("result", {}).get("data_array", []):
            if col_names:
                matches.append(dict(zip(col_names, row)))
            else:
                matches.append({"data": row})
        return {"matches": matches, "count": len(matches), "query_type": req.query_type}
    except Exception as e:
        raise HTTPException(500, detail=f"Vector search failed: {e}")


@app.post("/api/vector/sync")
def vector_sync():
    """Trigger a sync of the metadata VS index."""
    vs_index_name = f"{CATALOG}.{SCHEMA}.{VS_INDEX_SUFFIX}"
    try:
        ws = get_workspace_client()
        ws.vector_search_indexes.sync_index(index_name=vs_index_name)
        return {"status": "sync_triggered", "index": vs_index_name}
    except Exception as e:
        raise HTTPException(500, detail=f"Sync failed: {e}")


# ---------------------------------------------------------------------------
# SQL Analyst Agent (Blind vs Enriched)
# ---------------------------------------------------------------------------

_analyst_tasks: dict[str, dict] = {}


@app.post("/api/analyst/chat")
def analyst_chat(req: dict):
    """Run the analyst agent in a single mode (blind or enriched)."""
    question = req.get("question", "")
    mode = req.get("mode", "enriched")
    history = req.get("history", [])
    if mode not in ("blind", "enriched"):
        raise HTTPException(400, detail="mode must be 'blind' or 'enriched'")
    from agent.guardrails import validate_input
    ok, err = validate_input(question)
    if not ok:
        raise HTTPException(400, detail=err)
    from agent.analyst_agent import run_analyst_single
    task_id = str(_uuid.uuid4())[:12]
    _analyst_tasks[task_id] = {"status": "running", "stage": f"{mode}_running", "created": time.time()}

    def _run():
        try:
            result = run_analyst_single(question, mode, history)
            _analyst_tasks[task_id] = {"status": "done", "result": result, "created": _analyst_tasks[task_id]["created"]}
        except Exception as exc:
            logger.exception("Analyst single-mode failed")
            _analyst_tasks[task_id] = {"status": "error", "error": str(exc), "created": _analyst_tasks[task_id]["created"]}

    threading.Thread(target=_run, daemon=True).start()
    return {"task_id": task_id}


@app.post("/api/analyst/compare")
def analyst_compare(req: dict):
    """Run both blind and enriched analysts in parallel for side-by-side comparison."""
    question = req.get("question", "")
    from agent.guardrails import validate_input
    ok, err = validate_input(question)
    if not ok:
        raise HTTPException(400, detail=err)
    from agent.analyst_agent import run_analyst_compare
    import queue as _queue
    task_id = str(_uuid.uuid4())[:12]
    _analyst_tasks[task_id] = {"status": "running", "stage": "starting", "created": time.time()}
    progress_q = _queue.Queue()

    def _run():
        try:
            result = run_analyst_compare(question, progress_q)
            _analyst_tasks[task_id] = {"status": "done", "result": result, "created": _analyst_tasks[task_id]["created"]}
        except Exception as exc:
            logger.exception("Analyst compare failed")
            _analyst_tasks[task_id] = {"status": "error", "error": str(exc), "created": _analyst_tasks[task_id]["created"]}

    def _monitor():
        while True:
            try:
                event = progress_q.get(timeout=300)
                _analyst_tasks[task_id]["stage"] = event.get("stage", _analyst_tasks[task_id].get("stage"))
            except Exception:
                break
            if _analyst_tasks[task_id]["status"] in ("done", "error"):
                break

    threading.Thread(target=_run, daemon=True).start()
    threading.Thread(target=_monitor, daemon=True).start()
    return {"task_id": task_id}


@app.get("/api/analyst/task/{task_id}")
def analyst_task(task_id: str):
    """Poll analyst task status."""
    task = _analyst_tasks.get(task_id)
    if not task:
        raise HTTPException(404, detail="Task not found")
    if task["created"] < time.time() - 600:
        _analyst_tasks.pop(task_id, None)
        raise HTTPException(410, detail="Task expired")
    return task


@app.post("/api/analyst/plot")
def analyst_plot(req: dict):
    """Generate a chart specification from an analyst response."""
    content = req.get("content", "")
    sql = req.get("sql")
    history = req.get("history")
    if not content:
        raise HTTPException(400, detail="No content provided")
    try:
        from agent.analyst_agent import create_analyst_plot_spec
        spec = create_analyst_plot_spec(content, sql, history)
        return spec
    except Exception as e:
        logger.exception("Analyst plot error")
        return {"no_data": True, "reason": str(e)}


@app.post("/api/analyst/stream")
def analyst_stream(req: dict):
    """SSE streaming endpoint for analyst agent (single or compare mode)."""
    question = req.get("question", "")
    mode = req.get("mode", "compare")
    history = req.get("history", [])
    from agent.guardrails import validate_input
    ok, err = validate_input(question)
    if not ok:
        raise HTTPException(400, detail=err)

    def _sse(event: str, data: dict) -> str:
        return f"data: {json.dumps({'event': event, **data})}\n\n"

    def _single_generator(m: str):
        from agent.analyst_agent import run_analyst_single
        yield _sse("stage", {"stage": f"{m}_running"})
        try:
            result = run_analyst_single(question, m, history)
            yield _sse("done", {"result": result})
        except Exception as exc:
            logger.exception("Analyst stream single-mode failed")
            yield _sse("error", {"error": str(exc)})

    def _compare_generator():
        from agent.analyst_agent import run_analyst_single, generate_comparison_analysis
        yield _sse("stage", {"stage": "starting"})
        results = {"blind": None, "enriched": None}
        errors = {"blind": None, "enriched": None}
        done_q = queue.Queue()

        def _run_mode(m):
            try:
                results[m] = run_analyst_single(question, m)
            except Exception as exc:
                logger.exception("Analyst stream %s failed", m)
                errors[m] = str(exc)
            done_q.put(m)

        threading.Thread(target=_run_mode, args=("enriched",), daemon=True).start()
        yield _sse("stage", {"stage": "enriched_running"})

        time.sleep(4)
        threading.Thread(target=_run_mode, args=("blind",), daemon=True).start()
        yield _sse("stage", {"stage": "blind_running"})

        for _ in range(2):
            finished = done_q.get(timeout=600)
            if results[finished]:
                yield _sse("partial", {"mode": finished, "result": results[finished]})
            elif errors[finished]:
                yield _sse("partial", {"mode": finished, "error": errors[finished]})

        blind_res = results["blind"] or {"error": errors["blind"] or "Timeout"}
        enriched_res = results["enriched"] or {"error": errors["enriched"] or "Timeout"}

        comparison = None
        if not blind_res.get("error") and not enriched_res.get("error"):
            yield _sse("stage", {"stage": "comparing"})
            try:
                comparison = generate_comparison_analysis(question, blind_res, enriched_res)
            except Exception as exc:
                logger.warning("Comparison analysis failed: %s", exc)

        yield _sse("done", {
            "result": {"blind": blind_res, "enriched": enriched_res, "comparison_analysis": comparison},
        })

    gen = _compare_generator() if mode == "compare" else _single_generator(mode)
    return StreamingResponse(gen, media_type="text/event-stream")


# ---------------------------------------------------------------------------
# Governance & Compliance Explorer
# ---------------------------------------------------------------------------

@app.get("/api/governance/summary")
def governance_summary():
    """Per-schema sensitivity counts: PII/PHI/PCI columns, unclassified."""
    try:
        rows = execute_sql(f"""
            SELECT c.schema, c.classification_type,
                   COUNT(*) AS column_count,
                   COUNT(DISTINCT c.table_name) AS table_count
            FROM {CATALOG}.{SCHEMA}.column_knowledge_base c
            WHERE c.classification_type IS NOT NULL
            GROUP BY c.schema, c.classification_type
            ORDER BY c.schema, c.classification_type
        """)
        return {"summary": rows}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@app.get("/api/governance/gaps")
def governance_gaps():
    """Columns where profiling patterns suggest PII but classification is missing."""
    try:
        rows = execute_sql(f"""
            SELECT cs.table_name, cs.column_name, cs.pattern_detected,
                   cs.distinct_count, cs.null_rate,
                   ck.classification, ck.classification_type
            FROM {CATALOG}.{SCHEMA}.column_profiling_stats cs
            INNER JOIN (
                SELECT snapshot_id, table_name FROM (
                    SELECT snapshot_id, table_name,
                           ROW_NUMBER() OVER (PARTITION BY table_name ORDER BY snapshot_time DESC) rn
                    FROM {CATALOG}.{SCHEMA}.profiling_snapshots
                ) WHERE rn = 1
            ) latest ON cs.snapshot_id = latest.snapshot_id AND cs.table_name = latest.table_name
            LEFT JOIN {CATALOG}.{SCHEMA}.column_knowledge_base ck
              ON cs.table_name = ck.table_name AND cs.column_name = ck.column_name
            WHERE cs.pattern_detected IN ('email', 'phone', 'ssn', 'uuid', 'ip_address', 'credit_card')
              AND (ck.classification IS NULL OR ck.classification = 'none' OR ck.classification = '')
            ORDER BY cs.pattern_detected, cs.table_name
            LIMIT 200
        """)
        return {"gaps": rows}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@app.get("/api/governance/masking")
def governance_masking():
    """Classified columns that lack column mask policies."""
    try:
        rows = execute_sql(f"""
            SELECT ck.table_name, ck.column_name, ck.classification, ck.classification_type,
                   em.column_mask_policies
            FROM {CATALOG}.{SCHEMA}.column_knowledge_base ck
            LEFT JOIN {CATALOG}.{SCHEMA}.extended_table_metadata em
              ON ck.table_name = em.table_name
            WHERE ck.classification_type IN ('pii', 'phi', 'pci')
            ORDER BY ck.classification_type DESC, ck.table_name
            LIMIT 200
        """)
        return {"masking_audit": rows}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@app.get("/api/governance/lineage")
def governance_lineage(table: str = ""):
    """Sensitive data lineage: upstream/downstream of tables with classified columns."""
    where = f"AND ck.table_name LIKE '%{table.split('.')[-1]}%'" if table else ""
    try:
        rows = execute_sql(f"""
            SELECT DISTINCT ck.table_name, ck.classification_type,
                   em.upstream_tables, em.downstream_tables
            FROM {CATALOG}.{SCHEMA}.column_knowledge_base ck
            INNER JOIN {CATALOG}.{SCHEMA}.extended_table_metadata em
              ON ck.table_name = em.table_name
            WHERE ck.classification_type IN ('pii', 'phi', 'pci') {where}
            LIMIT 100
        """)
        return {"lineage": rows}
    except Exception as e:
        raise HTTPException(500, detail=str(e))


@app.post("/api/governance/chat")
async def governance_chat(req: dict):
    """Conversational governance agent."""
    question = req.get("question", "")
    history = req.get("history", [])
    from agent.guardrails import validate_input
    ok, err = validate_input(question)
    if not ok:
        raise HTTPException(400, detail=err)
    from agent.governance_agent import run_governance_agent
    result = await run_governance_agent(question, history)
    return result


# ---------------------------------------------------------------------------
# Impact Analysis Agent
# ---------------------------------------------------------------------------

_impact_tasks: dict[str, dict] = {}


@app.post("/api/impact/analyze")
def impact_analyze(req: dict):
    """Submit an impact analysis request."""
    question = req.get("question", "")
    from agent.guardrails import validate_input
    ok, err = validate_input(question)
    if not ok:
        raise HTTPException(400, detail=err)
    from agent.impact_agent import run_impact_analysis
    task_id = str(_uuid.uuid4())[:12]
    _impact_tasks[task_id] = {"status": "running", "stage": "starting", "created": time.time()}

    def _run():
        try:
            result = run_impact_analysis(question)
            _impact_tasks[task_id] = {"status": "done", "result": result, "created": _impact_tasks[task_id]["created"]}
        except Exception as exc:
            logger.exception("Impact analysis failed")
            _impact_tasks[task_id] = {"status": "error", "error": str(exc), "created": _impact_tasks[task_id]["created"]}

    threading.Thread(target=_run, daemon=True).start()
    return {"task_id": task_id}


@app.get("/api/impact/task/{task_id}")
def impact_task(task_id: str):
    """Poll impact analysis task."""
    task = _impact_tasks.get(task_id)
    if not task:
        raise HTTPException(404, detail="Task not found")
    if task["created"] < time.time() - 600:
        _impact_tasks.pop(task_id, None)
        raise HTTPException(410, detail="Task expired")
    return task


@app.post("/api/impact/chat")
async def impact_chat(req: dict):
    """Conversational follow-up for impact analysis."""
    question = req.get("question", "")
    history = req.get("history", [])
    from agent.guardrails import validate_input
    ok, err = validate_input(question)
    if not ok:
        raise HTTPException(400, detail=err)
    from agent.impact_agent import run_impact_chat
    result = await run_impact_chat(question, history)
    return result


# ---------------------------------------------------------------------------
# Serve React static files (production build)
# ---------------------------------------------------------------------------

static_dir = os.path.join(os.path.dirname(__file__), "src", "dist")
if os.path.isdir(static_dir):
    app.mount("/", StaticFiles(directory=static_dir, html=True), name="static")
